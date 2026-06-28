#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

ROUND4_PRIORITY120 = EXPORTS / "issue19-round4-city-gradient-priority120-groups.csv"
ROUND4_MAJOR_DETAILS = EXPORTS / "issue19-round4-city-gradient-major-details.csv"
GROUP_READINESS = WORKING / "issue19-stable-foundation-group-readiness-bridge.csv"
CLOSURE_V1_PRIORITY55 = EXPORTS / "issue19-closure-and-shortlist-v1-priority55-groups.csv"

OUT_PREFIX = "issue19-round4-priority-focus55"
OUT_SUMMARY = EXPORTS / f"{OUT_PREFIX}-summary.json"
OUT_WORKBOOK = EXPORTS / f"{OUT_PREFIX}.xlsx"
OUT_GROUPS = EXPORTS / f"{OUT_PREFIX}-groups.csv"
OUT_MAJORS = EXPORTS / f"{OUT_PREFIX}-major-details.csv"
OUT_PAUSED = EXPORTS / f"{OUT_PREFIX}-paused65-groups.csv"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
DATA_STAGE = "issue19_round4_priority_focus55"
GENERATED_AT = "2026-06-29"

PREFERRED_CITIES = {"武汉市": 10, "西安市": 8, "成都市": 8, "北京市": 8}
SECONDARY_CITIES = {"天津市": 3, "重庆市": 3, "郑州市": 2, "南京市": 2, "杭州市": 2, "长沙市": 2}
GRADIENT_QUOTA = {
    "保底观察": 1,
    "稳妥观察": 24,
    "稳冲观察": 18,
    "冲刺观察": 12,
}
GRADIENT_ORDER = ["保底观察", "稳妥观察", "稳冲观察", "冲刺观察"]

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "可否进入最终志愿方案",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

GROUP_FIELDS = [
    "Round4重点核验组ID",
    "来源Round4优先120组",
    "来源Round4完整组内专业",
    "来源稳定基座专业组就绪桥接表",
    "来源ClosureV1重点55组",
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "任务粒度",
    *FALSE_FIELDS,
    "压缩序号",
    "压缩建议层级",
    "是否进入40到60重点核验",
    "是否建议进入下一轮重点核验",
    "压缩综合分",
    "梯度配额标签",
    "梯度内排序",
    "城市偏好层级",
    "历史稳定性等级",
    "专业匹配等级",
    "核验成本等级",
    "调剂接受风险等级",
    "结构字段阻断等级",
    "费用限制风险",
    "压缩主理由",
    "压缩副理由",
    "家庭最小决策动作",
    "核验最小动作",
    "第四轮候选ID",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
    "稳定基座专业组就绪桥接ID",
    "城市",
    "城市来源",
    "省份",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组号",
    "历史线索分层",
    "冲稳保",
    "冲稳保解释",
    "与515关系",
    "2026考生分数",
    "2026考生位次",
    "2025投档分",
    "2025等位分差",
    "2024投档分",
    "2024等位分差",
    "2023投档分",
    "2023等位分差",
    "同代码命中年份数",
    "同代码命中年份",
    "三年投档分范围",
    "等位分差摘要",
    "历史线使用口径",
    "再选要求跨年变化",
    "历史组名疑似不一致",
    "公办民办机器线索",
    "家庭底线属性动作",
    "专业明细行数",
    "优先了解专业数",
    "可调剂了解专业数",
    "待核后判断专业数",
    "不能接受提示专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "计算机AI软件专业数",
    "电子信息网络专业数",
    "机械自动化机器人专业数",
    "师范类相关专业数",
    "环境工程科学专业数",
    "农业不含动物相关专业数",
    "工商旅游管理专业数",
    "高收费或超预算专业数",
    "护理助产专业数",
    "动物医学兽医专业数",
    "临床口腔中医暂缓专业数",
    "医技护理康复专业数",
    "特殊限制待核专业数",
    "调剂初判",
    "完整专业组接受度初判",
    "组内专业摘录",
    "办学地点线索",
    "来源页码",
    "版面列",
    "保真核验层级",
    "候选讨论就绪层级",
    "主要阻断原因",
    "字段缺口专业行数",
    "结构或归属未闭环专业行数",
    "P0人工核验专业行数",
    "P1人工核验专业行数",
    "必须100%人工核验专业行数",
    "湖北官方系统待核专业行数",
    "PDF原页待核专业行数",
    "官网辅证命中专业行数",
    "自动官网辅证任务数",
    "C0冲突任务数",
    "C1官网补缺任务数",
    "C4部分来源待结构化任务数",
    "C6继续补源任务数",
    "C7官网未匹配任务数",
    "第一闭环覆盖状态",
    "第一闭环明细任务数",
    "第一闭环双人复核任务数",
    "高校侧刷新任务数",
    "机器核验下一步",
    "人工核验下一步",
    "下一步核验动作",
    "公开安全策略",
]

MAJOR_FIELDS = [
    "Round4重点核验专业ID",
    "来源Round4重点核验组",
    "来源Round4完整组内专业",
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    *FALSE_FIELDS,
    "压缩序号",
    "第四轮候选ID",
    "专业组出现ID",
    "城市",
    "冲稳保",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组内专业序号",
    "专业代号",
    "专业名称及备注",
    "专业偏好方向",
    "第四轮方向标签",
    "专业角色判断",
    "机器专业接受度初判",
    "家庭接受度初判V2",
    "调剂风险初判V2",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "来源页码",
    "版面列",
    "专业行ID",
]

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/home/",
    "/var/folders/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    "file://",
    "Authorization",
    "Bearer ",
    "Cookie",
    "Set-Cookie",
    "access_token",
    "refresh_token",
    "password",
    "secret",
    "api_key",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "人工读数",
    "已确认",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
    "可排序",
]


def clean(value):
    text = "" if value is None else str(value)
    return text.replace("\r", " ").replace("\n", " ").strip()


def as_int(value):
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def as_float(value):
    try:
        return float(clean(value) or "0")
    except ValueError:
        return 0.0


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in fields} for row in rows])


def source_path(path):
    return str(path.relative_to(ROOT))


def stable_id(prefix, parts):
    text = "|".join(clean(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(text.encode('utf-8')).hexdigest()[:16]}"


def city_bonus(city):
    city = clean(city)
    if city in PREFERRED_CITIES:
        return PREFERRED_CITIES[city]
    return SECONDARY_CITIES.get(city, 0)


def city_level(city):
    city = clean(city)
    if city in PREFERRED_CITIES:
        return f"C0-初始优先城市：{city}"
    if city in SECONDARY_CITIES:
        return f"C1-次级可讨论城市：{city}"
    if city in {"武汉市", "孝感市", "黄冈市", "黄石市", "荆州市", "荆门市", "襄阳市", "十堰市"}:
        return f"C2-湖北省内或近距离可讨论：{city}"
    return "C3-非当前优先城市，作为梯度和专业匹配讨论"


def direction_score(row):
    fields = [
        ("数字媒体技术专业数", 8),
        ("计算机AI软件专业数", 5),
        ("计算机类相关专业数", 4),
        ("电子信息网络专业数", 3),
        ("师范类相关专业数", 3),
        ("机械自动化机器人专业数", 2),
        ("环境工程科学专业数", 1),
        ("农业不含动物相关专业数", 1),
    ]
    return sum(as_int(row.get(field)) * weight for field, weight in fields)


def direction_summary(row):
    parts = []
    mapping = [
        ("数字媒体技术", "数字媒体技术专业数"),
        ("计算机/软件/AI", "计算机AI软件专业数"),
        ("计算机类", "计算机类相关专业数"),
        ("电子信息/网络", "电子信息网络专业数"),
        ("师范", "师范类相关专业数"),
        ("机械自动化", "机械自动化机器人专业数"),
        ("环境", "环境工程科学专业数"),
        ("农业不含动物", "农业不含动物相关专业数"),
        ("工商旅游管理", "工商旅游管理专业数"),
    ]
    for label, field in mapping:
        n = as_int(row.get(field))
        if n:
            parts.append(f"{label}{n}")
    return "；".join(parts) or "无明显当前偏好专业，主要作为梯度/调剂讨论"


def major_acceptance(row):
    machine = clean(row.get("机器专业接受度初判"))
    risk = clean(row.get("专业风险类型"))
    direction = clean(row.get("专业偏好方向"))
    if "默认不能接受" in machine or "医学/护理" in risk:
        return "不能接受-当前硬排除或医学护理相关"
    if "特殊限制" in machine or risk:
        return "待核后判断-有体检/语种/单科/专项等限制"
    if direction:
        return "优先了解-命中当前专业方向"
    return "可调剂了解-未命中偏好但暂未见硬排除"


def transfer_risk_for_major(row, group_row):
    acceptance = major_acceptance(row)
    if "不能接受" in acceptance:
        return "同组若存在此类专业，服从调剂需谨慎"
    if as_int(group_row.get("特殊限制待核专业数")):
        return "组内有特殊限制待核，服从调剂前需核完整专业组"
    if as_int(group_row.get("专业明细行数")) > 12:
        return "组内专业较多，需逐条确认调剂接受度"
    return "暂未见硬排除项，但仍需原页/章程闭环"


def risk_penalty(row):
    penalty = 0
    penalty += as_int(row.get("_hard_major_count")) * 25
    penalty += min(as_int(row.get("特殊限制待核专业数")) * 2, 10)
    penalty += min(as_int(row.get("专业明细行数")) / 8, 4)
    if clean(row.get("历史组名疑似不一致")) not in {"", "false", "无"}:
        penalty += 4
    if clean(row.get("再选要求跨年变化")) not in {"", "false", "无"}:
        penalty += 4
    return penalty


def compression_score(row):
    return as_float(row.get("讨论排序分")) + city_bonus(row.get("城市")) + direction_score(row) - risk_penalty(row)


def professional_match_level(row):
    if as_int(row.get("数字媒体技术专业数")):
        return "P0-命中数字媒体技术"
    if as_int(row.get("计算机AI软件专业数")) or as_int(row.get("计算机类相关专业数")):
        return "P1-命中计算机/软件/AI"
    if as_int(row.get("电子信息网络专业数")) or as_int(row.get("师范类相关专业数")):
        return "P2-命中电子信息/网络或师范"
    if direction_score(row):
        return "P3-命中备选方向"
    return "P4-专业匹配弱"


def history_level(row):
    history = clean(row.get("历史线索分层"))
    if history.startswith("H1"):
        return "H1-保底历史线索"
    if history.startswith("H2"):
        return "H2-稳妥历史线索"
    if history.startswith("H3"):
        return "H3-稳冲历史线索"
    if history.startswith("H4"):
        return "H4-冲刺历史线索"
    return "H9-历史线索待补"


def verification_cost_level(row):
    p0 = as_int(row.get("P0人工核验专业行数"))
    field_gap = as_int(row.get("字段缺口专业行数"))
    structural = as_int(row.get("结构或归属未闭环专业行数"))
    first_closure = as_int(row.get("第一闭环明细任务数"))
    if structural or p0 >= 8 or first_closure:
        return "V0-高核验成本，先核页和组边界"
    if p0 or field_gap >= 8:
        return "V1-中高核验成本，需集中核字段"
    if field_gap:
        return "V2-中等核验成本，保留字段缺口"
    return "V3-相对低核验成本，仍未定稿"


def transfer_risk_level(row):
    hard = as_int(row.get("_hard_major_count"))
    if hard:
        return f"T0-有{hard}条不能接受提示，服从调剂高风险"
    if as_int(row.get("特殊限制待核专业数")):
        return "T1-特殊限制待核，先核章程/原页"
    if "结构异常" in clean(row.get("调剂初判")):
        return "T2-结构异常，先核完整组内专业"
    if as_int(row.get("专业明细行数")) > 12:
        return "T3-组内专业较多，需逐条确认"
    return "T4-暂未见硬阻断，仍需人工接受度判断"


def structure_level(row):
    structural = as_int(row.get("结构或归属未闭环专业行数"))
    field_gap = as_int(row.get("字段缺口专业行数"))
    if structural:
        return f"S0-结构或归属未闭环{structural}行"
    if field_gap:
        return f"S1-字段事实缺口{field_gap}行"
    return "S2-结构线索相对完整但仍未官方闭环"


def fee_risk(row):
    if as_int(row.get("高收费或超预算专业数")):
        return "F0-存在高收费或超预算线索"
    return "F1-暂未见高收费或超预算线索"


def main_reason(row):
    return "；".join(
        part
        for part in [
            clean(row.get("冲稳保")),
            history_level(row),
            professional_match_level(row),
            city_level(row.get("城市")),
        ]
        if part
    )


def sub_reason(row):
    parts = [direction_summary(row), transfer_risk_level(row), verification_cost_level(row)]
    if as_int(row.get("特殊限制待核专业数")):
        parts.append(f"特殊限制待核{as_int(row.get('特殊限制待核专业数'))}个")
    if as_int(row.get("C0冲突任务数")) or as_int(row.get("C1官网补缺任务数")):
        parts.append(
            f"高校辅证冲突/补缺任务{as_int(row.get('C0冲突任务数')) + as_int(row.get('C1官网补缺任务数'))}条"
        )
    if as_int(row.get("第一闭环明细任务数")):
        parts.append(f"已进入第一闭环任务{as_int(row.get('第一闭环明细任务数'))}条")
    return "；".join(parts)


def family_action(row, selected):
    if not selected:
        return "暂不投入人工核验时间；家庭点名喜欢时再拉回。"
    return "先看完整组内专业，逐条标可接受/勉强接受/不能接受；保留后再做PDF原页、湖北官方侧和章程核验。"


def verification_action(row, selected):
    if not selected:
        return "暂缓；若后续回捞，先核完整专业组和调剂风险。"
    if as_int(row.get("第一闭环明细任务数")):
        return "先处理已排入第一闭环的页列和字段，再核湖北官方侧。"
    if as_int(row.get("结构或归属未闭环专业行数")):
        return "先核PDF原页中的专业组边界、专业归属和完整组内专业。"
    if as_int(row.get("特殊限制待核专业数")):
        return "先核章程、体检、语种、单科、校区和调剂限制。"
    return "进入下一轮候选后，按完整组内专业做100%原页和湖北官方侧核验。"


def select_groups(priority_rows):
    by_gradient = defaultdict(list)
    for row in priority_rows:
        row["_compression_score"] = compression_score(row)
        by_gradient[clean(row.get("冲稳保"))].append(row)

    selected = []
    selected_by_gradient = Counter()
    rank_by_gradient = {}
    for gradient, quota in GRADIENT_QUOTA.items():
        candidates = [
            row for row in by_gradient.get(gradient, [])
            if as_int(row.get("_hard_major_count")) == 0
        ]
        if len(candidates) < quota:
            candidates = by_gradient.get(gradient, [])
        candidates = sorted(
            candidates,
            key=lambda r: (-r["_compression_score"], clean(r.get("院校代码")), clean(r.get("院校专业组代码"))),
        )
        for rank, row in enumerate(candidates, 1):
            rank_by_gradient[clean(row.get("第四轮候选ID"))] = rank
        for row in candidates[:quota]:
            selected.append(row)
            selected_by_gradient[gradient] += 1

    selected.sort(
        key=lambda r: (
            GRADIENT_ORDER.index(clean(r.get("冲稳保"))) if clean(r.get("冲稳保")) in GRADIENT_ORDER else 9,
            -r["_compression_score"],
            clean(r.get("院校代码")),
        )
    )
    selected_ids = {clean(row.get("第四轮候选ID")) for row in selected}
    paused = [row for row in priority_rows if clean(row.get("第四轮候选ID")) not in selected_ids]
    paused.sort(key=lambda r: (-r["_compression_score"], clean(r.get("冲稳保")), clean(r.get("院校代码"))))
    return selected, paused, selected_by_gradient, rank_by_gradient


def build_base_rows(priority_rows, major_rows, readiness_by_group_id):
    hard_major_count = Counter()
    accept_count = Counter()
    fallback_count = Counter()
    pending_count = Counter()
    for row in major_rows:
        cid = clean(row.get("第四轮候选ID"))
        acceptance = major_acceptance(row)
        if acceptance.startswith("不能接受"):
            hard_major_count[cid] += 1
        elif acceptance.startswith("优先了解"):
            accept_count[cid] += 1
        elif acceptance.startswith("可调剂了解"):
            fallback_count[cid] += 1
        else:
            pending_count[cid] += 1

    out = []
    for row in priority_rows:
        bridge = readiness_by_group_id.get(clean(row.get("专业组出现ID")), {})
        merged = {**row}
        for field in [
            "稳定基座专业组就绪桥接ID",
            "候选讨论就绪层级",
            "主要阻断原因",
            "字段缺口专业行数",
            "结构或归属未闭环专业行数",
            "P0人工核验专业行数",
            "P1人工核验专业行数",
            "必须100%人工核验专业行数",
            "湖北官方系统待核专业行数",
            "PDF原页待核专业行数",
            "官网辅证命中专业行数",
            "自动官网辅证任务数",
            "C0冲突任务数",
            "C1官网补缺任务数",
            "C4部分来源待结构化任务数",
            "C6继续补源任务数",
            "C7官网未匹配任务数",
            "第一闭环覆盖状态",
            "第一闭环明细任务数",
            "第一闭环双人复核任务数",
            "高校侧刷新任务数",
        ]:
            merged[field] = bridge.get(field, "")
        cid = clean(row.get("第四轮候选ID"))
        merged["_hard_major_count"] = str(hard_major_count[cid])
        merged["_accept_count"] = str(accept_count[cid])
        merged["_fallback_count"] = str(fallback_count[cid])
        merged["_pending_count"] = str(pending_count[cid])
        out.append(merged)
    return out


def group_row(row, index, selected, rank_by_gradient):
    cid = clean(row.get("第四轮候选ID"))
    layer = "G0-建议重点核验55组" if selected else "G2-暂缓保留65组"
    return {
        "Round4重点核验组ID": stable_id("R4FOCUSGROUP", [cid, "selected" if selected else "paused"]),
        "来源Round4优先120组": source_path(ROUND4_PRIORITY120),
        "来源Round4完整组内专业": source_path(ROUND4_MAJOR_DETAILS),
        "来源稳定基座专业组就绪桥接表": source_path(GROUP_READINESS),
        "来源ClosureV1重点55组": source_path(CLOSURE_V1_PRIORITY55),
        "来源期号": SOURCE_ISSUE,
        "来源PDF_SHA256": SOURCE_PDF_SHA256,
        "生成日期": GENERATED_AT,
        "数据阶段": DATA_STAGE,
        "主表粒度": "Round4优先120组压缩到55组重点核验",
        "任务粒度": "院校专业组×公开压缩理由；完整组内专业另表展开",
        **{field: "false" for field in FALSE_FIELDS},
        "压缩序号": str(index) if selected else "",
        "压缩建议层级": layer,
        "是否进入40到60重点核验": "true" if selected else "false",
        "是否建议进入下一轮重点核验": "true" if selected else "false",
        "压缩综合分": f"{as_float(row.get('_compression_score')):.2f}",
        "梯度配额标签": f"{clean(row.get('冲稳保'))}配额{GRADIENT_QUOTA.get(clean(row.get('冲稳保')), 0)}",
        "梯度内排序": str(rank_by_gradient.get(cid, "")),
        "城市偏好层级": city_level(row.get("城市")),
        "历史稳定性等级": history_level(row),
        "专业匹配等级": professional_match_level(row),
        "核验成本等级": verification_cost_level(row),
        "调剂接受风险等级": transfer_risk_level(row),
        "结构字段阻断等级": structure_level(row),
        "费用限制风险": fee_risk(row),
        "压缩主理由": main_reason(row),
        "压缩副理由": sub_reason(row),
        "家庭最小决策动作": family_action(row, selected),
        "核验最小动作": verification_action(row, selected),
        "第四轮候选ID": cid,
        "专业组出现ID": clean(row.get("专业组出现ID")),
        "稳定基座专业组筛选ID": clean(row.get("稳定基座专业组筛选ID")),
        "稳定基座专业组就绪桥接ID": clean(row.get("稳定基座专业组就绪桥接ID")),
        "优先了解专业数": row.get("_accept_count", "0"),
        "可调剂了解专业数": row.get("_fallback_count", "0"),
        "待核后判断专业数": row.get("_pending_count", "0"),
        "不能接受提示专业数": row.get("_hard_major_count", "0"),
        "公开安全策略": (
            "公开层只保存组级压缩理由、计数、状态和非最终门禁；"
            "不保存人工记录、私有路径、登录态或定稿排序。"
        ),
        **{
            field: clean(row.get(field))
            for field in GROUP_FIELDS
            if field
            not in {
                "Round4重点核验组ID",
                "来源Round4优先120组",
                "来源Round4完整组内专业",
                "来源稳定基座专业组就绪桥接表",
                "来源ClosureV1重点55组",
                "来源期号",
                "来源PDF_SHA256",
                "生成日期",
                "数据阶段",
                "主表粒度",
                "任务粒度",
                *FALSE_FIELDS,
                "压缩序号",
                "压缩建议层级",
                "是否进入40到60重点核验",
                "是否建议进入下一轮重点核验",
                "压缩综合分",
                "梯度配额标签",
                "梯度内排序",
                "城市偏好层级",
                "历史稳定性等级",
                "专业匹配等级",
                "核验成本等级",
                "调剂接受风险等级",
                "结构字段阻断等级",
                "费用限制风险",
                "压缩主理由",
                "压缩副理由",
                "家庭最小决策动作",
                "核验最小动作",
                "第四轮候选ID",
                "专业组出现ID",
                "稳定基座专业组筛选ID",
                "稳定基座专业组就绪桥接ID",
                "优先了解专业数",
                "可调剂了解专业数",
                "待核后判断专业数",
                "不能接受提示专业数",
                "公开安全策略",
            }
        },
    }


def major_row(row, group, selected_order):
    cid = clean(row.get("第四轮候选ID"))
    return {
        "Round4重点核验专业ID": stable_id("R4FOCUSMAJOR", [cid, row.get("专业行ID")]),
        "来源Round4重点核验组": source_path(OUT_GROUPS),
        "来源Round4完整组内专业": source_path(ROUND4_MAJOR_DETAILS),
        "来源期号": SOURCE_ISSUE,
        "来源PDF_SHA256": SOURCE_PDF_SHA256,
        "生成日期": GENERATED_AT,
        "数据阶段": DATA_STAGE,
        "主表粒度": "Round4重点核验组×完整组内专业",
        **{field: "false" for field in FALSE_FIELDS},
        "压缩序号": str(selected_order[cid]),
        "家庭接受度初判V2": major_acceptance(row),
        "调剂风险初判V2": transfer_risk_for_major(row, group),
        **{
            field: clean(row.get(field))
            for field in MAJOR_FIELDS
            if field
            not in {
                "Round4重点核验专业ID",
                "来源Round4重点核验组",
                "来源Round4完整组内专业",
                "来源期号",
                "来源PDF_SHA256",
                "生成日期",
                "数据阶段",
                "主表粒度",
                *FALSE_FIELDS,
                "压缩序号",
                "家庭接受度初判V2",
                "调剂风险初判V2",
            }
        },
    }


def append_sheet(wb, title, rows, fields, tab_color=None):
    ws = wb.create_sheet(title)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(fields)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([clean(row.get(field)) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, 1):
        values = [len(clean(field))] + [len(clean(row.get(field))) for row in rows[:200]]
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(10, min(max(values) + 2, 42))
    return ws


def write_workbook(summary_rows, group_rows, major_rows, paused_rows):
    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(wb, "00_摘要", summary_rows, ["项目", "当前值", "说明"], "305496")
    append_sheet(wb, "01_重点核验55组", group_rows, GROUP_FIELDS, "70AD47")
    append_sheet(wb, "02_重点组完整专业", major_rows, MAJOR_FIELDS, "5B9BD5")
    append_sheet(wb, "03_暂缓65组", paused_rows, GROUP_FIELDS, "A6A6A6")
    wb.save(OUT_WORKBOOK)


def validate_public_safety(rows, summary):
    text = json.dumps(summary, ensure_ascii=False) + "\n"
    for rowset, fields in rows:
        text += "\n".join(",".join(clean(row.get(field)) for field in fields) for row in rowset)
    hits = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]
    if hits:
        raise ValueError(f"public output contains forbidden tokens: {hits}")


def build_summary(group_rows, major_rows, paused_rows, selected_by_gradient, existing_selected_ids):
    selected_ids = {row["第四轮候选ID"] for row in group_rows}
    summary_rows = [
        {"项目": "Round4优先讨论组", "当前值": "120", "说明": "从第四轮城市冲稳保候选池进入压缩的组数"},
        {"项目": "建议重点核验组", "当前值": str(len(group_rows)), "说明": "本轮固定压缩为55组，位于40-60目标区间内"},
        {"项目": "暂缓组", "当前值": str(len(paused_rows)), "说明": "仍保留在Round4优先120组备选，不投入第一批人工核验"},
        {"项目": "重点组完整专业", "当前值": str(len(major_rows)), "说明": "55组内完整组内专业展开，用于调剂接受度判断"},
        {"项目": "字段事实门禁", "当前值": "0", "说明": "字段写回、推荐依据和最终可用均为0"},
    ]
    summary = {
        "status": f"{DATA_STAGE}_not_final",
        "generated_by": Path(__file__).name,
        "source_round4_priority120": source_path(ROUND4_PRIORITY120),
        "source_round4_major_details": source_path(ROUND4_MAJOR_DETAILS),
        "source_group_readiness_bridge": source_path(GROUP_READINESS),
        "source_closure_v1_priority55": source_path(CLOSURE_V1_PRIORITY55),
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "generated_at": GENERATED_AT,
        "output_group_table": source_path(OUT_GROUPS),
        "output_major_table": source_path(OUT_MAJORS),
        "output_paused_table": source_path(OUT_PAUSED),
        "output_workbook": source_path(OUT_WORKBOOK),
        "target_focus_range": "40-60",
        "selected_group_count": len(group_rows),
        "paused_group_count": len(paused_rows),
        "source_priority120_group_count": len(group_rows) + len(paused_rows),
        "selected_major_detail_count": len(major_rows),
        "unique_selected_school_count": len({row["院校代码"] for row in group_rows}),
        "selected_gradient_counts": dict(sorted(Counter(row["冲稳保"] for row in group_rows).items())),
        "paused_gradient_counts": dict(sorted(Counter(row["冲稳保"] for row in paused_rows).items())),
        "selected_history_counts": dict(sorted(Counter(row["历史线索分层"] for row in group_rows).items())),
        "selected_city_top20": dict(Counter(row["城市"] for row in group_rows).most_common(20)),
        "selected_professional_match_counts": dict(sorted(Counter(row["专业匹配等级"] for row in group_rows).items())),
        "selected_verification_cost_counts": dict(sorted(Counter(row["核验成本等级"] for row in group_rows).items())),
        "selected_transfer_risk_counts": dict(sorted(Counter(row["调剂接受风险等级"] for row in group_rows).items())),
        "selected_structure_block_counts": dict(sorted(Counter(row["结构字段阻断等级"] for row in group_rows).items())),
        "selected_major_acceptance_counts": dict(sorted(Counter(row["家庭接受度初判V2"] for row in major_rows).items())),
        "selected_by_gradient_quota": dict(sorted(selected_by_gradient.items())),
        "selected_overlap_with_closure_v1_priority55": len(selected_ids & existing_selected_ids),
        "field_writeback_ready_count": 0,
        "final_available_count": 0,
        "next_stage_available_count": 0,
        "recommendation_basis_allowed_count": 0,
        "school_major_suggestion_allowed_count": 0,
        "official_plan_replacement_allowed_count": 0,
        "summary_rows": summary_rows,
        "public_boundary": (
            "该表只把Round4优先120组压缩为55个重点核验入口；"
            "所有组仍需第19期原页、湖北官方侧、高校章程和完整调剂范围核验后才可进入定稿表。"
        ),
    }
    return summary, summary_rows


def main():
    priority_rows = read_csv(ROUND4_PRIORITY120)
    major_rows = read_csv(ROUND4_MAJOR_DETAILS)
    readiness_rows = read_csv(GROUP_READINESS)
    existing_rows = read_csv(CLOSURE_V1_PRIORITY55) if CLOSURE_V1_PRIORITY55.exists() else []
    readiness_by_group_id = {clean(row.get("专业组出现ID")): row for row in readiness_rows}
    existing_selected_ids = {clean(row.get("第四轮候选ID")) for row in existing_rows}

    base_rows = build_base_rows(priority_rows, major_rows, readiness_by_group_id)
    selected, paused, selected_by_gradient, rank_by_gradient = select_groups(base_rows)
    selected_order = {clean(row.get("第四轮候选ID")): idx for idx, row in enumerate(selected, 1)}
    selected_by_id = {clean(row.get("第四轮候选ID")): row for row in selected}

    group_rows = [group_row(row, idx, True, rank_by_gradient) for idx, row in enumerate(selected, 1)]
    paused_rows = [group_row(row, "", False, rank_by_gradient) for row in paused]

    major_output_rows = []
    for row in major_rows:
        cid = clean(row.get("第四轮候选ID"))
        if cid in selected_by_id:
            major_output_rows.append(major_row(row, selected_by_id[cid], selected_order))
    major_output_rows.sort(key=lambda r: (as_int(r["压缩序号"]), as_int(r["专业组内专业序号"])))

    summary, summary_rows = build_summary(group_rows, major_output_rows, paused_rows, selected_by_gradient, existing_selected_ids)
    validate_public_safety(
        [(group_rows, GROUP_FIELDS), (major_output_rows, MAJOR_FIELDS), (paused_rows, GROUP_FIELDS)],
        summary,
    )

    write_csv(OUT_GROUPS, group_rows, GROUP_FIELDS)
    write_csv(OUT_MAJORS, major_output_rows, MAJOR_FIELDS)
    write_csv(OUT_PAUSED, paused_rows, GROUP_FIELDS)
    OUT_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(summary_rows, group_rows, major_output_rows, paused_rows)


if __name__ == "__main__":
    main()
