#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

PERSONAL_GROUPS = EXPORTS / "issue19-personal-fit-v1-groups.csv"
ROUND2_MAIN_SHORTLIST_GROUPS = EXPORTS / "issue19-round2-updated-preferences-main-shortlist-groups.csv"
ROUND2_MAIN_SHORTLIST_MAJORS = EXPORTS / "issue19-round2-updated-preferences-main-shortlist-majors.csv"
ROUND1_SHORTLIST_GROUPS = EXPORTS / "issue19-round1-shortlist-groups.csv"
STABLE_MAJOR_BROWSER = EXPORTS / "issue19-stable-foundation-major-browser.csv"

MAIN_TABLE = EXPORTS / "issue19-volunteer-table-v1-draft-main.csv"
MAJOR_CHOICES = EXPORTS / "issue19-volunteer-table-v1-draft-major-choices.csv"
DISCUSSION_BATCHES = EXPORTS / "issue19-volunteer-table-v1-discussion-batches.csv"
SPECIAL_OPTIONS = EXPORTS / "issue19-volunteer-table-v1-special-options.csv"
SUMMARY_JSON = EXPORTS / "issue19-volunteer-table-v1-draft-summary.json"
WORKBOOK = EXPORTS / "issue19-volunteer-table-v1-draft.xlsx"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

MAIN_FIELDS = [
    "志愿序号",
    "草案梯度",
    "讨论批次",
    "二次核验优先级",
    "是否进入主草表",
    "是否可作为定稿依据",
    "推荐进入原因",
    "核心风险",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "专业明细行数",
    "建议专业1",
    "建议专业2",
    "建议专业3",
    "建议专业4",
    "建议专业5",
    "建议专业6",
    "组内待确认专业摘录",
    "服从调剂草案",
    "二次核验动作",
    "第19期页码",
    "版面列",
    "专业组出现ID",
    "来源候选层",
]

MAJOR_FIELDS = [
    "志愿序号",
    "草案梯度",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "是否建议填入6专业",
    "专业选择顺位",
    "专业角色判断",
    "专业偏好方向",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "第19期页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
]

DISCUSSION_FIELDS = [
    "讨论批次",
    "批次目标",
    "包含志愿序号",
    "院校专业组数量",
    "优先讨论问题",
    "完成条件",
]

SPECIAL_FIELDS = [
    "专项序号",
    "专项层级",
    "家庭讨论动作",
    "是否进入专项讨论",
    "是否可作为定稿依据",
    "适配理由",
    "关键风险",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "城市候选",
    "历史线索分层",
    "历史最高等位分差",
    "场景判断",
    "场景内最低学费",
    "场景内最高学费",
    "预算内中外合作或高收费专业数",
    "超过预算专业数",
    "待核费用专业数",
    "第19期页码",
    "版面列",
    "组内完整专业清单索引",
    "专业组出现ID",
    "下一步核验动作",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def int_value(row, field, default=0):
    try:
        return int(str(row.get(field, "")).strip() or str(default))
    except ValueError:
        return default


def float_value(row, field, default=999.0):
    try:
        text = str(row.get(field, "")).strip()
        return float(text) if text else default
    except ValueError:
        return default


def code(row):
    return row.get("院校专业组代码OCR规范化", "")


def group_key(row):
    return row.get("专业组出现ID", "") or code(row)


def stable_id(prefix, key):
    return f"{prefix}-{hashlib.sha1(key.encode('utf-8')).hexdigest()[:16]}"


def history_code(row):
    return row.get("历史线索分层", "")[:2]


def is_personal_normal(row):
    return (
        row.get("候选主线") == "普通公办低费用主线"
        and row.get("是否进入本轮家庭讨论") == "true"
    )


def discussion_entered(row):
    return row.get("是否进入本轮家庭讨论") == "true"


def public_normal(row):
    return row.get("家庭底线属性动作") == "继续核公办普通学费-非民办线索"


def hard_main_ok(row):
    family_match = row.get("机器家庭匹配初判", "")
    return (
        public_normal(row)
        and int_value(row, "医学护理排除专业数") == 0
        and int_value(row, "临床口腔中医暂缓专业数") == 0
        and int_value(row, "护理助产专业数") == 0
        and int_value(row, "医技护理康复专业数") == 0
        and int_value(row, "动物医学兽医专业数") == 0
        and int_value(row, "高收费或超预算专业数") == 0
        and int_value(row, "专业明细行数") > 0
        and "默认不进主方案" not in family_match
    )


def preference_count(row):
    return int_value(row, "偏好专业数")


def risk_text(row):
    pieces = [
        row.get("风险提示", ""),
        row.get("关键风险", ""),
        row.get("专项风险提示", ""),
        row.get("体检和章程核验提示", ""),
        row.get("调剂初判", ""),
        row.get("机器家庭匹配初判", ""),
    ]
    return "；".join(piece for piece in pieces if piece)


def bucket(row):
    h = history_code(row)
    diff = float_value(row, "历史最高等位分差")
    if h == "H4" or 15 < diff <= 30:
        return "冲刺"
    if h == "H0":
        return "冲刺待补历史"
    if h == "H3" or 0 < diff <= 15:
        return "稳冲"
    if h == "H1" or diff <= -10:
        return "保底"
    if h == "H2" or diff <= 0:
        return "稳妥"
    return "备选待核"


def group_score(row):
    city_bonus = 12 if row.get("城市候选") in {"武汉", "成都", "西安", "北京"} else 0
    personal_bonus = 20 if row.get("来自个人适配") == "true" else 0
    digital_bonus = int_value(row, "数字媒体技术专业数") * 9
    comp_bonus = int_value(row, "计算机类相关专业数") * 4
    ai_bonus = int_value(row, "计算机AI软件专业数") * 5
    electronic_bonus = int_value(row, "电子信息网络专业数") * 4
    mech_bonus = int_value(row, "机械自动化机器人专业数") * 3
    teacher_bonus = int_value(row, "师范类相关专业数") * 2
    wider_bonus = int_value(row, "环境工程科学专业数") + int_value(row, "工商旅游管理专业数")
    risk_penalty = 0
    text = risk_text(row)
    if "结构" in text:
        risk_penalty += 8
    if "特殊限制" in text:
        risk_penalty += 5
    return (
        personal_bonus
        + city_bonus
        + digital_bonus
        + comp_bonus
        + ai_bonus
        + electronic_bonus
        + mech_bonus
        + teacher_bonus
        + wider_bonus
        - risk_penalty
    )


def select_main_groups(personal_rows, candidate_rows, source_label):
    personal_by_code = {code(row): row for row in personal_rows}
    combined = []
    seen = set()
    for row in candidate_rows:
        if not hard_main_ok(row):
            continue
        out = dict(row)
        p = personal_by_code.get(code(row))
        out["来自个人适配"] = "true" if p and is_personal_normal(p) else "false"
        if p:
            out["适配理由"] = p.get("适配理由", "")
            out["关键风险"] = p.get("关键风险", "")
            out["家庭讨论动作"] = p.get("家庭讨论动作", "")
        out["来源候选层"] = f"personal-fit-v1+{source_label}" if p else source_label
        key = group_key(out)
        if key and key not in seen:
            combined.append(out)
            seen.add(key)

    selected = []
    selected_keys = set()

    def add_from(predicate, limit):
        candidates = [row for row in combined if group_key(row) not in selected_keys and predicate(row)]
        candidates.sort(
            key=lambda r: (
                -group_score(r),
                float_value(r, "历史最高等位分差"),
                r.get("院校代码", ""),
                r.get("院校专业组代码OCR规范化", ""),
            )
        )
        for row in candidates[:limit]:
            selected.append(row)
            selected_keys.add(group_key(row))

    add_from(lambda r: bucket(r) in {"冲刺", "冲刺待补历史"}, 15)
    add_from(lambda r: bucket(r) == "稳冲", 15)
    add_from(lambda r: bucket(r) in {"稳妥", "保底"}, 15)

    if len(selected) < 45:
        add_from(lambda r: True, 45 - len(selected))
    if len(selected) != 45:
        raise SystemExit(f"主草表未能形成45个院校专业组：{len(selected)}")

    # Ensure official-style order: front冲, middle稳, back保.
    selected.sort(
        key=lambda r: (
            {"冲刺": 0, "冲刺待补历史": 0, "稳冲": 1, "稳妥": 2, "保底": 2}.get(bucket(r), 3),
            -group_score(r),
            float_value(r, "历史最高等位分差"),
            r.get("院校代码", ""),
        )
    )
    return selected[:45]


def major_score(row):
    name = row.get("专业名称及备注短摘", "")
    pref = row.get("专业偏好方向", "")
    direction = row.get("第二轮方向标签", "")
    role = row.get("专业角色判断", "")
    risk = "；".join([row.get("专业风险类型", ""), row.get("机器专业接受度初判", "")])
    score = 0
    if "数字媒体技术" in pref or "数字媒体" in name or "数字媒体" in direction:
        score += 120
    if "计算机类相关" in pref or "计算机AI软件" in direction:
        score += 90
    if "电子信息网络" in direction:
        score += 72
    if "机械自动化机器人" in direction:
        score += 58
    if "师范类相关" in pref or "师范" in role:
        score += 55
    if "环境工程科学" in direction:
        score += 24
    if "工商旅游管理" in direction:
        score += 16
    for token in ["软件", "数据", "人工智能", "物联网", "网络", "信息安全", "智能科学", "大数据"]:
        if token in name:
            score += 12
    for token in ["医学", "护理", "康复", "药学", "临床", "动物医学", "兽医", "助产", "中外合作", "高收费"]:
        if token in name or token in risk:
            score -= 500
    if row.get("字段缺口数", "") not in {"", "0"}:
        score -= 3
    return score


def select_six_majors(rows):
    ordered = sorted(
        rows,
        key=lambda r: (
            -major_score(r),
            int_value(r, "专业组内专业序号"),
            r.get("专业代号OCR", ""),
        ),
    )
    return ordered[:6]


def format_major(row):
    code_text = row.get("专业代号OCR", "")
    name = row.get("专业名称及备注短摘", "")
    return f"{code_text} {name}".strip()


def pending_major_excerpt(rows, selected_ids):
    pending = []
    for row in rows:
        if row.get("专业行ID") in selected_ids:
            continue
        name = row.get("专业名称及备注短摘", "")
        role = row.get("专业角色判断", "")
        if not role:
            role = row.get("机器专业接受度初判", "") or "待判断"
        pending.append(f"{row.get('专业代号OCR', '')} {name}({role})")
    return "；".join(pending[:8])


def verification_priority(row):
    text = risk_text(row)
    if bucket(row) == "冲刺待补历史" or "结构" in text:
        return "V0-先核结构边界和历史映射"
    if bucket(row) == "冲刺":
        return "V1-先核冲刺线和章程限制"
    if bucket(row) == "稳冲":
        return "V1-核计划数、组内专业和调剂"
    if bucket(row) in {"稳妥", "保底"}:
        return "V2-核保底安全和完整调剂"
    return "V3-备选核验"


def discussion_batch(volunteer_no):
    if volunteer_no <= 15:
        return "B1-冲刺和城市观察批"
    if volunteer_no <= 30:
        return "B2-稳冲主线批"
    return "B3-稳妥和保底批"


def build_outputs(main_groups, major_rows, personal_rows):
    majors_by_group = defaultdict(list)
    for row in major_rows:
        majors_by_group[group_key(row)].append(row)
        majors_by_group[code(row)].append(row)

    main_rows = []
    major_choice_rows = []
    for idx, row in enumerate(main_groups, start=1):
        group_majors = majors_by_group.get(row.get("专业组出现ID", "")) or majors_by_group.get(code(row), [])
        six = select_six_majors(group_majors)
        selected_ids = {m.get("专业行ID") for m in six}
        reason = row.get("适配理由") or row.get("入选理由", "")
        risk = row.get("关键风险") or row.get("风险提示", "")
        out = {
            "志愿序号": str(idx),
            "草案梯度": bucket(row),
            "讨论批次": discussion_batch(idx),
            "二次核验优先级": verification_priority(row),
            "是否进入主草表": "true",
            "是否可作为定稿依据": "false",
            "推荐进入原因": reason,
            "核心风险": risk,
            "院校代码": row.get("院校代码", ""),
            "院校名称OCR": row.get("院校名称OCR", ""),
            "院校专业组代码OCR规范化": code(row),
            "专业组号OCR": row.get("专业组号OCR", ""),
            "城市候选": row.get("城市候选", ""),
            "公办民办机器线索": row.get("公办民办机器线索", ""),
            "家庭底线属性动作": row.get("家庭底线属性动作", ""),
            "历史线索分层": row.get("历史线索分层", ""),
            "历史最高等位分差": row.get("历史最高等位分差", ""),
            "历史最低等位分差": row.get("历史最低等位分差", ""),
            "偏好专业数": row.get("偏好专业数", ""),
            "数字媒体技术专业数": row.get("数字媒体技术专业数", ""),
            "计算机类相关专业数": row.get("计算机类相关专业数", ""),
            "师范类相关专业数": row.get("师范类相关专业数", ""),
            "专业明细行数": row.get("专业明细行数", ""),
            "组内待确认专业摘录": pending_major_excerpt(group_majors, selected_ids),
            "服从调剂草案": "待定-完整专业组核验且家庭确认组内专业可接受后，原则上选择服从以降低退档风险。",
            "二次核验动作": "核第19期原页、湖北官方系统或省招办计划、高校章程、完整专业组、6个专业顺序、学费、计划数、选科、校区和限制条件。",
            "第19期页码": row.get("来源页码", "") or row.get("第19期页码", ""),
            "版面列": row.get("版面列", ""),
            "专业组出现ID": row.get("专业组出现ID", ""),
            "来源候选层": row.get("来源候选层", ""),
        }
        for major_idx in range(6):
            out[f"建议专业{major_idx + 1}"] = format_major(six[major_idx]) if major_idx < len(six) else ""
        main_rows.append(out)

        for major_rank, major in enumerate(
            sorted(
                group_majors,
                key=lambda r: (
                    0 if r.get("专业行ID") in selected_ids else 1,
                    -major_score(r),
                    int_value(r, "专业组内专业序号"),
                ),
            ),
            start=1,
        ):
            item = {
                "志愿序号": str(idx),
                "草案梯度": bucket(row),
                "是否建议填入6专业": "true" if major.get("专业行ID") in selected_ids else "false",
                "专业选择顺位": str(major_rank) if major.get("专业行ID") in selected_ids else "",
            }
            for field in MAJOR_FIELDS:
                item.setdefault(field, major.get(field, ""))
            item["人工核验优先级"] = item.get("人工核验优先级", "").replace("最终候选", "本轮候选")
            item["院校代码"] = row.get("院校代码", "")
            item["院校名称OCR"] = row.get("院校名称OCR", "")
            item["院校专业组代码OCR规范化"] = code(row)
            item["第19期页码"] = major.get("来源页码", "")
            major_choice_rows.append(item)

    special_rows = []
    for idx, row in enumerate(
        [r for r in personal_rows if r.get("候选主线") == "7万内中外合作/高收费专项"],
        start=1,
    ):
        special_rows.append(
            {
                "专项序号": str(idx),
                "专项层级": row.get("本轮分层", ""),
                "家庭讨论动作": row.get("家庭讨论动作", ""),
                "是否进入专项讨论": row.get("是否进入本轮家庭讨论", ""),
                "是否可作为定稿依据": "false",
                "适配理由": row.get("适配理由", ""),
                "关键风险": row.get("关键风险", ""),
                "院校代码": row.get("院校代码", ""),
                "院校名称OCR": row.get("院校名称OCR", ""),
                "院校专业组代码OCR规范化": code(row),
                "城市候选": row.get("城市候选", ""),
                "历史线索分层": row.get("历史线索分层", ""),
                "历史最高等位分差": row.get("历史最高等位分差", ""),
                "场景判断": row.get("场景判断", ""),
                "场景内最低学费": row.get("场景内最低学费", ""),
                "场景内最高学费": row.get("场景内最高学费", ""),
                "预算内中外合作或高收费专业数": row.get("预算内中外合作或高收费专业数", ""),
                "超过预算专业数": row.get("超过预算专业数", ""),
                "待核费用专业数": row.get("待核费用专业数", ""),
                "第19期页码": row.get("第19期页码", ""),
                "版面列": row.get("版面列", ""),
                "组内完整专业清单索引": row.get("组内完整专业清单索引", ""),
                "专业组出现ID": row.get("专业组出现ID", ""),
                "下一步核验动作": row.get("下一步核验动作", ""),
            }
        )

    discussion_rows = [
        {
            "讨论批次": "B1-冲刺和城市观察批",
            "批次目标": "先判断愿不愿意把更理想但风险更高的组放在前面。",
            "包含志愿序号": "1-15",
            "院校专业组数量": "15",
            "优先讨论问题": "历史线是否偏高、H0是否能补到替代历史线、是否接受城市/学校/调剂代价。",
            "完成条件": "每个组完成原页和历史线二次核验；高冲项不挤占稳保名额。",
        },
        {
            "讨论批次": "B2-稳冲主线批",
            "批次目标": "决定主方案核心学校和专业方向。",
            "包含志愿序号": "16-30",
            "院校专业组数量": "15",
            "优先讨论问题": "6个专业顺序、完整组内调剂可接受度、计划数和章程限制。",
            "完成条件": "家庭确认可接受专业和不可接受专业；可决定是否服从调剂。",
        },
        {
            "讨论批次": "B3-稳妥和保底批",
            "批次目标": "保证录取成功率，避免全部集中在冲稳。",
            "包含志愿序号": "31-45",
            "院校专业组数量": "15",
            "优先讨论问题": "学校/城市是否能接受、保底组是否真的没有调剂硬伤。",
            "完成条件": "至少保留足够数量的低风险保底组，且每组完整核验调剂范围。",
        },
        {
            "讨论批次": "B4-高费专项备选",
            "批次目标": "只在家庭明确接受更高预算时讨论，不进入普通主草表。",
            "包含志愿序号": "专项表",
            "院校专业组数量": str(len(special_rows)),
            "优先讨论问题": "预算、证书、出国安排、英语/单科限制、不得转专业、真实性价比。",
            "完成条件": "家庭明确预算和培养模式接受边界后，再决定是否替换主草表中的某些位置。",
        },
    ]
    return main_rows, major_choice_rows, discussion_rows, special_rows


def write_workbook(main_rows, major_rows, discussion_rows, special_rows):
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name, rows, fields):
        ws = wb.create_sheet(name)
        ws.append(fields)
        fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, field in enumerate(fields, start=1):
            sample = [len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:120]]
            ws.column_dimensions[get_column_letter(idx)].width = max(10, min(42, max(sample) + 2))
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "本科普通批院校专业组 V1 草案，只用于家庭讨论和二次核验，不作为定稿依据。"),
        ("主草表", "01_主草表45组按前冲、中稳、后保组织，仍需逐组核验。"),
        ("专业顺序", "02_专业选择草案给出每组建议优先看的专业，正式6个专业顺序须核完整专业组后再定。"),
        ("专项", "04_高费专项备选不进入普通主草表，除非家庭明确接受预算和培养模式。"),
        ("二次核验", "每个入围组必须核第19期原页、湖北官方侧、章程、学费、计划数、选科、校区、限制和调剂范围。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 100

    add_sheet("01_主草表45组", main_rows, MAIN_FIELDS)
    add_sheet("02_专业选择草案", major_rows, MAJOR_FIELDS)
    add_sheet("03_分批讨论与核验", discussion_rows, DISCUSSION_FIELDS)
    add_sheet("04_高费专项备选", special_rows, SPECIAL_FIELDS)
    wb.save(WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "private/",
        "private\\",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "人工读数",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "最终候选",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"志愿草表公开文件包含禁止内容：{hit}")


def main():
    personal_rows = read_csv(PERSONAL_GROUPS)
    if ROUND2_MAIN_SHORTLIST_GROUPS.exists() and ROUND2_MAIN_SHORTLIST_MAJORS.exists():
        candidate_rows = read_csv(ROUND2_MAIN_SHORTLIST_GROUPS)
        major_rows = read_csv(ROUND2_MAIN_SHORTLIST_MAJORS)
        source_label = "round2-updated-preferences"
    else:
        candidate_rows = read_csv(ROUND1_SHORTLIST_GROUPS)
        major_rows = read_csv(STABLE_MAJOR_BROWSER)
        source_label = "round1-shortlist"

    main_groups = select_main_groups(personal_rows, candidate_rows, source_label)
    main_rows, major_choice_rows, discussion_rows, special_rows = build_outputs(
        main_groups,
        major_rows,
        personal_rows,
    )

    write_csv(MAIN_TABLE, main_rows, MAIN_FIELDS)
    write_csv(MAJOR_CHOICES, major_choice_rows, MAJOR_FIELDS)
    write_csv(DISCUSSION_BATCHES, discussion_rows, DISCUSSION_FIELDS)
    write_csv(SPECIAL_OPTIONS, special_rows, SPECIAL_FIELDS)
    write_workbook(main_rows, major_choice_rows, discussion_rows, special_rows)

    summary = {
        "status": "issue19_volunteer_table_v1_draft_ready",
        "generated_by": "build_issue19_volunteer_table_v1_draft.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "usage_boundary": "本科普通批院校专业组草表，只用于家庭讨论和二次核验，不作为定稿依据。",
        "fixed_inputs": {
            "省份": "湖北",
            "年份": 2026,
            "类别": "普通类首选物理",
            "总分": 515,
            "累计位次": 91723,
            "等位分": {"2025": 494, "2024": 497, "2023": 481},
            "主线": "公办普通学费优先，不学医",
        },
        "official_rule_basis": [
            "本科普通批可填45个院校专业组志愿。",
            "每个院校专业组志愿内可选不超过6个专业及是否服从专业调剂。",
            "平行志愿按分数优先、遵循志愿顺序检索。",
        ],
        "main_table": str(MAIN_TABLE.relative_to(ROOT)),
        "major_choices": str(MAJOR_CHOICES.relative_to(ROOT)),
        "discussion_batches": str(DISCUSSION_BATCHES.relative_to(ROOT)),
        "special_options": str(SPECIAL_OPTIONS.relative_to(ROOT)),
        "workbook": str(WORKBOOK.relative_to(ROOT)),
        "candidate_source": source_label,
        "main_group_count": len(main_rows),
        "major_choice_row_count": len(major_choice_rows),
        "discussion_batch_count": len(discussion_rows),
        "special_option_count": len(special_rows),
        "gradient_counts": dict(Counter(row["草案梯度"] for row in main_rows)),
        "discussion_batch_counts": dict(Counter(row["讨论批次"] for row in main_rows)),
        "risk_attention_counts": {
            "保底组数量": sum(row["草案梯度"] == "保底" for row in main_rows),
            "冲刺待补历史组数量": sum(row["草案梯度"] == "冲刺待补历史" for row in main_rows),
        },
        "gate": {
            "all_main_rows_not_final_basis": all(row["是否可作为定稿依据"] == "false" for row in main_rows),
            "all_main_rows_enter_main_draft": all(row["是否进入主草表"] == "true" for row in main_rows),
            "special_options_separate": True,
        },
        "must_verify": [
            "第19期原页或纸质原页",
            "湖北官方系统或省招办计划",
            "高校招生章程",
            "完整院校专业组和调剂范围",
            "6个专业顺序、计划数、学费、选科、校区、语种/单科/体检限制",
        ],
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    public_text_safe([MAIN_TABLE, MAJOR_CHOICES, DISCUSSION_BATCHES, SPECIAL_OPTIONS, SUMMARY_JSON])
    print(f"写出本科普通批主草表：{MAIN_TABLE.relative_to(ROOT)}")
    print(f"写出专业选择草案：{MAJOR_CHOICES.relative_to(ROOT)}")
    print(f"写出讨论批次：{DISCUSSION_BATCHES.relative_to(ROOT)}")
    print(f"写出专项备选：{SPECIAL_OPTIONS.relative_to(ROOT)}")
    print(f"写出工作簿：{WORKBOOK.relative_to(ROOT)}")
    print(f"主草表院校专业组数量：{len(main_rows)}")


if __name__ == "__main__":
    main()
