#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

RESULT = WORKING / "issue19-first-closure-verification-result-public-ledger.csv"
PAGE_RESULT = WORKING / "issue19-first-closure-verification-result-page-summary.csv"
FIELD_STATUS = WORKING / "issue19-first-closure-field-verification-status-public-ledger.csv"

PAGE_OUTPUT = EXPORTS / "issue19-first-closure-manual-verification-page-packets.csv"
TASK_OUTPUT = EXPORTS / "issue19-first-closure-manual-verification-task-items.csv"
FIELD_OUTPUT = EXPORTS / "issue19-first-closure-manual-verification-field-items.csv"
SUMMARY_OUTPUT = EXPORTS / "issue19-first-closure-manual-verification-workbook-summary.json"
WORKBOOK_OUTPUT = EXPORTS / "issue19-first-closure-manual-verification-workbook.xlsx"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
DATA_STAGE = "issue19_first_closure_manual_verification_workbook"
GENERATED_AT = "2026-06-29"

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "可否进入最终志愿方案",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

BASE_FIELDS = [
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "任务粒度",
    *FALSE_FIELDS,
]

PAGE_FIELDS = [
    "第一闭环人工核验页列包ID",
    *BASE_FIELDS,
    "人工核验包序号",
    "来源页码",
    "版面列",
    "页码版面键",
    "执行泳道",
    "页列任务数",
    "字段核验行数",
    "涉及专业行数",
    "涉及院校代码数",
    "待核字段分布",
    "任务结果桶分布",
    "PDF原页证据状态分布",
    "OCR提示状态分布",
    "机器坐标提示状态分布",
    "高校辅证证据状态分布",
    "冲突状态分布",
    "需要双人复核任务数",
    "需要人工直接看图任务数",
    "建议人工核验小包类型",
    "建议每包任务上限",
    "建议核验顺序",
    "PDF原页核验操作",
    "湖北官方侧核验操作",
    "高校官网辅证操作",
    "完成条件",
    "公开填写边界",
    "关联任务集合SHA16",
    "关联字段集合SHA16",
    "公开安全策略",
]

TASK_FIELDS = [
    "第一闭环人工核验任务ID",
    *BASE_FIELDS,
    "人工核验包序号",
    "稳定基座第一闭环明细任务ID",
    "稳定基座第一闭环页列包ID",
    "来源页码",
    "版面列",
    "页码版面键",
    "专业行ID",
    "专业组出现ID",
    "院校代码",
    "任务来源类型",
    "字段名",
    "执行泳道",
    "核验动作层级",
    "任务结果桶",
    "PDF原页证据状态",
    "OCR提示状态",
    "机器坐标提示状态",
    "高校辅证证据状态",
    "湖北官方侧状态",
    "冲突状态",
    "三方闭环状态",
    "字段写回门禁",
    "是否需要双人复核",
    "是否需要人工直接看图",
    "完成证据要求",
    "当前阻断原因",
    "下一步核验动作",
    "公开人工记录栏",
    "公开填写边界",
    "公开安全策略",
]

FIELD_FIELDS = [
    "第一闭环人工核验字段ID",
    *BASE_FIELDS,
    "人工核验包序号",
    "稳定基座第一闭环明细任务ID",
    "来源页码",
    "版面列",
    "页码版面键",
    "专业行ID",
    "专业组出现ID",
    "院校代码",
    "字段名",
    "字段映射状态",
    "字段事实状态",
    "字段核验优先级",
    "字段PDF核验状态",
    "字段湖北官方核验状态",
    "字段事实闭环等级",
    "字段事实阻断等级",
    "字段事实缺口类型",
    "PDF原页证据状态",
    "OCR提示状态",
    "机器坐标提示状态",
    "高校辅证证据状态",
    "冲突状态",
    "核验动作层级",
    "执行泳道",
    "当前阻断原因",
    "下一步核验动作",
    "公开人工记录栏",
    "公开填写边界",
    "公开安全策略",
]

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/home/",
    "/var/folders/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    "file://",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".tif",
    ".tiff",
    ".heic",
    "Authorization",
    "Bearer ",
    "Cookie",
    "Set-Cookie",
    "access_token",
    "refresh_token",
    "password",
    "secret",
    "api_key",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "OCR行文本",
    "字段确认值",
    "人工读数",
    "候选值",
    "PDF原页人工读数",
    "湖北官方字段值",
    "高校官网或招生章程字段值",
    "已确认",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
    "可排序",
]


def clean(value):
    return "" if value is None else str(value).replace("\r", " ").replace("\n", " ").strip()


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in fields} for row in rows])


def stable_id(prefix, parts):
    text = "|".join(clean(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(text.encode('utf-8')).hexdigest()[:16]}"


def sha16(values):
    normalized = "\n".join(sorted({clean(value) for value in values if clean(value)}))
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:16]


def as_int(value):
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def counter_text(rows, field):
    counter = Counter(clean(row.get(field)) for row in rows if clean(row.get(field)))
    return "；".join(f"{key}×{value}" for key, value in sorted(counter.items()))


def base_row(grain, task_grain):
    row = {
        "来源期号": SOURCE_ISSUE,
        "来源PDF_SHA256": SOURCE_PDF_SHA256,
        "生成日期": GENERATED_AT,
        "数据阶段": DATA_STAGE,
        "主表粒度": grain,
        "任务粒度": task_grain,
    }
    row.update({field: "false" for field in FALSE_FIELDS})
    return row


def page_action(row):
    lane = clean(row.get("执行泳道"))
    if lane.startswith("E0"):
        return "先做双人回看PDF原页，再核湖北官方侧；冲突字段不得由OCR或高校源直接定案。"
    if lane.startswith("E1"):
        return "先补计划数、学费或选科缺口，再用高校源作定位提示，最后核湖北官方侧。"
    return "先核专业名、专业组边界和备注归属，再决定是否需要逐字段复核。"


def school_action(row):
    tasks = as_int(row.get("高校辅证线索任务数"))
    if tasks:
        return "高校官网只作double check：记录来源是否能解释差异，但不能替代湖北官方计划或PDF原页。"
    return "本页列当前无高校字段线索，先以PDF原页和湖北官方侧为准。"


def manual_boundary():
    return "公开表只保留状态、计数、ID和SHA；具体核验内容、字段内容、OCR原文和截图路径不得填入公开仓库。"


def public_policy():
    return "not_final；for_manual_execution_only；no_field_values；no_private_paths；no_ocr_text；no_recommendation"


def workbook_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([clean(row.get(field, "")) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 28,
        "B": 28,
        "C": 66,
        "D": 18,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    return ws


def assert_public_safe(rows, label):
    text = json.dumps(rows, ensure_ascii=False)
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]
    if found:
        raise SystemExit(f"{label} contains forbidden public tokens: {found[:5]}")


def main():
    result_rows = read_csv(RESULT)
    page_rows = read_csv(PAGE_RESULT)
    field_rows = read_csv(FIELD_STATUS)

    tasks_by_page = defaultdict(list)
    fields_by_page = defaultdict(list)
    for row in result_rows:
        tasks_by_page[clean(row.get("页码版面键"))].append(row)
    for row in field_rows:
        fields_by_page[clean(row.get("页码版面键"))].append(row)

    page_packets = []
    for index, source in enumerate(page_rows, start=1):
        key = clean(source.get("页码版面键"))
        tasks = tasks_by_page[key]
        fields = fields_by_page[key]
        row = base_row("第一闭环页列人工核验包", "PDF页码×版面列")
        row.update(
            {
                "第一闭环人工核验页列包ID": stable_id("FCMV-PAGE", [SOURCE_PDF_SHA256, key]),
                "人工核验包序号": index,
                "来源页码": source.get("来源页码", ""),
                "版面列": source.get("版面列", ""),
                "页码版面键": key,
                "执行泳道": source.get("执行泳道", ""),
                "页列任务数": len(tasks),
                "字段核验行数": len(fields),
                "涉及专业行数": len({row.get("专业行ID", "") for row in tasks if row.get("专业行ID")}),
                "涉及院校代码数": len({row.get("院校代码", "") for row in tasks if row.get("院校代码")}),
                "待核字段分布": counter_text(fields, "字段名"),
                "任务结果桶分布": counter_text(tasks, "任务结果桶"),
                "PDF原页证据状态分布": counter_text(tasks, "PDF原页证据状态"),
                "OCR提示状态分布": counter_text(tasks, "OCR提示状态"),
                "机器坐标提示状态分布": counter_text(tasks, "机器坐标提示状态"),
                "高校辅证证据状态分布": counter_text(tasks, "高校辅证证据状态"),
                "冲突状态分布": counter_text(tasks, "冲突状态"),
                "需要双人复核任务数": sum(row.get("是否需要双人复核") == "true" for row in tasks),
                "需要人工直接看图任务数": sum(row.get("是否需要人工直接看图") == "true" for row in tasks),
                "建议人工核验小包类型": source.get("建议人工核验小包类型", ""),
                "建议每包任务上限": source.get("建议每包任务上限", ""),
                "建议核验顺序": page_action(source),
                "PDF原页核验操作": "按页码和版面列打开第19期原页，逐字段核专业组边界、专业行归属、计划数、学费和选科。",
                "湖北官方侧核验操作": "在湖北官方系统或省招办计划可用时逐项核同一专业组和专业行；未核前保持pending。",
                "高校官网辅证操作": school_action(source),
                "完成条件": source.get("完成条件", ""),
                "公开填写边界": manual_boundary(),
                "关联任务集合SHA16": sha16(row.get("稳定基座第一闭环明细任务ID") for row in tasks),
                "关联字段集合SHA16": sha16(row.get("第一闭环字段核验状态ID") for row in fields),
                "公开安全策略": public_policy(),
            }
        )
        page_packets.append(row)

    page_order = {row["页码版面键"]: row["人工核验包序号"] for row in page_packets}

    task_items = []
    for source in result_rows:
        row = base_row("第一闭环明细任务人工核验项", "逐专业招生明细×第一闭环任务")
        key = clean(source.get("页码版面键"))
        row.update(
            {
                "第一闭环人工核验任务ID": stable_id(
                    "FCMV-TASK",
                    [SOURCE_PDF_SHA256, source.get("稳定基座第一闭环明细任务ID", "")],
                ),
                "人工核验包序号": page_order.get(key, ""),
                "稳定基座第一闭环明细任务ID": source.get("稳定基座第一闭环明细任务ID", ""),
                "稳定基座第一闭环页列包ID": source.get("稳定基座第一闭环页列包ID", ""),
                "来源页码": source.get("来源页码", ""),
                "版面列": source.get("版面列", ""),
                "页码版面键": key,
                "专业行ID": source.get("专业行ID", ""),
                "专业组出现ID": source.get("专业组出现ID", ""),
                "院校代码": source.get("院校代码", ""),
                "任务来源类型": source.get("任务来源类型", ""),
                "字段名": source.get("字段名", ""),
                "执行泳道": source.get("执行泳道", ""),
                "核验动作层级": source.get("核验动作层级", ""),
                "任务结果桶": source.get("任务结果桶", ""),
                "PDF原页证据状态": source.get("PDF原页证据状态", ""),
                "OCR提示状态": source.get("OCR提示状态", ""),
                "机器坐标提示状态": source.get("机器坐标提示状态", ""),
                "高校辅证证据状态": source.get("高校辅证证据状态", ""),
                "湖北官方侧状态": source.get("湖北官方侧状态", ""),
                "冲突状态": source.get("冲突状态", ""),
                "三方闭环状态": source.get("三方闭环状态", ""),
                "字段写回门禁": source.get("字段写回门禁", ""),
                "是否需要双人复核": source.get("是否需要双人复核", ""),
                "是否需要人工直接看图": source.get("是否需要人工直接看图", ""),
                "完成证据要求": source.get("完成证据要求", ""),
                "当前阻断原因": source.get("当前阻断原因", ""),
                "下一步核验动作": source.get("下一步核验动作", ""),
                "公开人工记录栏": "留空；具体核验内容只允许写入私有复核表，不写入公开仓库。",
                "公开填写边界": manual_boundary(),
                "公开安全策略": public_policy(),
            }
        )
        task_items.append(row)

    field_items = []
    for source in field_rows:
        row = base_row("第一闭环字段人工核验项", "第一闭环明细任务×待核字段")
        key = clean(source.get("页码版面键"))
        row.update(
            {
                "第一闭环人工核验字段ID": stable_id(
                    "FCMV-FIELD",
                    [
                        SOURCE_PDF_SHA256,
                        source.get("稳定基座第一闭环明细任务ID", ""),
                        source.get("字段名", ""),
                    ],
                ),
                "人工核验包序号": page_order.get(key, ""),
                "稳定基座第一闭环明细任务ID": source.get("稳定基座第一闭环明细任务ID", ""),
                "来源页码": source.get("来源页码", ""),
                "版面列": source.get("版面列", ""),
                "页码版面键": key,
                "专业行ID": source.get("专业行ID", ""),
                "专业组出现ID": source.get("专业组出现ID", ""),
                "院校代码": source.get("院校代码", ""),
                "字段名": source.get("字段名", ""),
                "字段映射状态": source.get("字段映射状态", ""),
                "字段事实状态": source.get("字段事实状态", ""),
                "字段核验优先级": source.get("字段核验优先级", ""),
                "字段PDF核验状态": source.get("字段PDF核验状态", ""),
                "字段湖北官方核验状态": source.get("字段湖北官方核验状态", ""),
                "字段事实闭环等级": source.get("字段事实闭环等级", ""),
                "字段事实阻断等级": source.get("字段事实阻断等级", ""),
                "字段事实缺口类型": source.get("字段事实缺口类型", ""),
                "PDF原页证据状态": source.get("PDF原页证据状态", ""),
                "OCR提示状态": source.get("OCR提示状态", ""),
                "机器坐标提示状态": source.get("机器坐标提示状态", ""),
                "高校辅证证据状态": source.get("高校辅证证据状态", ""),
                "冲突状态": source.get("冲突状态", ""),
                "核验动作层级": source.get("核验动作层级", ""),
                "执行泳道": source.get("执行泳道", ""),
                "当前阻断原因": source.get("当前阻断原因", ""),
                "下一步核验动作": source.get("下一步核验动作", ""),
                "公开人工记录栏": "留空；字段具体内容只允许写入私有复核表，不写入公开仓库。",
                "公开填写边界": manual_boundary(),
                "公开安全策略": public_policy(),
            }
        )
        field_items.append(row)

    for rows, label in [
        (page_packets, "page_packets"),
        (task_items, "task_items"),
        (field_items, "field_items"),
    ]:
        assert_public_safe(rows, label)

    write_csv(PAGE_OUTPUT, page_packets, PAGE_FIELDS)
    write_csv(TASK_OUTPUT, task_items, TASK_FIELDS)
    write_csv(FIELD_OUTPUT, field_items, FIELD_FIELDS)

    wb = Workbook()
    wb.remove(wb.active)
    workbook_sheet(wb, "页列核验包37", page_packets, PAGE_FIELDS)
    workbook_sheet(wb, "任务核验项206", task_items, TASK_FIELDS)
    workbook_sheet(wb, "字段核验项354", field_items, FIELD_FIELDS)
    wb.save(WORKBOOK_OUTPUT)

    summary = {
        "status": "issue19_first_closure_manual_verification_workbook_ready_not_final",
        "generated_by": "build_issue19_first_closure_manual_verification_workbook.py",
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "page_packet_output": str(PAGE_OUTPUT.relative_to(ROOT)),
        "task_item_output": str(TASK_OUTPUT.relative_to(ROOT)),
        "field_item_output": str(FIELD_OUTPUT.relative_to(ROOT)),
        "workbook_output": str(WORKBOOK_OUTPUT.relative_to(ROOT)),
        "page_packet_count": len(page_packets),
        "task_item_count": len(task_items),
        "field_item_count": len(field_items),
        "field_counts": dict(Counter(row["字段名"] for row in field_items)),
        "page_packet_lane_counts": dict(Counter(row["执行泳道"] for row in page_packets)),
        "double_review_task_count": sum(row["是否需要双人复核"] == "true" for row in task_items),
        "direct_image_review_task_count": sum(row["是否需要人工直接看图"] == "true" for row in task_items),
        "final_available_count": 0,
        "recommendation_basis_allowed_count": 0,
        "field_writeback_allowed_count": 0,
        "policy": "公开工作簿只作人工核验执行清单；不得填写具体核验内容、字段内容、OCR原文、截图路径或最终建议。",
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text = (
        PAGE_OUTPUT.read_text(encoding="utf-8", errors="ignore")
        + TASK_OUTPUT.read_text(encoding="utf-8", errors="ignore")
        + FIELD_OUTPUT.read_text(encoding="utf-8", errors="ignore")
        + SUMMARY_OUTPUT.read_text(encoding="utf-8", errors="ignore")
    )
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in public_text]
    if found:
        raise SystemExit(f"public output contains forbidden tokens: {found[:5]}")

    print(f"wrote {PAGE_OUTPUT}")
    print(f"wrote {TASK_OUTPUT}")
    print(f"wrote {FIELD_OUTPUT}")
    print(f"wrote {SUMMARY_OUTPUT}")
    print(f"wrote {WORKBOOK_OUTPUT}")


if __name__ == "__main__":
    main()
