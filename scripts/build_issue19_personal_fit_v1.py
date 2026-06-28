#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

STABLE_GROUP_BROWSER = EXPORTS / "issue19-stable-foundation-group-browser.csv"
STABLE_MAJOR_BROWSER = EXPORTS / "issue19-stable-foundation-major-browser.csv"
ROUND1_GROUPS = EXPORTS / "issue19-round1-candidate-groups.csv"
EXPANDED_GROUPS = EXPORTS / "issue19-expanded-budget-coop-groups.csv"

GROUP_OUTPUT = EXPORTS / "issue19-personal-fit-v1-groups.csv"
MAJOR_OUTPUT = EXPORTS / "issue19-personal-fit-v1-major-details.csv"
SUMMARY_OUTPUT = EXPORTS / "issue19-personal-fit-v1-summary.json"
WORKBOOK_OUTPUT = EXPORTS / "issue19-personal-fit-v1.xlsx"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

GROUP_FIELDS = [
    "个性化候选ID",
    "候选主线",
    "本轮分层",
    "家庭讨论动作",
    "是否进入本轮家庭讨论",
    "是否可作为定稿依据",
    "适配理由",
    "关键风险",
    "下一步核验动作",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "医学护理排除专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "历史线索分层",
    "同代码命中年份最大值",
    "历史最高等位分差",
    "历史最低等位分差",
    "场景判断",
    "场景内最低学费",
    "场景内最高学费",
    "预算内中外合作或高收费专业数",
    "超过预算专业数",
    "待核费用专业数",
    "第19期页码",
    "版面列",
    "组内完整专业清单索引",
    "专业组出现ID",
    "来源表",
]

MAJOR_FIELDS = [
    "个性化候选ID",
    "候选主线",
    "本轮分层",
    "家庭讨论动作",
    "是否进入本轮家庭讨论",
    "是否可作为定稿依据",
    "专业角色判断",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "第19期页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
    "下一步核验动作",
]


NORMAL_CANDIDATES = [
    ("C13107", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办工科，计算机/人工智能/自动化/机器人等方向集中，H2 历史线索贴近等位分。", "组内 15 个专业，需确认汽车产业方向、校区和完整调剂范围。"),
    ("C13908", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办，计算机/物联网/智能科学/机器人与师范方向都有线索，H2。", "组内含机械、材料、飞行器制造等，家庭必须判断能否接受。"),
    ("F30508", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "命中数字媒体技术，同时有计算机与师范线索，H2。", "城市和学校层级一般，需确认数字媒体专业确在该组且计划数无误。"),
    ("F59704", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "电气、电子、自动化、机器人、计算机、软件、数据科学、人工智能方向较完整，H2。", "组内专业多，土木/焊接/水利等非偏好方向需判断调剂接受度。"),
    ("C13810", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办，信息与计算、大数据、人工智能等方向贴近，H3 稳冲。", "需核创新班、计划数、专业代号和近三年同组稳定性。"),
    ("C12903", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办，信息安全、计算机、电子、智能制造等方向可讨论，H3 贴线。", "偏贴线，组内非计算机方向也不少，需做整组接受度判断。"),
    ("C12706", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办，数据科学、人工智能、软件、网络空间安全方向贴合，H3。", "带稳冲属性，需确认计划数和网络空间安全方向接受度。"),
    ("C12610", "普通公办低费用主线", "A-主线优先讨论", "进入家庭讨论", "湖北省内公办，软件、物联网、电子信息方向可讨论，H3。", "组内车辆/交通方向较多，需要判断是否愿意服从调剂。"),
    ("F30805", "普通公办低费用主线", "B-保底备选讨论", "进入家庭讨论", "H2 且最高等位分差较低，人工智能/计算机/软件与师范方向同组。", "需核 OCR 是否把专业名连读，食品等非偏好专业也要接受。"),
    ("K78401", "普通公办低费用主线", "B-保底备选讨论", "进入家庭讨论", "H1 保底线索，计算机、物联网、数据科学等方向存在。", "地理距离远，组内材料、测绘、新能源、资源勘查等需接受。"),
    ("C15108", "普通公办低费用主线", "B-保底备选讨论", "进入家庭讨论", "湖北省内公办师范线，H2，若愿意了解教师路线可讨论。", "组内有特殊限制待核，且师范意愿尚未完全明确。"),
    ("K27203", "普通公办低费用主线", "B-保底备选讨论", "进入家庭讨论", "H2，数学/生物师范与物联网方向可作保底讨论。", "城市远，冶金等非偏好专业需接受。"),
    ("K75703", "普通公办低费用主线", "B-保底备选讨论", "进入家庭讨论", "H2，通信、物联网方向，组小。", "需核产教融合或合作班性质、费用和培养模式。"),
    ("C12815", "普通公办低费用主线", "C-冲刺观察讨论", "进入家庭讨论", "武汉公办，计算机、软件、物联网、人工智能、电子信息等方向强匹配。", "H4，历史最高线高于等位分较多，只能作为冲刺观察。"),
    ("H89207", "普通公办低费用主线", "C-冲刺观察讨论", "进入家庭讨论", "电子信息/计算机口碑线索较强，计算机、软件、智能科学、网安方向贴合。", "H4 且最高差较高，冲刺属性强，不能占稳保位置。"),
    ("K15306", "普通公办低费用主线", "C-城市专业观察", "进入家庭讨论", "成都优先城市，数字媒体、软件、计算机、信息安全、人工智能等方向高度贴合。", "H0 缺同代码历史线索，且 22 专业大组内非偏好专业多，需优先补历史映射。"),
    ("K46904", "普通公办低费用主线", "C-城市专业观察", "进入家庭讨论", "西安优先城市，软件、智能科学、人工智能三项集中。", "H0 缺历史映射，需核是否新组、计划数和近年替代组。"),
    ("K48103", "普通公办低费用主线", "C-城市专业观察", "谨慎进入家庭讨论", "西安优先城市，软件、物联网、电子信息方向。", "H4 且结构边界需核，需确认没有串校或串组。"),
    ("F31303", "普通公办低费用主线", "D-谨慎核验后再讨论", "暂缓进入家庭讨论", "计算机、数据科学、物联网、智能科学、自动化、机器人方向密集，H2。", "结构边界异常，必须先重读原页。"),
    ("K58501", "普通公办低费用主线", "D-谨慎核验后再讨论", "暂缓进入家庭讨论", "网络、数据科学、物联网、集成电路方向可讨论，H2。", "OCR 串页或组边界风险，先核第 19 期原页。"),
    ("H13502", "普通公办低费用主线", "D-谨慎核验后再讨论", "暂缓进入家庭讨论", "计算机与师范方向都较多，H3。", "18 专业大组，特殊限制和语种限制待核，调剂风险较大。"),
    ("H45004", "普通公办低费用主线", "D-谨慎核验后再讨论", "暂缓进入家庭讨论", "含数字媒体、人工智能、计算机、数据科学、自动化、电子等方向，H3。", "结构边界异常，统计/安全/应急方向是否接受待判断。"),
    ("C15003", "普通公办低费用主线", "D-谨慎核验后再讨论", "暂缓进入家庭讨论", "武汉城市线索，含大数据管理与应用。", "22 专业大组仅 1 个偏好相关，英语/语种限制和调剂接受度压力大。"),
]

COOP_CANDIDATES = [
    ("K48703", "7万内中外合作/高收费专项", "A-专项优先核验", "进入专项讨论", "西安、公办线索、计算机中外合作方向贴合，历史线索在专项里相对可研究。", "高收费且计划数字段需核，校区、证书、调剂和培养模式必须闭环。"),
    ("K15308", "7万内中外合作/高收费专项", "A-专项优先核验", "进入专项讨论", "成都优先城市，软件工程中外合作方向强贴合，费用候选 55000。", "H0 无同代码历史，计划数小，非英语慎报等限制需核。"),
    ("C13909", "7万内中外合作/高收费专项", "A-专项优先核验", "进入专项讨论", "数字媒体技术中外合作，第一偏好专业，费用候选 25000，H2。", "学校层级一般，需核是否单专业组、培养模式和证书。"),
    ("A36603", "7万内中外合作/高收费专项", "A-专项优先核验", "进入专项讨论", "厦门大学马来西亚校区，计算机/软件/AI/电子信息/自动化方向集中，品牌价值高。", "海外校区、币种和生活成本、全英文、英语单科门槛都要逐条确认。"),
    ("A08521", "7万内中外合作/高收费专项", "A-专项优先核验", "进入专项讨论", "哈尔滨工程大学层次高，电子信息方向值得单列研究。", "H0，3+1、不得转专业、出国费用和真实录取概率需补证。"),
    ("C10212", "7万内中外合作/高收费专项", "B-高价值高冲观察", "谨慎进入专项讨论", "武汉科技大学，信息与计算科学、机器人工程方向强，费用候选卡 70000。", "H5 高冲，费用正好上限，组边界和调剂需先核。"),
    ("C10211", "7万内中外合作/高收费专项", "B-高价值高冲观察", "谨慎进入专项讨论", "武汉科技大学，网络工程方向贴合，学校和城市有吸引力。", "H5 且历史差距大，只能作为高冲观察，不占稳妥名额。"),
    ("F58906", "7万内中外合作/高收费专项", "B-专项可研究", "进入专项讨论", "吉林师范大学软件工程方向，费用候选 20000，H3。", "同组环境工程有体检限制线索，需判断调剂可接受性。"),
    ("F60017", "7万内中外合作/高收费专项", "B-专项可研究", "进入专项讨论", "长春大学电子信息工程方向，费用候选 26000，H2。", "计划数和合作院校、证书需核。"),
    ("H31310", "7万内中外合作/高收费专项", "B-专项可研究", "进入专项讨论", "南昌航空大学软件工程方向，费用候选 28000，工科背景可研究。", "H0，英文课程/双学位/历史替代组需核。"),
    ("F44904", "7万内中外合作/高收费专项", "B-专项可研究", "进入专项讨论", "山东交通学院计算机方向，费用候选 45000，专业贴合。", "H0，小计划风险和双证真实性需核。"),
    ("C13815", "7万内中外合作/高收费专项", "C-专项保守观察", "谨慎进入专项讨论", "湖北理工学院中外合作费用低，H2 接近。", "专业为机械，不是最高偏好，需确认是否愿意读。"),
    ("F58907", "7万内中外合作/高收费专项", "C-专项保守观察", "谨慎进入专项讨论", "吉林师范大学电子信息方向，费用可控，H3。", "存在串页/组边界风险，先确认学校与专业组归属。"),
    ("F60016", "7万内中外合作/高收费专项", "C-专项保守观察", "谨慎进入专项讨论", "长春大学自动化方向，费用候选 26000，H3。", "俄方合作，授课语言、4+0/出国安排和证书需核。"),
    ("K27302", "7万内中外合作/高收费专项", "C-专项保守观察", "谨慎进入专项讨论", "贵州商学院数据科学方向，H2，录取安全线索较好。", "费用 60000 但学校层级不占优，性价比需要家庭判断。"),
    ("H91203", "7万内中外合作/高收费专项", "C-专项保守观察", "谨慎进入专项讨论", "桂林旅游学院数据科学方向，H3 接近。", "有只录有专业志愿线索和串页风险，先核组边界。"),
    ("H89304", "7万内中外合作/高收费专项", "D-专项暂缓", "暂缓进入专项讨论", "桂林理工大学软件/网络方向贴合，费用在预算内。", "H5 高冲且 OCR 有串页风险，先不投入主核验。"),
    ("F27306", "7万内中外合作/高收费专项", "D-专项暂缓", "暂缓进入专项讨论", "厦门大学嘉庚学院计算机/软件方向和费用看似友好。", "民办独立学院，不能按厦门大学主校理解，除非保底不足再讨论。"),
    ("A14509", "7万内中外合作/高收费专项", "E-排除或仅留风险记录", "不进入家庭讨论", "中国地质大学（武汉）学校和城市有吸引力。", "费用线索超过 70000 或待核且 H5 明显高冲，当前口径不进入讨论。"),
    ("H31011", "7万内中外合作/高收费专项", "E-排除或仅留风险记录", "不进入家庭讨论", "南昌大学 211 和计算机/软件方向有吸引力。", "费用文本写 80000 待核且有体检限制线索，按超预算风险处理。"),
    ("A03208", "7万内中外合作/高收费专项", "E-排除或仅留风险记录", "不进入家庭讨论", "北京语言大学线索保留为负例。", "费用候选 100000，超过预算，且专业方向不构成优先收益。"),
    ("A03209", "7万内中外合作/高收费专项", "E-排除或仅留风险记录", "不进入家庭讨论", "北京语言大学线索保留为负例。", "费用候选 82500，超过预算，且涉及康复等与不学医底线冲突的风险。"),
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def stable_id(prefix, code):
    return f"{prefix}-{hashlib.sha1(code.encode('utf-8')).hexdigest()[:16]}"


def by_code(rows):
    result = {}
    for row in rows:
        code = row.get("院校专业组代码OCR规范化", "")
        if code and code not in result:
            result[code] = row
    return result


def candidate_source_row(code, track, round1_by_code, expanded_by_code, stable_by_code):
    if track == "普通公办低费用主线":
        return dict(round1_by_code.get(code) or stable_by_code.get(code) or {})
    return dict(expanded_by_code.get(code) or stable_by_code.get(code) or {})


def build_group_rows(stable_by_code, round1_by_code, expanded_by_code):
    rows = []
    seen = set()
    for code, track, tier, action, reason, risk in NORMAL_CANDIDATES + COOP_CANDIDATES:
        if code in seen:
            raise SystemExit(f"重复候选代码：{code}")
        seen.add(code)
        base = candidate_source_row(code, track, round1_by_code, expanded_by_code, stable_by_code)
        if not base:
            raise SystemExit(f"未在结构化底座中找到候选：{code}")
        out = dict(base)
        out["个性化候选ID"] = stable_id("PFITGROUP", code)
        out["候选主线"] = track
        out["本轮分层"] = tier
        out["家庭讨论动作"] = action
        out["是否进入本轮家庭讨论"] = (
            "false" if action.startswith("暂缓") or action.startswith("不进入") else "true"
        )
        out["是否可作为定稿依据"] = "false"
        out["适配理由"] = reason
        out["关键风险"] = risk
        out["下一步核验动作"] = (
            "回看第19期原页和完整专业组；核湖北官方系统或省招办计划；核高校招生章程、学费、"
            "培养模式、语种/单科/体检限制、校区、计划数和调剂范围。"
        )
        out["第19期页码"] = base.get("来源页码", "")
        out["来源表"] = "round1普通主线" if track == "普通公办低费用主线" else "expanded_budget_coop专项"
        rows.append(out)
    return rows


def major_role(row):
    text = "；".join(
        [
            row.get("专业名称及备注短摘", ""),
            row.get("专业偏好方向", ""),
            row.get("机器专业接受度初判", ""),
            row.get("专业风险类型", ""),
        ]
    )
    if any(token in text for token in ["临床", "护理", "医学", "检验", "影像", "药学", "康复", "口腔", "中医"]):
        return "组内风险-医学相关待确认"
    if row.get("专业偏好方向"):
        return "偏好或可研究专业"
    if any(token in text for token in ["中外合作", "高收费", "合作办学"]):
        return "费用或培养模式需核专业"
    if row.get("字段缺口数", "") not in {"", "0"}:
        return "字段缺口待核专业"
    return "调剂接受度待判断专业"


def build_major_rows(group_rows, stable_major_rows):
    group_by_id = {row.get("专业组出现ID", ""): row for row in group_rows}
    rows = []
    for row in stable_major_rows:
        meta = group_by_id.get(row.get("专业组出现ID", ""))
        if not meta:
            continue
        out = dict(row)
        out["个性化候选ID"] = meta.get("个性化候选ID", "")
        out["候选主线"] = meta.get("候选主线", "")
        out["本轮分层"] = meta.get("本轮分层", "")
        out["家庭讨论动作"] = meta.get("家庭讨论动作", "")
        out["是否进入本轮家庭讨论"] = meta.get("是否进入本轮家庭讨论", "")
        out["是否可作为定稿依据"] = "false"
        out["专业角色判断"] = major_role(out)
        out["第19期页码"] = row.get("来源页码", "")
        out["下一步核验动作"] = "整组逐专业核第19期原页、湖北官方侧、章程限制、学费、计划数和调剂接受度。"
        rows.append(out)
    return sorted(
        rows,
        key=lambda row: (
            row.get("候选主线", ""),
            row.get("本轮分层", ""),
            row.get("院校专业组代码OCR规范化", ""),
            int(row.get("专业组内专业序号", "0") or "0"),
        ),
    )


def append_sheet(ws, rows, fields):
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, start=1):
        max_len = min(
            52,
            max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:200]]),
        )
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(group_rows, major_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "00_说明"
    ws.append(["项目", "说明"])
    for item in [
        ("定位", "结合个人分数、位次、家庭偏好和第19期结构化底座生成的讨论池，不作定稿依据。"),
        ("先看", "先看 01_个人适配专业组，再看 02_完整组内专业明细。"),
        ("原则", "只要准备服从调剂，就必须看完整院校专业组，不只看拟填的6个专业。"),
        ("下一步", "对愿意讨论的组做原页、省招办计划、章程、学费、计划数、选科、备注限制和调剂范围闭环。"),
    ]:
        ws.append(item)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100
    append_sheet(wb.create_sheet("01_个人适配专业组"), group_rows, GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_完整组内专业明细"), major_rows, MAJOR_FIELDS)
    WORKBOOK_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_OUTPUT)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "/home/",
        "/var/folders/",
        "/private/",
        "private/",
        "private\\",
        "ocr-runs",
        "rendered-pages",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "人工读数",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"个人适配候选公开文件包含禁止内容：{hit}")


def main():
    stable_groups = read_csv(STABLE_GROUP_BROWSER)
    stable_majors = read_csv(STABLE_MAJOR_BROWSER)
    round1_groups = read_csv(ROUND1_GROUPS)
    expanded_groups = read_csv(EXPANDED_GROUPS)

    group_rows = build_group_rows(
        by_code(stable_groups),
        by_code(round1_groups),
        by_code(expanded_groups),
    )
    major_rows = build_major_rows(group_rows, stable_majors)

    write_csv(GROUP_OUTPUT, group_rows, GROUP_FIELDS)
    write_csv(MAJOR_OUTPUT, major_rows, MAJOR_FIELDS)
    write_workbook(group_rows, major_rows)

    summary = {
        "status": "issue19_personal_fit_v1_discussion_pool_ready",
        "generated_by": "build_issue19_personal_fit_v1.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "usage_boundary": "结合个人与家庭偏好的候选讨论池，不作定稿依据。",
        "fixed_inputs": {
            "省份": "湖北",
            "年份": 2026,
            "类别": "普通类首选物理",
            "再选科目": "化学、地理",
            "总分": 515,
            "累计位次": 91723,
            "等位分": {"2025": 494, "2024": 497, "2023": 481},
            "普通主线": "优先公办普通学费",
            "专项预算上限": 70000,
            "明确不接受": ["医学/护理默认不进主线"],
        },
        "group_output": str(GROUP_OUTPUT.relative_to(ROOT)),
        "major_output": str(MAJOR_OUTPUT.relative_to(ROOT)),
        "workbook": str(WORKBOOK_OUTPUT.relative_to(ROOT)),
        "group_count": len(group_rows),
        "major_detail_count": len(major_rows),
        "discussion_group_count": sum(row["是否进入本轮家庭讨论"] == "true" for row in group_rows),
        "not_enter_discussion_group_count": sum(row["是否进入本轮家庭讨论"] == "false" for row in group_rows),
        "track_counts": dict(Counter(row["候选主线"] for row in group_rows)),
        "tier_counts": dict(Counter(row["本轮分层"] for row in group_rows)),
        "action_counts": dict(Counter(row["家庭讨论动作"] for row in group_rows)),
        "all_rows_gate": {
            "是否可作为定稿依据": "false",
            "must_verify": [
                "第19期原页或纸质原页",
                "湖北官方系统或省招办计划",
                "高校招生章程",
                "完整专业组和调剂范围",
                "学费、计划数、选科、校区、语种/单科/体检限制",
            ],
        },
        "source_tables": [
            str(STABLE_GROUP_BROWSER.relative_to(ROOT)),
            str(STABLE_MAJOR_BROWSER.relative_to(ROOT)),
            str(ROUND1_GROUPS.relative_to(ROOT)),
            str(EXPANDED_GROUPS.relative_to(ROOT)),
        ],
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    public_text_safe([GROUP_OUTPUT, MAJOR_OUTPUT, SUMMARY_OUTPUT])

    print(f"写出个人适配专业组：{GROUP_OUTPUT.relative_to(ROOT)}")
    print(f"写出完整组内专业明细：{MAJOR_OUTPUT.relative_to(ROOT)}")
    print(f"写出个人适配工作簿：{WORKBOOK_OUTPUT.relative_to(ROOT)}")
    print(f"专业组={len(group_rows)}；逐专业明细={len(major_rows)}")


if __name__ == "__main__":
    main()
