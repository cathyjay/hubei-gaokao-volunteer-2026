#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"
WORKING = ROOT / "data/working"

ACTION_PACK = EXPORTS / "issue19-next-closure-family-review-v1-first-closure-action-pack.csv"
TASK_STATUS = EXPORTS / "issue19-next-closure-family-review-v1-first-closure-task-status.csv"
FIELD_LEDGER = WORKING / "issue19-stable-foundation-first-closure-field-confirmation-public-ledger.csv"
SCHOOL_OPP = WORKING / "issue19-school-source-opportunity-queue.csv"
SCHOOL_LIVE = WORKING / "issue19-school-source-live-20260629-ledger.csv"
GROUP_REVIEW = EXPORTS / "issue19-next-closure-family-review-v1-priority55-group-review.csv"
MAJOR_REVIEW = EXPORTS / "issue19-next-closure-family-review-v1-priority55-major-review.csv"

OUT_PREFIX = "issue19-data-foundation-next-execution-v1"
OUT_WORKBOOK = EXPORTS / f"{OUT_PREFIX}.xlsx"
OUT_P0_PACKAGES = EXPORTS / f"{OUT_PREFIX}-p0-conflict-packages.csv"
OUT_P0_TASKS = EXPORTS / f"{OUT_PREFIX}-p0-conflict-tasks.csv"
OUT_SCHOOL_NEXT20 = EXPORTS / f"{OUT_PREFIX}-school-source-next20.csv"
OUT_GROUP_RISK = EXPORTS / f"{OUT_PREFIX}-priority55-transfer-risk.csv"
OUT_SUMMARY = EXPORTS / f"{OUT_PREFIX}-summary.json"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".heic",
    "Authorization",
    "Bearer ",
    "Cookie",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
]

P0_PACKAGE_FIELDS = [
    "执行序号",
    "核验包编号",
    "页码版面",
    "来源页码",
    "版面列",
    "涉及院校",
    "涉及专业组",
    "涉及专业代号",
    "包内任务数",
    "涉及字段范围",
    "双人复核任务数",
    "高校辅证任务数",
    "人工看图任务数",
    "PDFOCR与高校辅证冲突任务数",
    "原页记录状态",
    "湖北官方记录状态",
    "高校辅证记录状态",
    "三方一致性状态",
    "字段写回评估状态",
    "公开冲突判断",
    "优先级理由",
    "自动可继续动作",
    "人工最小核验动作",
    "完成后可进入状态",
    "仍不可越过的门禁",
    "关联任务ID集合SHA256",
    "公开安全边界",
]

P0_TASK_FIELDS = [
    "执行序号",
    "核验包编号",
    "页码版面键",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业代号",
    "专业名称短摘",
    "字段范围",
    "任务来源类型",
    "公开冲突桶",
    "候选提示桶",
    "PDFOCR提示状态",
    "机器坐标提示状态",
    "高校辅证线索",
    "是否需要双人复核",
    "是否需要人工看图",
    "PDF原页记录状态",
    "湖北官方记录状态",
    "高校辅证记录状态",
    "三方一致性状态",
    "字段写回状态",
    "下一步核验动作",
    "完成条件",
    "专业行ID",
    "第一闭环任务ID",
]

SCHOOL_NEXT20_FIELDS = [
    "执行建议序号",
    "院校代码",
    "院校名称",
    "机会优先级",
    "机会类型",
    "自动收益分",
    "官网辅证自动动作",
    "高校侧刷新任务类型",
    "涉及招生明细数",
    "涉及专业组数",
    "计划数冲突行数",
    "官网补缺候选行数",
    "官网未匹配行数",
    "C4C6计划数冲突候选数",
    "C4C6官网可补OCR计划数候选数",
    "C4C6无结构化官网源明细数",
    "已有live补源状态",
    "已有live湖北物理计划状态",
    "已有live结构化输出",
    "来源质量判断",
    "自动化下一步",
    "人工最小核验动作",
    "保真边界",
    "PDF原页核页状态",
    "湖北官方核验状态",
    "高校官网源刷新状态",
    "字段事实写回状态",
    "是否可作为定稿依据",
]

GROUP_RISK_FIELDS = [
    "下一轮序号",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "城市",
    "冲稳保",
    "下一轮建议状态",
    "专业明细行数",
    "可接受专业数",
    "勉强接受专业数",
    "待核后判断专业数",
    "不能接受专业数",
    "特殊限制专业数",
    "高费或中外合作提示数",
    "字段缺口专业数",
    "100%核验专业数",
    "调剂风险等级",
    "家庭最小决策动作",
    "核验最小动作",
    "是否建议进入下一轮重点核验",
    "是否可作为定稿依据",
]


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, fields, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def as_int(value):
    try:
        return int(str(value).strip())
    except Exception:
        return 0


def yes(value):
    return str(value).strip().lower() == "true"


def digest(values):
    if isinstance(values, str):
        text = values
    else:
        text = "；".join(sorted({str(value) for value in values if value}))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def short_counter(values):
    counter = Counter(value or "空" for value in values)
    return "；".join(f"{key}×{value}" for key, value in counter.most_common())


def split_semicolon(value):
    if not value:
        return []
    return [part.strip() for part in str(value).replace("；", ";").split(";") if part.strip()]


def p0_priority_reason(package_row, tasks):
    reasons = []
    if as_int(package_row.get("双人复核任务数")) >= 3:
        reasons.append("双人复核任务集中")
    if as_int(package_row.get("包内任务数")) >= 5:
        reasons.append("同页列冲突任务多")
    if as_int(package_row.get("高校辅证需复核任务数")):
        reasons.append("高校辅证可用于解释差异")
    if len({row.get("院校代码", "") for row in tasks}) == 1:
        reasons.append("单校单页列适合先打样")
    return "；".join(reasons) or "P0冲突包按页列顺序推进"


def p0_auto_action(tasks):
    schools = sorted({row.get("院校名称", "") for row in tasks if row.get("院校名称", "")})
    if len(schools) == 1:
        return f"优先复用或补找{schools[0]}官网2026湖北物理计划、章程或计划图，仅作差异解释。"
    return "按包内院校分别补高校官网辅证，先形成差异解释，不写回字段。"


def p0_manual_action(package_row):
    return (
        f"回看第19期PDF第{package_row.get('来源页码')}页{package_row.get('版面列')}栏；"
        "记录专业代号、专业名称、计划数、学费、选科和专业组边界；冲突字段做第二人复核；再查湖北官方侧。"
    )


def build_p0_outputs(action_rows, task_rows, field_rows):
    field_by_detail = {row.get("稳定基座第一闭环明细任务ID", ""): row for row in field_rows}
    package_by_detail = {}
    p0_packages = [row for row in action_rows if row.get("包类型") == "P0候选冲突"]
    for package in p0_packages:
        for detail_id in split_semicolon(package.get("关联明细任务编号集合", "")):
            package_by_detail[detail_id] = package

    p0_task_rows = []
    tasks_by_package = defaultdict(list)
    for task in task_rows:
        detail_id = task.get("第一闭环任务ID", "")
        package = package_by_detail.get(detail_id)
        if not package:
            continue
        field = field_by_detail.get(detail_id, {})
        output = {
            "执行序号": package.get("执行顺序", ""),
            "核验包编号": package.get("核验包编号", ""),
            "页码版面键": task.get("页码版面键", ""),
            "院校代码": task.get("院校代码", ""),
            "院校名称": task.get("院校名称", ""),
            "院校专业组代码": task.get("院校专业组代码", ""),
            "专业代号": task.get("专业代号", ""),
            "专业名称短摘": task.get("专业名称短摘", ""),
            "字段范围": task.get("字段范围", ""),
            "任务来源类型": task.get("任务来源类型", ""),
            "公开冲突桶": task.get("公开冲突桶", ""),
            "候选提示桶": task.get("候选提示桶", ""),
            "PDFOCR提示状态": task.get("PDFOCR提示", ""),
            "机器坐标提示状态": task.get("机器坐标提示", ""),
            "高校辅证线索": task.get("高校辅证线索", ""),
            "是否需要双人复核": task.get("是否需要双人复核", ""),
            "是否需要人工看图": field.get("是否需要人工直接看图", ""),
            "PDF原页记录状态": task.get("PDF原页记录状态", ""),
            "湖北官方记录状态": task.get("湖北官方记录状态", ""),
            "高校辅证记录状态": task.get("高校辅证记录状态", ""),
            "三方一致性状态": field.get("三方字段一致性公开状态", ""),
            "字段写回状态": task.get("字段写回状态", ""),
            "下一步核验动作": task.get("人工动作", ""),
            "完成条件": "PDF原页记录、湖北官方侧记录、高校辅证记录和必要双人复核齐全后，才进入字段写回评估；仍不作为最终志愿依据。",
            "专业行ID": task.get("专业行ID", ""),
            "第一闭环任务ID": detail_id,
        }
        p0_task_rows.append(output)
        tasks_by_package[package.get("核验包编号", "")].append(output)

    p0_package_rows = []
    for package in p0_packages:
        tasks = tasks_by_package.get(package.get("核验包编号", ""), [])
        p0_package_rows.append(
            {
                "执行序号": package.get("执行顺序", ""),
                "核验包编号": package.get("核验包编号", ""),
                "页码版面": package.get("页码版面", ""),
                "来源页码": package.get("来源页码", ""),
                "版面列": package.get("版面列", ""),
                "涉及院校": "；".join(sorted({row.get("院校名称", "") for row in tasks if row.get("院校名称", "")})),
                "涉及专业组": "；".join(sorted({row.get("院校专业组代码", "") for row in tasks if row.get("院校专业组代码", "")})),
                "涉及专业代号": "；".join(sorted({row.get("专业代号", "") for row in tasks if row.get("专业代号", "")})),
                "包内任务数": str(len(tasks)),
                "涉及字段范围": short_counter(row.get("字段范围", "") for row in tasks),
                "双人复核任务数": str(sum(yes(row.get("是否需要双人复核", "")) for row in tasks)),
                "高校辅证任务数": str(sum(yes(row.get("高校辅证线索", "")) for row in tasks)),
                "人工看图任务数": str(sum(yes(row.get("是否需要人工看图", "")) for row in tasks)),
                "PDFOCR与高校辅证冲突任务数": str(len(tasks)),
                "原页记录状态": short_counter(row.get("PDF原页记录状态", "") for row in tasks),
                "湖北官方记录状态": short_counter(row.get("湖北官方记录状态", "") for row in tasks),
                "高校辅证记录状态": short_counter(row.get("高校辅证记录状态", "") for row in tasks),
                "三方一致性状态": short_counter(row.get("三方一致性状态", "") for row in tasks),
                "字段写回评估状态": short_counter(row.get("字段写回状态", "") for row in tasks),
                "公开冲突判断": "PDF OCR 与高校辅证存在公开状态层冲突；公开表不保存候选读数，必须回原页和官方侧。",
                "优先级理由": p0_priority_reason(package, tasks),
                "自动可继续动作": p0_auto_action(tasks),
                "人工最小核验动作": p0_manual_action(package),
                "完成后可进入状态": "完成三方记录和双人复核后，只能进入字段写回评估队列。",
                "仍不可越过的门禁": "不得用高校官网替代湖北官方计划；不得跳过第19期原页；不得在未核专业组边界前生成志愿建议。",
                "关联任务ID集合SHA256": digest(row.get("第一闭环任务ID", "") for row in tasks),
                "公开安全边界": "公开层只保存状态、计数、学校/组索引和任务ID哈希；字段读数、截图、OCR原文和人工备注留在私有工作台。",
            }
        )
    p0_package_rows.sort(key=lambda row: as_int(row.get("执行序号")))
    p0_task_rows.sort(key=lambda row: (as_int(row.get("执行序号")), row.get("页码版面键", ""), row.get("专业代号", "")))
    return p0_package_rows, p0_task_rows


def build_school_next20(opportunity_rows, live_rows):
    live_by_school = {row.get("院校代码", ""): row for row in live_rows}
    selected = sorted(
        opportunity_rows,
        key=lambda row: (
            {"P0-立即处理": 0, "P1-高收益自动补源": 1, "P2-常规自动补源或抽检": 2, "P3-低收益留存": 3}.get(row.get("机会优先级", ""), 9),
            -as_int(row.get("自动收益分", "")),
            as_int(row.get("执行建议序号", "")),
        ),
    )[:20]
    rows = []
    for row in selected:
        live = live_by_school.get(row.get("院校代码", ""), {})
        rows.append(
            {
                "执行建议序号": row.get("执行建议序号", ""),
                "院校代码": row.get("院校代码", ""),
                "院校名称": row.get("院校名称OCR", ""),
                "机会优先级": row.get("机会优先级", ""),
                "机会类型": row.get("机会类型", ""),
                "自动收益分": row.get("自动收益分", ""),
                "官网辅证自动动作": row.get("官网辅证自动动作", ""),
                "高校侧刷新任务类型": row.get("高校侧刷新任务类型", ""),
                "涉及招生明细数": row.get("涉及招生明细数", ""),
                "涉及专业组数": row.get("涉及专业组数", ""),
                "计划数冲突行数": row.get("计划数冲突行数", ""),
                "官网补缺候选行数": row.get("官网补缺候选行数", ""),
                "官网未匹配行数": row.get("官网未匹配行数", ""),
                "C4C6计划数冲突候选数": row.get("C4C6计划数冲突候选数", ""),
                "C4C6官网可补OCR计划数候选数": row.get("C4C6官网可补OCR计划数候选数", ""),
                "C4C6无结构化官网源明细数": row.get("C4C6无结构化官网源明细数", ""),
                "已有live补源状态": live.get("自动补源结论", "未进入2026-06-29 live补源样本"),
                "已有live湖北物理计划状态": live.get("湖北物理计划状态", ""),
                "已有live结构化输出": "有结构化输出" if live.get("结构化输出", "") else "无结构化输出",
                "来源质量判断": row.get("来源质量判断", ""),
                "自动化下一步": row.get("自动化下一步", ""),
                "人工最小核验动作": row.get("人工最小核验动作", ""),
                "保真边界": row.get("保真边界", ""),
                "PDF原页核页状态": row.get("PDF原页核页状态", ""),
                "湖北官方核验状态": row.get("湖北官方系统或省招办计划核验状态", ""),
                "高校官网源刷新状态": row.get("高校官网源刷新状态", ""),
                "字段事实写回状态": row.get("字段事实写回状态", ""),
                "是否可作为定稿依据": "false",
            }
        )
    return rows


def transfer_risk_level(group, majors):
    unacceptable = as_int(group.get("不能接受提示专业数", ""))
    pending = as_int(group.get("待核后判断专业数", ""))
    barely = as_int(group.get("可调剂了解专业数", ""))
    special = as_int(group.get("特殊限制专业数", ""))
    fee = as_int(group.get("高费或中外合作提示数", ""))
    field_gap = as_int(group.get("字段缺口专业数", ""))
    if unacceptable:
        return "T3-存在不能接受专业需谨慎调剂"
    if special or fee:
        return "T2-先核限制费用再谈调剂"
    if pending >= 2 or field_gap >= max(1, len(majors) // 2):
        return "T2-字段或限制待核较多"
    if barely > as_int(group.get("优先了解专业数", "")):
        return "T1-可调剂但需家庭逐专业确认"
    return "T0-当前机器口径调剂阻力较低"


def build_group_risk(group_rows, major_rows):
    majors_by_group = defaultdict(list)
    for row in major_rows:
        key = row.get("第四轮候选ID", "") or f"{row.get('院校代码')}|{row.get('院校专业组代码')}"
        majors_by_group[key].append(row)

    rows = []
    for group in group_rows:
        key = group.get("第四轮候选ID", "") or f"{group.get('院校代码')}|{group.get('院校专业组代码')}"
        majors = majors_by_group.get(key, [])
        risk = transfer_risk_level(group, majors)
        enter_next = "true" if group.get("下一轮建议状态") in {"优先家庭讨论", "可讨论但先看调剂"} and not risk.startswith("T3") else "false"
        rows.append(
            {
                "下一轮序号": group.get("下一轮序号", ""),
                "院校代码": group.get("院校代码", ""),
                "院校名称": group.get("院校名称", ""),
                "院校专业组代码": group.get("院校专业组代码", ""),
                "城市": group.get("城市", ""),
                "冲稳保": group.get("冲稳保", ""),
                "下一轮建议状态": group.get("下一轮建议状态", ""),
                "专业明细行数": group.get("专业明细行数", ""),
                "可接受专业数": group.get("优先了解专业数", ""),
                "勉强接受专业数": group.get("可调剂了解专业数", ""),
                "待核后判断专业数": group.get("待核后判断专业数", ""),
                "不能接受专业数": group.get("不能接受提示专业数", ""),
                "特殊限制专业数": group.get("特殊限制专业数", ""),
                "高费或中外合作提示数": group.get("高费或中外合作提示数", ""),
                "字段缺口专业数": group.get("字段缺口专业数", ""),
                "100%核验专业数": group.get("100%核验专业数", ""),
                "调剂风险等级": risk,
                "家庭最小决策动作": group.get("家庭先决策问题", ""),
                "核验最小动作": group.get("组级核验动作", ""),
                "是否建议进入下一轮重点核验": enter_next,
                "是否可作为定稿依据": "false",
            }
        )
    rows.sort(key=lambda row: (row["是否建议进入下一轮重点核验"] != "true", as_int(row["下一轮序号"])))
    return rows


def add_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(len(header) + 4, 12), 44)


def write_workbook(summary, p0_packages, p0_tasks, school_next20, group_risk):
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = []
    for section, value in summary.items():
        if isinstance(value, dict):
            for key, subvalue in value.items():
                if isinstance(subvalue, (dict, list)):
                    subvalue = json.dumps(subvalue, ensure_ascii=False)
                summary_rows.append({"板块": section, "指标": key, "值": subvalue})
        else:
            summary_rows.append({"板块": "summary", "指标": section, "值": value})
    add_sheet(wb, "00_摘要", summary_rows, ["板块", "指标", "值"])
    add_sheet(wb, "01_P0冲突包", p0_packages, P0_PACKAGE_FIELDS)
    add_sheet(wb, "02_P0冲突逐任务", p0_tasks, P0_TASK_FIELDS)
    add_sheet(wb, "03_官网辅证next20", school_next20, SCHOOL_NEXT20_FIELDS)
    add_sheet(wb, "04_55组调剂风险", group_risk, GROUP_RISK_FIELDS)
    wb.save(OUT_WORKBOOK)


def assert_public_safe(paths):
    for path in paths:
        if path.suffix == ".xlsx":
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        hits = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]
        if hits:
            raise SystemExit(f"{path} contains forbidden public tokens: {hits}")


def main():
    action_rows = read_csv(ACTION_PACK)
    task_rows = read_csv(TASK_STATUS)
    field_rows = read_csv(FIELD_LEDGER)
    opportunity_rows = read_csv(SCHOOL_OPP)
    live_rows = read_csv(SCHOOL_LIVE)
    group_rows = read_csv(GROUP_REVIEW)
    major_rows = read_csv(MAJOR_REVIEW)

    p0_packages, p0_tasks = build_p0_outputs(action_rows, task_rows, field_rows)
    school_next20 = build_school_next20(opportunity_rows, live_rows)
    group_risk = build_group_risk(group_rows, major_rows)

    write_csv(OUT_P0_PACKAGES, P0_PACKAGE_FIELDS, p0_packages)
    write_csv(OUT_P0_TASKS, P0_TASK_FIELDS, p0_tasks)
    write_csv(OUT_SCHOOL_NEXT20, SCHOOL_NEXT20_FIELDS, school_next20)
    write_csv(OUT_GROUP_RISK, GROUP_RISK_FIELDS, group_risk)

    summary = {
        "status": "issue19_data_foundation_next_execution_v1_ready_not_final",
        "generated_by": Path(__file__).name,
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "usage_boundary": "用于推进第一闭环P0冲突包、高校官网辅证next20和55组调剂风险复核；不确认字段事实，不作为最终志愿方案。",
        "outputs": {
            "workbook": str(OUT_WORKBOOK.relative_to(ROOT)),
            "p0_conflict_packages": str(OUT_P0_PACKAGES.relative_to(ROOT)),
            "p0_conflict_tasks": str(OUT_P0_TASKS.relative_to(ROOT)),
            "school_source_next20": str(OUT_SCHOOL_NEXT20.relative_to(ROOT)),
            "priority55_transfer_risk": str(OUT_GROUP_RISK.relative_to(ROOT)),
        },
        "p0_conflict": {
            "package_count": len(p0_packages),
            "task_count": len(p0_tasks),
            "double_review_task_count": sum(yes(row.get("是否需要双人复核", "")) for row in p0_tasks),
            "school_evidence_task_count": sum(yes(row.get("高校辅证线索", "")) for row in p0_tasks),
            "pdf_reading_completed_count": 0,
            "hubei_official_completed_count": 0,
            "field_writeback_ready_count": 0,
        },
        "school_source_next20": {
            "row_count": len(school_next20),
            "unique_school_count": len({row.get("院校代码", "") for row in school_next20 if row.get("院校代码", "")}),
            "opportunity_type_counts": dict(Counter(row.get("机会类型", "") for row in school_next20)),
            "live_structured_output_count": sum(row.get("已有live结构化输出") == "有结构化输出" for row in school_next20),
            "recommendation_basis_allowed_count": 0,
        },
        "priority55_transfer_risk": {
            "group_count": len(group_risk),
            "enter_next_review_count": sum(row.get("是否建议进入下一轮重点核验") == "true" for row in group_risk),
            "risk_counts": dict(Counter(row.get("调剂风险等级", "") for row in group_risk)),
            "final_basis_allowed_count": 0,
        },
        "hashes": {
            "p0_packages_sha16": digest(json.dumps(p0_packages, ensure_ascii=False, sort_keys=True))[:16],
            "p0_tasks_sha16": digest(json.dumps(p0_tasks, ensure_ascii=False, sort_keys=True))[:16],
            "school_next20_sha16": digest(json.dumps(school_next20, ensure_ascii=False, sort_keys=True))[:16],
            "group_risk_sha16": digest(json.dumps(group_risk, ensure_ascii=False, sort_keys=True))[:16],
        },
    }
    OUT_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(summary, p0_packages, p0_tasks, school_next20, group_risk)
    assert_public_safe([OUT_P0_PACKAGES, OUT_P0_TASKS, OUT_SCHOOL_NEXT20, OUT_GROUP_RISK, OUT_SUMMARY])
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
