#!/usr/bin/env python3
import csv
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from build_issue19_round2_updated_preferences import (
    GROUP_BROWSER,
    MAJOR_BROWSER,
    ROOT,
    SOURCE_PDF_SHA256,
    aggregate_major_signals,
    agriculture_without_animal_count,
    append_sheet,
    build_major_rows as build_round2_major_rows,
    count,
    direction_summary,
    direction_tags,
    float_value,
    has_clinical_pause,
    has_health_or_vet_special,
    has_medtech_rehab_special,
    has_nursing_or_animal_block,
    history_band,
    history_code,
    int_value,
    main_ok,
    major_role,
    merge_signals,
    public_normal,
    public_value,
    read_csv,
    risk_notes,
    specific_watch_tags,
    stable_hash,
)


EXPORTS = ROOT / "data/exports"

ROUND3_GROUPS = EXPORTS / "issue19-round3-unrestricted-region-candidate-groups.csv"
ROUND3_MAIN_SHORTLIST = EXPORTS / "issue19-round3-unrestricted-region-main-shortlist-groups.csv"
ROUND3_DISCUSSION_PRIORITY = EXPORTS / "issue19-round3-unrestricted-region-discussion-priority-groups.csv"
ROUND3_SPECIAL_GROUPS = EXPORTS / "issue19-round3-unrestricted-region-special-low-priority-groups.csv"
ROUND3_MAIN_MAJORS = EXPORTS / "issue19-round3-unrestricted-region-main-shortlist-majors.csv"
ROUND3_SPECIAL_MAJORS = EXPORTS / "issue19-round3-unrestricted-region-special-majors.csv"
ROUND3_CITY_DISTRIBUTION = EXPORTS / "issue19-round3-unrestricted-region-city-distribution.csv"
ROUND3_WORKBOOK = EXPORTS / "issue19-round3-unrestricted-region.xlsx"
ROUND3_SUMMARY = EXPORTS / "issue19-round3-unrestricted-region-summary.json"

CITY_POLICY = "城市仅展示不参与筛选排序"

ROUND3_GROUP_FIELDS = [
    "第三轮候选ID",
    "第三轮池",
    "第三轮层级",
    "第三轮建议梯度",
    "是否进入家庭讨论",
    "是否可作为定稿依据",
    "不限地区排序分",
    "城市筛选口径",
    "入选理由",
    "专项风险提示",
    "下一步核验动作",
    "体检和章程核验提示",
    "重点观察标签",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "机械自动化机器人专业数",
    "电子信息网络专业数",
    "计算机AI软件专业数",
    "农业动医兽医专业数",
    "动物医学兽医专业数",
    "护理助产专业数",
    "工商旅游管理专业数",
    "环境工程科学专业数",
    "医技护理康复专业数",
    "临床口腔中医暂缓专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "机器筛选价值层级",
    "保真核验层级",
    "历史线索分层",
    "同代码命中年份最大值",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "组内完整专业清单索引",
    "机器核验下一步",
    "人工核验下一步",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

ROUND3_MAJOR_FIELDS = [
    "第三轮候选ID",
    "第三轮池",
    "第三轮层级",
    "第三轮建议梯度",
    "专业角色判断",
    "第三轮方向标签",
    "是否可作为定稿依据",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
]

CITY_DISTRIBUTION_FIELDS = [
    "城市候选",
    "完整候选池专业组数",
    "主线候选池专业组数",
    "主线精选专业组数",
    "优先讨论专业组数",
    "低优先级专项专业组数",
]


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: public_value(row.get(field, "")) for field in fields} for row in rows])


def score_unrestricted(row, special=False):
    value = {
        "H1": 44,
        "H2": 42,
        "H3": 30,
        "H4": 15,
        "H0": 9,
        "H5": -25,
    }.get(history_code(row), 0)
    value += count(row, "数字媒体技术专业数") * 14
    value += count(row, "计算机类相关专业数") * 6
    value += count(row, "计算机AI软件专业数") * 6
    value += count(row, "电子信息网络专业数") * 5
    value += count(row, "机械自动化机器人专业数") * 5
    value += count(row, "师范类相关专业数") * 4
    value += count(row, "环境工程科学专业数") * 3
    value += count(row, "工商旅游管理专业数") * 2
    value += agriculture_without_animal_count(row) * (3 if special else 2)
    value += count(row, "偏好专业数") * 2
    value += count(row, "医技护理康复专业数") * (4 if special else -10)
    value -= count(row, "动物医学兽医专业数") * 60
    value -= count(row, "护理助产专业数") * 60
    if specific_watch_tags(row):
        value += 6
    if row.get("调剂初判", "") == "可进入人工调剂接受度判断":
        value += 8
    if count(row, "特殊限制待核专业数") > 0:
        value -= 8
    if "结构异常" in row.get("调剂初判", ""):
        value -= 10
    if has_clinical_pause(row):
        value -= 40
    value -= max(float_value(row, "历史最高等位分差", 0), 0) / 5
    return round(value, 3)


def sorted_unrestricted(rows, special=False):
    return sorted(
        rows,
        key=lambda row: (
            -score_unrestricted(row, special=special),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
            row.get("专业组出现ID", ""),
        ),
    )


def take_unique_unrestricted(source, limit, selected_ids, special=False):
    picked = []
    for row in sorted_unrestricted(source, special=special):
        group_id = row.get("专业组出现ID", "")
        if group_id in selected_ids:
            continue
        picked.append(row)
        selected_ids.add(group_id)
        if len(picked) >= limit:
            break
    return picked


def special_low_priority_ok(row, main_group_ids):
    return (
        row.get("专业组出现ID", "") not in main_group_ids
        and public_normal(row)
        and count(row, "高收费或超预算专业数") == 0
        and count(row, "专业明细行数") > 0
        and not has_clinical_pause(row)
        and not has_nursing_or_animal_block(row)
        and (
            has_medtech_rehab_special(row)
            or count(row, "环境工程科学专业数") > 0
            or agriculture_without_animal_count(row) > 0
            or specific_watch_tags(row)
        )
    )


def selection_reason(row, pool):
    reasons = [
        f"命中方向：{direction_summary(row)}",
        f"历史线索：{row.get('历史线索分层')}",
        f"城市口径：{CITY_POLICY}",
        "家庭口径：优先公办本科、普通学费，年学费上限15000；护理/助产、动物医学/兽医/动物科学暂不纳入。",
    ]
    tags = specific_watch_tags(row)
    if tags:
        reasons.append("点名观察：" + "、".join(tags))
    if pool != "主线":
        reasons.append("低优先级专项只用于了解，不和普通主线混排。")
    reasons.append("仍需回看第19期原页、湖北官方侧和招生章程。")
    return "；".join(reasons)


def body_exam_note(row):
    if has_nursing_or_animal_block(row):
        return "护理/助产、动物医学/兽医/动物科学本轮暂不纳入；若未来重启，需先核职业接受度、动物接触或夜班实习、职业暴露、体检要求和章程。"
    if has_health_or_vet_special(row) or has_clinical_pause(row):
        return "体检摘要显示色觉正常、矫正视力4.80/4.80；医技、康复、农学等仍必须逐校核招生章程和体检指导意见。"
    if count(row, "特殊限制待核专业数") > 0:
        return "体检或特殊限制字段待核；需看招生章程、专业备注和第19期原页。"
    return "无新增体检自动阻断线索；仍需按最终入围专业逐项核章程。"


def decorate(row, layer, pool, enter_discussion=True, special=False):
    out = dict(row)
    out["第三轮候选ID"] = f"R3GROUP-{stable_hash('issue19-round3-unrestricted', row)}"
    out["第三轮池"] = pool
    out["第三轮层级"] = layer
    out["第三轮建议梯度"] = history_band(row)
    out["是否进入家庭讨论"] = "true" if enter_discussion else "false"
    out["是否可作为定稿依据"] = "false"
    out["不限地区排序分"] = str(score_unrestricted(row, special=special))
    out["城市筛选口径"] = CITY_POLICY
    out["入选理由"] = selection_reason(row, pool)
    out["专项风险提示"] = risk_notes(row, special=special)
    out["下一步核验动作"] = "回看第19期原页和完整专业组；核湖北官方系统/省招办计划；核高校招生章程、学费、计划数、选科、校区、语种/单科/体检限制和调剂范围。"
    out["体检和章程核验提示"] = body_exam_note(row)
    out["重点观察标签"] = "；".join(specific_watch_tags(row))
    return out


def build_main_shortlist(main_pool):
    selected_ids = set()
    shortlist = []
    quota_plan = [
        ("U1-保底稳妥不限地区", [row for row in main_pool if history_code(row) in {"H1", "H2"}], 50),
        ("U2-稳冲不限地区", [row for row in main_pool if history_code(row) == "H3"], 35),
        ("U3-冲刺不限地区", [row for row in main_pool if history_code(row) == "H4"], 25),
        (
            "U4-历史待补或点名观察不限地区",
            [row for row in main_pool if history_code(row) == "H0" or specific_watch_tags(row)],
            10,
        ),
    ]
    for layer, rows, limit in quota_plan:
        shortlist.extend(
            decorate(row, layer, "主线", special=False)
            for row in take_unique_unrestricted(rows, limit, selected_ids, special=False)
        )
    return shortlist


def build_discussion_priority(main_shortlist):
    selected_ids = set()
    rows = []
    quota_plan = [
        ("D1-先讨论保底稳妥", [row for row in main_shortlist if history_code(row) in {"H1", "H2"}], 24),
        ("D2-先讨论稳冲", [row for row in main_shortlist if history_code(row) == "H3"], 22),
        ("D3-先讨论冲刺", [row for row in main_shortlist if history_code(row) == "H4"], 10),
        (
            "D4-历史待补或点名待核",
            [row for row in main_shortlist if history_code(row) == "H0" or row.get("重点观察标签", "")],
            4,
        ),
    ]
    for layer, source, limit in quota_plan:
        for row in take_unique_unrestricted(source, limit, selected_ids, special=False):
            out = dict(row)
            out["第三轮层级"] = layer
            rows.append(out)
    return rows


def build_special_groups(special_pool):
    selected_ids = set()
    rows = []
    quota_plan = [
        (
            "T1-医技康复低优先级了解",
            [row for row in special_pool if has_medtech_rehab_special(row)],
            24,
        ),
        (
            "T2-农业环境低优先级了解",
            [
                row
                for row in special_pool
                if agriculture_without_animal_count(row) > 0
                or count(row, "环境工程科学专业数") > 0
            ],
            30,
        ),
        (
            "T3-点名学校低优先级观察",
            [row for row in special_pool if specific_watch_tags(row)],
            6,
        ),
    ]
    for layer, source, limit in quota_plan:
        rows.extend(
            decorate(row, layer, "低优先级专项", special=True)
            for row in take_unique_unrestricted(source, limit, selected_ids, special=True)
        )
    return rows


def build_all_candidates(main_pool, special_pool):
    rows = []
    seen = set()
    for pool_name, layer, source, special in [
        ("主线", "C1-不限地区主线完整候选池", main_pool, False),
        ("低优先级专项", "C2-不限地区低优先级专项候选池", special_pool, True),
    ]:
        for row in sorted_unrestricted(source, special=special):
            group_id = row.get("专业组出现ID", "")
            if group_id in seen:
                continue
            rows.append(decorate(row, layer, pool_name, enter_discussion=False, special=special))
            seen.add(group_id)
    return rows


def build_major_rows(group_rows, major_rows):
    round2_style_groups = []
    for row in group_rows:
        out = dict(row)
        out["第二轮候选ID"] = row.get("第三轮候选ID", "")
        out["第二轮池"] = row.get("第三轮池", "")
        out["第二轮层级"] = row.get("第三轮层级", "")
        out["建议梯度"] = row.get("第三轮建议梯度", "")
        round2_style_groups.append(out)
    round2_major_rows = build_round2_major_rows(round2_style_groups, major_rows)
    rows = []
    for row in round2_major_rows:
        out = dict(row)
        out["第三轮候选ID"] = out.pop("第二轮候选ID", "")
        out["第三轮池"] = out.pop("第二轮池", "")
        out["第三轮层级"] = out.pop("第二轮层级", "")
        out["第三轮建议梯度"] = out.pop("建议梯度", "")
        out["第三轮方向标签"] = "；".join(direction_tags(out))
        out["专业角色判断"] = major_role(out)
        out["是否可作为定稿依据"] = "false"
        rows.append(out)
    return rows


def build_city_distribution(all_candidates, main_pool, main_shortlist, discussion_priority, special_groups):
    counters = {
        "完整候选池专业组数": Counter(row.get("城市候选", "") or "未识别" for row in all_candidates),
        "主线候选池专业组数": Counter(row.get("城市候选", "") or "未识别" for row in main_pool),
        "主线精选专业组数": Counter(row.get("城市候选", "") or "未识别" for row in main_shortlist),
        "优先讨论专业组数": Counter(row.get("城市候选", "") or "未识别" for row in discussion_priority),
        "低优先级专项专业组数": Counter(row.get("城市候选", "") or "未识别" for row in special_groups),
    }
    cities = sorted({city for counter in counters.values() for city in counter})
    return [
        {
            "城市候选": city,
            "完整候选池专业组数": counters["完整候选池专业组数"].get(city, 0),
            "主线候选池专业组数": counters["主线候选池专业组数"].get(city, 0),
            "主线精选专业组数": counters["主线精选专业组数"].get(city, 0),
            "优先讨论专业组数": counters["优先讨论专业组数"].get(city, 0),
            "低优先级专项专业组数": counters["低优先级专项专业组数"].get(city, 0),
        }
        for city in cities
    ]


def write_workbook(all_candidates, main_shortlist, discussion_priority, special_groups, main_majors, special_majors, city_distribution):
    wb = Workbook()
    wb.remove(wb.active)
    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "第三轮不限地区候选池只用于讨论和核验，不作为定稿依据。"),
        ("城市口径", CITY_POLICY + "；本轮不因为武汉、成都、西安、北京等城市加分，也不设置优先城市名额。"),
        ("家庭口径", "优先公办本科、普通学费，年学费上限15000；民办、独立学院、职业本科、中外合作或高收费不进普通主线。"),
        ("专业底线", "护理/助产、动物医学/兽医/动物科学暂不纳入；临床/口腔/中医等医学主线暂缓；医技/康复只作低优先级专项了解。"),
        ("下一步", "先看优先讨论60组，再看主线120组；任何保留组都必须核第19期原页、湖北官方侧、招生章程和完整组内调剂。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 118

    append_sheet(wb.create_sheet("01_不限地区主线120组"), main_shortlist, ROUND3_GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_优先讨论60组"), discussion_priority, ROUND3_GROUP_FIELDS)
    append_sheet(wb.create_sheet("03_低优先级专项"), special_groups, ROUND3_GROUP_FIELDS)
    append_sheet(wb.create_sheet("04_主线组内专业明细"), main_majors, ROUND3_MAJOR_FIELDS)
    append_sheet(wb.create_sheet("05_专项组内专业明细"), special_majors, ROUND3_MAJOR_FIELDS)
    append_sheet(wb.create_sheet("06_城市分布"), city_distribution, CITY_DISTRIBUTION_FIELDS)
    append_sheet(wb.create_sheet("07_完整候选池"), all_candidates, ROUND3_GROUP_FIELDS)
    ROUND3_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ROUND3_WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "/home/",
        "/var/folders/",
        "/private/",
        "private/",
        "private\\",
        "file://",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"第三轮不限地区候选公开文件包含禁止内容：{hit}")


def main():
    group_rows = read_csv(GROUP_BROWSER)
    major_rows = read_csv(MAJOR_BROWSER)
    major_signals, _ = aggregate_major_signals(major_rows)
    enriched_groups = merge_signals(group_rows, major_signals)

    main_pool = [row for row in enriched_groups if main_ok(row)]
    main_group_ids = {row.get("专业组出现ID", "") for row in main_pool}
    special_pool = [row for row in enriched_groups if special_low_priority_ok(row, main_group_ids)]

    main_shortlist = build_main_shortlist(main_pool)
    discussion_priority = build_discussion_priority(main_shortlist)
    special_groups = build_special_groups(special_pool)
    all_candidates = build_all_candidates(main_pool, special_pool)
    main_major_rows = build_major_rows(main_shortlist, major_rows)
    special_major_rows = build_major_rows(special_groups, major_rows)
    city_distribution = build_city_distribution(all_candidates, main_pool, main_shortlist, discussion_priority, special_groups)

    write_csv(ROUND3_GROUPS, all_candidates, ROUND3_GROUP_FIELDS)
    write_csv(ROUND3_MAIN_SHORTLIST, main_shortlist, ROUND3_GROUP_FIELDS)
    write_csv(ROUND3_DISCUSSION_PRIORITY, discussion_priority, ROUND3_GROUP_FIELDS)
    write_csv(ROUND3_SPECIAL_GROUPS, special_groups, ROUND3_GROUP_FIELDS)
    write_csv(ROUND3_MAIN_MAJORS, main_major_rows, ROUND3_MAJOR_FIELDS)
    write_csv(ROUND3_SPECIAL_MAJORS, special_major_rows, ROUND3_MAJOR_FIELDS)
    write_csv(ROUND3_CITY_DISTRIBUTION, city_distribution, CITY_DISTRIBUTION_FIELDS)
    write_workbook(
        all_candidates,
        main_shortlist,
        discussion_priority,
        special_groups,
        main_major_rows,
        special_major_rows,
        city_distribution,
    )

    direction_group_counts = {}
    direction_major_counts = {}
    for field in [
        "机械自动化机器人专业数",
        "电子信息网络专业数",
        "计算机AI软件专业数",
        "农业动医兽医专业数",
        "动物医学兽医专业数",
        "护理助产专业数",
        "工商旅游管理专业数",
        "环境工程科学专业数",
        "医技护理康复专业数",
        "临床口腔中医暂缓专业数",
    ]:
        direction_group_counts[field] = sum(1 for row in enriched_groups if count(row, field) > 0)
        direction_major_counts[field] = sum(count(row, field) for row in enriched_groups)

    summary = {
        "status": "issue19_round3_unrestricted_region_ready",
        "generated_by": "build_issue19_round3_unrestricted_region_candidates.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "city_policy": CITY_POLICY,
        "selection_boundary": "第三轮不限地区候选池只用于家庭讨论和入围组核验，不作为定稿依据。",
        "family_constraints": [
            "优先公办本科、普通学费，普通主线年学费上限15000。",
            "城市地区本轮不限制、不加分、不设优先城市名额。",
            "临床/口腔/中医等医学主线暂缓；护理/助产、动物医学/兽医/动物科学暂不纳入。",
            "医技/康复只进入低优先级专项了解，不自动进入普通主线。",
        ],
        "candidate_group_rows": len(all_candidates),
        "main_pool_rows_before_shortlist": len(main_pool),
        "main_shortlist_group_rows": len(main_shortlist),
        "discussion_priority_group_rows": len(discussion_priority),
        "special_pool_rows_before_shortlist": len(special_pool),
        "special_group_rows": len(special_groups),
        "main_shortlist_major_rows": len(main_major_rows),
        "special_major_rows": len(special_major_rows),
        "city_distribution_rows": len(city_distribution),
        "discussion_gradient_distribution": dict(Counter(row.get("第三轮建议梯度", "") for row in discussion_priority)),
        "main_shortlist_gradient_distribution": dict(Counter(row.get("第三轮建议梯度", "") for row in main_shortlist)),
        "top_city_counts_in_main_shortlist": Counter(row.get("城市候选", "") or "未识别" for row in main_shortlist).most_common(20),
        "direction_group_counts": direction_group_counts,
        "direction_major_counts": direction_major_counts,
        "workbook": str(ROUND3_WORKBOOK.relative_to(ROOT)),
        "candidate_groups_csv": str(ROUND3_GROUPS.relative_to(ROOT)),
        "main_shortlist_csv": str(ROUND3_MAIN_SHORTLIST.relative_to(ROOT)),
        "discussion_priority_csv": str(ROUND3_DISCUSSION_PRIORITY.relative_to(ROOT)),
        "special_groups_csv": str(ROUND3_SPECIAL_GROUPS.relative_to(ROOT)),
        "main_majors_csv": str(ROUND3_MAIN_MAJORS.relative_to(ROOT)),
        "special_majors_csv": str(ROUND3_SPECIAL_MAJORS.relative_to(ROOT)),
        "city_distribution_csv": str(ROUND3_CITY_DISTRIBUTION.relative_to(ROOT)),
        "hard_filters_main": [
            "公办/普通学费机器线索",
            "组内高收费或超预算专业数为0",
            "必须有逐专业明细",
            "临床/口腔/中医等暂缓方向不进主线",
            "护理/助产、动物医学/兽医/动物科学不进主线",
            "医技/康复不混入主线",
        ],
    }
    ROUND3_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe(
        [
            ROUND3_GROUPS,
            ROUND3_MAIN_SHORTLIST,
            ROUND3_DISCUSSION_PRIORITY,
            ROUND3_SPECIAL_GROUPS,
            ROUND3_MAIN_MAJORS,
            ROUND3_SPECIAL_MAJORS,
            ROUND3_CITY_DISTRIBUTION,
            ROUND3_SUMMARY,
        ]
    )

    print(f"写出第三轮不限地区完整候选池：{ROUND3_GROUPS.relative_to(ROOT)}")
    print(f"写出第三轮不限地区主线精选：{ROUND3_MAIN_SHORTLIST.relative_to(ROOT)}")
    print(f"写出第三轮不限地区优先讨论：{ROUND3_DISCUSSION_PRIORITY.relative_to(ROOT)}")
    print(f"写出第三轮低优先级专项：{ROUND3_SPECIAL_GROUPS.relative_to(ROOT)}")
    print(f"写出第三轮工作簿：{ROUND3_WORKBOOK.relative_to(ROOT)}")
    print(
        f"主线池={len(main_pool)}；主线精选={len(main_shortlist)}；"
        f"优先讨论={len(discussion_priority)}；专项池={len(special_pool)}；专项精选={len(special_groups)}；"
        f"完整候选={len(all_candidates)}"
    )


if __name__ == "__main__":
    main()
