#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

PROGRESS = WORKING / "issue19-school-source-progress-board-public-ledger.csv"

OUTPUT = WORKING / "issue19-school-source-structured-ingestion-candidates-public-ledger.csv"
SUMMARY_OUTPUT = WORKING / "issue19-school-source-structured-ingestion-candidates-summary.json"
WORKBOOK_OUTPUT = EXPORTS / "issue19-school-source-structured-ingestion-candidates.xlsx"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
DATA_STAGE = "issue19_school_source_structured_ingestion_candidates"
GENERATED_AT = "2026-06-29"

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "可否进入最终志愿方案",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

FIELDS = [
    "高校源结构化接入候选ID",
    "来源高校源进度看板",
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "任务粒度",
    *FALSE_FIELDS,
    "接入优先序",
    "院校代码",
    "院校名称公开",
    "学校键SHA16",
    "接入批次",
    "推荐接入动作",
    "来源族",
    "来源文件类型",
    "本地公开证据文件集合",
    "本地公开证据文件集合SHA16",
    "公开URL线索数量",
    "进度看板任务数",
    "涉及招生明细数合计",
    "涉及专业组数最大值",
    "最新高校侧证据层级集合",
    "下一批推进优先级集合",
    "结构化适配器状态",
    "匹配规则状态",
    "候选diff优先级",
    "阻断原因",
    "下一脚本建议",
    "PDF原页核页状态",
    "湖北官方系统或省招办计划核验状态",
    "高校官网源状态",
    "字段事实写回状态",
    "公开安全策略",
]

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/home/",
    "/var/folders/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    "file://",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".tif",
    ".tiff",
    ".heic",
    "Authorization",
    "Bearer ",
    "Cookie",
    "Set-Cookie",
    "access_token",
    "refresh_token",
    "password",
    "secret",
    "api_key",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "OCR行文本",
    "字段确认值",
    "人工读数",
    "候选值",
    "PDF原页人工读数",
    "湖北官方字段值",
    "高校官网或招生章程字段值",
    "已确认",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
    "可排序",
]

PLAN = [
    {
        "rank": 1,
        "code": "A195",
        "school": "兰州大学",
        "batch": "B0B1_API_JSON",
        "action": "复跑API结构化源，补专业名归一和计划数diff闭合。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-b0-b1-official-sources/lzu-2026-hubei-physics-normal-zsjh.json"],
        "url_count": 1,
        "adapter": "existing_json_adapter_needs_schema_check",
        "match_rule": "needs_major_name_normalization",
        "diff_priority": "D1-已有JSON源，优先生成补缺或冲突diff",
        "blocker": "PDF原页、湖北官方侧和人工复核未闭环，不能写回。",
        "script_hint": "新增或复用 LZU JSON schema adapter，输出高校侧候选diff。",
    },
    {
        "rank": 2,
        "code": "C125",
        "school": "武汉轻工大学",
        "batch": "NEXT20_API_JSON",
        "action": "把next20智能答疑JSON接入C4/C6候选diff桥接。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-next20-official-sources/whpu-2026-hubei-physics-normal.json"],
        "url_count": 1,
        "adapter": "new_next20_json_bridge_needed",
        "match_rule": "needs_group_boundary_absent_policy",
        "diff_priority": "D0-next20高明细量优先接入",
        "blocker": "官网API未给湖北院校专业组边界，仍需PDF原页和湖北官方侧。",
        "script_hint": "写 next20 JSON -> school source normalized rows bridge。",
    },
    {
        "rank": 3,
        "code": "C133",
        "school": "湖北师范大学",
        "batch": "NEXT20_API_JSON",
        "action": "把next20智能答疑JSON接入候选diff；章程仅作规则侧证。",
        "source_type": "API/JSON；章程HTML",
        "evidence": [
            "data/external/issue19-next20-official-sources/hbnu-2026-hubei-physics-group-normal.json",
            "data/external/issue19-school-source-live-20260629/hbnu-2026-undergraduate-charter.html",
        ],
        "url_count": 2,
        "adapter": "new_next20_json_bridge_needed",
        "match_rule": "needs_batch_and_subject_note_policy",
        "diff_priority": "D0-next20高明细量优先接入",
        "blocker": "章程不能核分省分专业计划，官网API也缺湖北专业组边界。",
        "script_hint": "next20 JSON 进diff；章程另入限制规则侧账。",
    },
    {
        "rank": 4,
        "code": "K465",
        "school": "西安建筑科技大学",
        "batch": "C4C6_API_JSON",
        "action": "把已抓API结果接入统一结构化候选diff，并保留冲突分层。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-c4-c6-official-sources/xauat-2026-hubei-physics-normal.json"],
        "url_count": 1,
        "adapter": "existing_c4c6_json_adapter_needs_unified_bridge",
        "match_rule": "needs_conflict_priority_rules",
        "diff_priority": "D0-冲突候选优先回页",
        "blocker": "存在冲突/补缺候选，字段事实写回仍被阻断。",
        "script_hint": "接入 C4/C6 API normalized rows，再重算 diff 状态。",
    },
    {
        "rank": 5,
        "code": "A032",
        "school": "北京语言大学",
        "batch": "C4C6_API_JSON",
        "action": "补专业名、批次和类别规则，重跑候选diff。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-c4-c6-official-sources/blcu-2026-hubei-physics-normal.json"],
        "url_count": 1,
        "adapter": "existing_c4c6_json_adapter_needs_rule_patch",
        "match_rule": "needs_major_category_and_batch_rules",
        "diff_priority": "D0-冲突候选优先回页",
        "blocker": "计划数冲突线索必须回PDF原页和湖北官方侧。",
        "script_hint": "补 BLCU JSON 专业名/类别/批次匹配规则。",
    },
    {
        "rank": 6,
        "code": "F099",
        "school": "天津外国语大学",
        "batch": "B0B1_API_JSON",
        "action": "复跑API并统一JSON schema adapter，生成补缺diff。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-b0-b1-official-sources/tjfsu-2026-hubei-physics-normal-zsjh.json"],
        "url_count": 1,
        "adapter": "existing_json_adapter_needs_schema_check",
        "match_rule": "needs_school_major_group_absent_policy",
        "diff_priority": "D1-补缺线索优先回页",
        "blocker": "官网源不含湖北专业组边界，不能替代湖北官方计划。",
        "script_hint": "新增 TJFSU JSON schema adapter 并输出候选diff。",
    },
    {
        "rank": 7,
        "code": "F305",
        "school": "忻州师范学院",
        "batch": "B0B1_PDF_CSV",
        "action": "对PDF抽取CSV补字段映射校验和专业名归一。",
        "source_type": "PDF抽取CSV",
        "evidence": [
            "data/external/issue19-b0-b1-official-sources/xztu-2026-province-major-plan.pdf",
            "data/external/issue19-b0-b1-official-sources/xztu-2026-hubei-physics-plan-extracted.csv",
        ],
        "url_count": 1,
        "adapter": "existing_pdf_csv_adapter_needs_grid_check",
        "match_rule": "needs_major_name_and_subject_mapping",
        "diff_priority": "D1-PDF抽取源补缺优先",
        "blocker": "高校PDF附件不是湖北官方计划，需PDF原页和湖北官方侧复核。",
        "script_hint": "补 XZTU PDF/CSV grid parser 的字段映射校验。",
    },
    {
        "rank": 8,
        "code": "K487",
        "school": "西安航空学院",
        "batch": "LIVE_PDF_CSV",
        "action": "把live crosscheck转成标准化高校源输入，并加院校代码边界校验。",
        "source_type": "PDF抽取CSV",
        "evidence": [
            "data/external/issue19-school-source-live-20260629/xaau-2026-province-major-plan.pdf",
            "data/external/issue19-school-source-live-20260629/xaau-2026-hubei-physics-plan-extracted.csv",
            "data/working/issue19-school-source-live-20260629-xaau-crosscheck.csv",
        ],
        "url_count": 1,
        "adapter": "new_live_pdf_csv_bridge_needed",
        "match_rule": "needs_school_code_boundary_guard",
        "diff_priority": "D2-live源接入并查串行风险",
        "blocker": "live账本明确只作补缺线索，不能直接写主表。",
        "script_hint": "新增 XAAU live crosscheck -> normalized school source bridge。",
    },
    {
        "rank": 9,
        "code": "C108",
        "school": "江汉大学",
        "batch": "B0B1_API_JSON",
        "action": "对强辅证JSON做冲突抽检和未匹配规则补丁。",
        "source_type": "API/JSON",
        "evidence": ["data/external/issue19-b0-b1-official-sources/jhun-2026-hubei-physics-normal-lqxx2.json"],
        "url_count": 1,
        "adapter": "existing_json_adapter_needs_conflict_audit",
        "match_rule": "needs_unmatched_major_rules",
        "diff_priority": "D0-强辅证冲突优先",
        "blocker": "存在冲突候选，湖北官方侧未闭环。",
        "script_hint": "重跑 JHUN JSON diff，并输出冲突抽检包。",
    },
    {
        "rank": 10,
        "code": "K753",
        "school": "喀什大学",
        "batch": "B0B1_XLSX",
        "action": "拆XLSX省份列、科类和批次，生成抽检diff。",
        "source_type": "XLSX",
        "evidence": ["data/external/issue19-b0-b1-official-sources/ksu-2026-undergraduate-plan.xlsx"],
        "url_count": 1,
        "adapter": "new_xlsx_province_column_adapter_needed",
        "match_rule": "needs_subject_and_batch_split",
        "diff_priority": "D0-XLSX源冲突抽检",
        "blocker": "官网附件缺湖北专业组边界和专业代号。",
        "script_hint": "新增 KSU XLSX 省份列拆分和科类识别校验。",
    },
    {
        "rank": 11,
        "code": "H450",
        "school": "山东工商学院",
        "batch": "B0B1_PDF_CSV",
        "action": "补PDF表格规则和专业名归一后重跑diff。",
        "source_type": "PDF抽取CSV",
        "evidence": [
            "data/external/issue19-b0-b1-official-sources/sdtbu-2026-province-major-plan.pdf",
            "data/external/issue19-b0-b1-official-sources/sdtbu-2026-hubei-physics-plan-extracted.csv",
        ],
        "url_count": 1,
        "adapter": "existing_pdf_csv_adapter_needs_table_rule_patch",
        "match_rule": "needs_unmatched_major_rules",
        "diff_priority": "D0-PDF源冲突和未匹配优先",
        "blocker": "有冲突候选和未匹配项，不能自动写回。",
        "script_hint": "补 SDTBU PDF table parser 和专业名规则。",
    },
    {
        "rank": 12,
        "code": "H001",
        "school": "杭州电子科技大学",
        "batch": "B0B1_XLSX",
        "action": "XLSX adapter补专业组、批次和选科一致性抽检。",
        "source_type": "XLSX",
        "evidence": ["data/external/issue19-b0-b1-official-sources/hdu-2026-hubei-plan.xlsx"],
        "url_count": 1,
        "adapter": "existing_xlsx_adapter_needs_consistency_audit",
        "match_rule": "needs_group_batch_subject_audit",
        "diff_priority": "D0-XLSX强辅证抽检",
        "blocker": "官网附件未直接给湖北院校专业组代码和专业代号。",
        "script_hint": "补 HDU XLSX adapter 的组/批次/选科一致性抽检。",
    },
]


def clean(value):
    return "" if value is None else str(value).replace("\r", " ").replace("\n", " ").strip()


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in fields} for row in rows])


def stable_id(prefix, parts):
    text = "|".join(clean(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(text.encode('utf-8')).hexdigest()[:16]}"


def sha16(values):
    normalized = "\n".join(sorted({clean(value) for value in values if clean(value)}))
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:16]


def sha256_file(path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def as_int(value):
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def counter_join(rows, field):
    counter = Counter(clean(row.get(field)) for row in rows if clean(row.get(field)))
    return "；".join(f"{key}×{value}" for key, value in sorted(counter.items()))


def max_int(rows, field):
    return max([as_int(row.get(field)) for row in rows] or [0])


def base_row():
    row = {
        "来源高校源进度看板": str(PROGRESS.relative_to(ROOT)),
        "来源期号": SOURCE_ISSUE,
        "来源PDF_SHA256": SOURCE_PDF_SHA256,
        "生成日期": GENERATED_AT,
        "数据阶段": DATA_STAGE,
        "主表粒度": "高校官网结构化源接入候选学校",
        "任务粒度": "高校×结构化适配器候选",
    }
    row.update({field: "false" for field in FALSE_FIELDS})
    return row


def evidence_sha(files):
    parts = []
    for rel in files:
        path = ROOT / rel
        if not path.exists():
            raise SystemExit(f"missing evidence file: {rel}")
        parts.append(f"{rel}:{sha256_file(path)}")
    return sha16(parts)


def public_policy():
    return "not_final；structured_source_ingestion_only；for_double_check_only；no_field_values；no_official_replacement"


def workbook_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([clean(row.get(field, "")) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[column].width = 24
    ws.column_dimensions["T"].width = 52
    ws.column_dimensions["U"].width = 52
    return ws


def assert_public_safe(rows, label):
    text = json.dumps(rows, ensure_ascii=False)
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]
    if found:
        raise SystemExit(f"{label} contains forbidden public tokens: {found[:5]}")


def main():
    progress_rows = read_csv(PROGRESS)
    by_school = defaultdict(list)
    for row in progress_rows:
        by_school[clean(row.get("院校代码"))].append(row)

    output_rows = []
    for plan in PLAN:
        code = plan["code"]
        related = by_school[code]
        if not related:
            raise SystemExit(f"school not found in progress board: {code}")
        row = base_row()
        row.update(
            {
                "高校源结构化接入候选ID": stable_id("SS-INGEST", [SOURCE_PDF_SHA256, code, plan["batch"]]),
                "接入优先序": plan["rank"],
                "院校代码": code,
                "院校名称公开": plan["school"],
                "学校键SHA16": sha16([code, plan["school"]]),
                "接入批次": plan["batch"],
                "推荐接入动作": plan["action"],
                "来源族": counter_join(related, "最新证据来源族"),
                "来源文件类型": plan["source_type"],
                "本地公开证据文件集合": "；".join(plan["evidence"]),
                "本地公开证据文件集合SHA16": evidence_sha(plan["evidence"]),
                "公开URL线索数量": plan["url_count"],
                "进度看板任务数": len(related),
                "涉及招生明细数合计": sum(as_int(item.get("涉及招生明细数")) for item in related),
                "涉及专业组数最大值": max_int(related, "涉及专业组数"),
                "最新高校侧证据层级集合": counter_join(related, "最新高校侧证据层级"),
                "下一批推进优先级集合": counter_join(related, "下一批推进优先级"),
                "结构化适配器状态": plan["adapter"],
                "匹配规则状态": plan["match_rule"],
                "候选diff优先级": plan["diff_priority"],
                "阻断原因": plan["blocker"],
                "下一脚本建议": plan["script_hint"],
                "PDF原页核页状态": "pending_manual_pdf_review",
                "湖北官方系统或省招办计划核验状态": "pending_hubei_official_review",
                "高校官网源状态": "structured_source_candidate_for_double_check_only",
                "字段事实写回状态": "blocked_until_pdf_hubei_official_review",
                "公开安全策略": public_policy(),
            }
        )
        output_rows.append(row)

    assert_public_safe(output_rows, "structured_ingestion_candidates")
    write_csv(OUTPUT, output_rows, FIELDS)

    wb = Workbook()
    wb.remove(wb.active)
    workbook_sheet(wb, "结构化接入候选12校", output_rows, FIELDS)
    wb.save(WORKBOOK_OUTPUT)

    summary = {
        "status": "issue19_school_source_structured_ingestion_candidates_ready_not_final",
        "generated_by": "build_issue19_school_source_structured_ingestion_candidates.py",
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "output": str(OUTPUT.relative_to(ROOT)),
        "workbook_output": str(WORKBOOK_OUTPUT.relative_to(ROOT)),
        "row_count": len(output_rows),
        "unique_school_count": len({row["院校代码"] for row in output_rows}),
        "source_type_counts": dict(Counter(row["来源文件类型"] for row in output_rows)),
        "batch_counts": dict(Counter(row["接入批次"] for row in output_rows)),
        "total_attached_progress_tasks": sum(as_int(row["进度看板任务数"]) for row in output_rows),
        "total_attached_major_detail_count": sum(as_int(row["涉及招生明细数合计"]) for row in output_rows),
        "final_available_count": 0,
        "official_plan_replacement_allowed_count": 0,
        "field_writeback_allowed_count": 0,
        "policy": "本表只定义高校公开结构化源接入候选，不写回第19期主表，不替代湖北官方计划。",
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text = OUTPUT.read_text(encoding="utf-8", errors="ignore") + SUMMARY_OUTPUT.read_text(
        encoding="utf-8", errors="ignore"
    )
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in public_text]
    if found:
        raise SystemExit(f"public output contains forbidden tokens: {found[:5]}")

    print(f"wrote {OUTPUT}")
    print(f"wrote {SUMMARY_OUTPUT}")
    print(f"wrote {WORKBOOK_OUTPUT}")


if __name__ == "__main__":
    main()
