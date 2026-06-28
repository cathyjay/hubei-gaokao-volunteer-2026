#!/usr/bin/env python3
import csv
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

STATUS_CSV = WORKING / "issue19-stable-foundation-v0-status.csv"
GROUP_VIEW_CSV = WORKING / "issue19-stable-foundation-group-screening-view.csv"
MAJOR_VIEW_CSV = WORKING / "issue19-stable-foundation-major-screening-view.csv"
GROUP_READY_CSV = WORKING / "issue19-stable-foundation-group-readiness-bridge.csv"

GROUP_BROWSER_CSV = EXPORTS / "issue19-stable-foundation-group-browser.csv"
MAJOR_BROWSER_CSV = EXPORTS / "issue19-stable-foundation-major-browser.csv"
OBSERVATION_GROUP_CSV = EXPORTS / "issue19-stable-foundation-observation-groups.csv"
MACHINE_SIGNAL_CSV = EXPORTS / "issue19-stable-foundation-machine-signal-majors.csv"
QUICK_SAMPLE_CSV = EXPORTS / "issue19-stable-foundation-quick-verification-sample.csv"
CANDIDATE_TEMPLATE_CSV = EXPORTS / "issue19-stable-foundation-candidate-discussion-template.csv"
WORKBOOK_XLSX = EXPORTS / "issue19-stable-foundation-browser.xlsx"
SUMMARY_JSON = EXPORTS / "issue19-stable-foundation-browser-summary.json"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

GROUP_BROWSER_FIELDS = [
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "医学护理排除专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "机器筛选价值层级",
    "是否进入机器初筛观察池",
    "保真核验层级",
    "历史线索分层",
    "同代码命中年份最大值",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "组内完整专业清单索引",
    "机器核验下一步",
    "人工核验下一步",
    "不得进入原因",
    "下一步",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

MAJOR_BROWSER_FIELDS = [
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器专业接受度初判",
    "专业风险类型",
    "是否超预算机器初判",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "是否可低风险抽检",
    "是否可作为机器初筛线索",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网来源状态",
    "高校官网证据匹配状态",
    "历史线索分层",
    "同代码命中年份数",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
    "机器核验下一步",
    "人工核验下一步",
    "不得进入原因",
    "下一步",
]

SAMPLE_EXTRA_FIELDS = [
    "抽样批次",
    "抽样理由",
    "原页核验结果",
    "湖北官方核验结果",
    "高校辅证核验结果",
    "是否需要升级同组核验",
    "核验备注",
]

CANDIDATE_TEMPLATE_EXTRA_FIELDS = [
    "家庭是否愿意讨论",
    "调剂是否可接受",
    "需要排除的组内专业",
    "待补充问题",
    "核验备注",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def int_value(row, field):
    try:
        return int(str(row.get(field, "")).strip() or "0")
    except ValueError:
        return 0


def float_value(row, field):
    try:
        return float(str(row.get(field, "")).strip() or "999999")
    except ValueError:
        return 999999.0


def stable_rank(row):
    key = "|".join(
        [
            row.get("专业行ID", ""),
            row.get("专业组出现ID", ""),
            row.get("院校代码", ""),
            row.get("专业代号OCR", ""),
        ]
    )
    return hashlib.sha256(("issue19-browser-sample|" + key).encode("utf-8")).hexdigest()


def group_sort_key(row):
    value_order = {
        "V1-偏好专业优先线索": 0,
        "V3-历史线索普通留存": 1,
        "V4-普通留存待了解": 2,
        "V0-默认不进主方案风险": 3,
        "V5-无逐专业明细先补结构": 4,
    }
    hist_order = {
        "H1-历史最高线低于等位分20分以上": 0,
        "H2-历史最高线不高于等位分": 1,
        "H3-历史最高线高于等位分15分内": 2,
        "H4-历史最高线高于等位分16-30分": 3,
        "H5-历史最高线高于等位分30分以上": 4,
        "H0-无同代码历史线索": 5,
    }
    return (
        value_order.get(row.get("机器筛选价值层级", ""), 9),
        hist_order.get(row.get("历史线索分层", ""), 9),
        -int_value(row, "偏好专业数"),
        int_value(row, "医学护理排除专业数"),
        int_value(row, "高收费或超预算专业数"),
        row.get("院校代码", ""),
        row.get("院校专业组代码OCR规范化", ""),
    )


def major_sort_key(row):
    priority_order = {
        "P0-最终候选/冲突/结构强阻断先核": 0,
        "P1-字段缺口和B0B1辅证优先核": 1,
        "P2-家庭费用调剂与三方闭环核": 2,
        "P3-低风险抽检但非最终": 3,
    }
    return (
        0 if row.get("专业偏好方向", "") else 1,
        priority_order.get(row.get("人工核验优先级", ""), 9),
        float_value(row, "历史最高等位分差"),
        row.get("院校代码", ""),
        row.get("院校专业组代码OCR规范化", ""),
        int_value(row, "专业组内专业序号"),
    )


def add_sample(rows, selected, selected_ids, limit, reason, predicate):
    matches = [row for row in rows if predicate(row) and row.get("专业行ID") not in selected_ids]
    for row in sorted(matches, key=stable_rank)[:limit]:
        sampled = dict(row)
        sampled["抽样理由"] = reason
        sampled["抽样批次"] = f"S{len(selected) + 1:03d}"
        sampled["原页核验结果"] = ""
        sampled["湖北官方核验结果"] = ""
        sampled["高校辅证核验结果"] = ""
        sampled["是否需要升级同组核验"] = ""
        sampled["核验备注"] = ""
        selected.append(sampled)
        selected_ids.add(row.get("专业行ID"))


def build_quick_sample(major_rows):
    selected = []
    selected_ids = set()
    add_sample(
        major_rows,
        selected,
        selected_ids,
        35,
        "偏好专业方向抽样：数字媒体技术/计算机/师范",
        lambda r: bool(r.get("专业偏好方向", "")),
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        30,
        "优先城市抽样：武汉/成都/西安/北京",
        lambda r: r.get("城市候选", "") in {"武汉", "成都", "西安", "北京"},
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        30,
        "P0强阻断或冲突抽样",
        lambda r: r.get("人工核验优先级", "").startswith("P0-"),
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        30,
        "P1字段缺口和高校辅证抽样",
        lambda r: r.get("人工核验优先级", "").startswith("P1-"),
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        20,
        "高校官网辅证命中或差异抽样",
        lambda r: bool(r.get("高校官网证据匹配状态", "")) or bool(r.get("高校官网来源状态", "")),
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        20,
        "低风险抽检样本",
        lambda r: r.get("是否可低风险抽检", "") == "true",
    )
    add_sample(
        major_rows,
        selected,
        selected_ids,
        25,
        "历史线索接近样本：H1/H2/H3",
        lambda r: r.get("历史线索分层", "").startswith(("H1-", "H2-", "H3-")),
    )
    return sorted(selected, key=lambda r: r["抽样批次"])


def build_candidate_template(group_rows):
    rows = []
    for row in group_rows:
        if row.get("机器筛选价值层级", "") != "V1-偏好专业优先线索":
            continue
        out = dict(row)
        for field in CANDIDATE_TEMPLATE_EXTRA_FIELDS:
            out[field] = ""
        rows.append(out)
    return sorted(rows, key=group_sort_key)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "private/",
        "private\\",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
        "人工读数",
        "已核准",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"公开浏览导出包含禁止内容：{hit}")


def append_sheet(ws, rows, fields):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, start=1):
        max_len = min(
            42,
            max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:200]]),
        )
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(status_rows, group_rows, major_rows, observation_rows, signal_rows, sample_rows, candidate_rows):
    wb = Workbook()
    wb.remove(wb.active)

    guide = wb.create_sheet("00_怎么看")
    guide_rows = [
        ("当前定位", "这是第19期稳定数据基座 V0 的浏览与核验工作簿，只用于候选发现和核验排队，不作为定稿依据。"),
        ("先看哪张", "先看 02_专业组浏览，再下钻到 03_逐专业浏览；想快速找线索看 05_初筛专业。"),
        ("怎么筛", "优先筛：城市候选、公办民办机器线索、专业偏好方向、调剂初判、历史线索分层、人工核验优先级。"),
        ("怎么核验", "先看 06_快速核验抽样；入围组必须回看第19期原页、专业组完整清单、学费、计划数、选科、备注限制和调剂范围。"),
        ("边界", "高校官网和第三方只做 double check；第19期原页和湖北官方侧仍是定稿前必须闭环的证据。"),
    ]
    guide.append(["项目", "说明"])
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 100

    append_sheet(wb.create_sheet("01_V0状态"), status_rows, list(status_rows[0].keys()))
    append_sheet(wb.create_sheet("02_专业组浏览"), group_rows, GROUP_BROWSER_FIELDS)
    append_sheet(wb.create_sheet("03_逐专业浏览"), major_rows, MAJOR_BROWSER_FIELDS)
    append_sheet(wb.create_sheet("04_观察池专业组"), observation_rows, GROUP_BROWSER_FIELDS)
    append_sheet(wb.create_sheet("05_初筛专业"), signal_rows, MAJOR_BROWSER_FIELDS)
    append_sheet(wb.create_sheet("06_快速核验抽样"), sample_rows, SAMPLE_EXTRA_FIELDS + MAJOR_BROWSER_FIELDS)
    append_sheet(
        wb.create_sheet("07_候选讨论模板"),
        candidate_rows,
        CANDIDATE_TEMPLATE_EXTRA_FIELDS + GROUP_BROWSER_FIELDS,
    )

    WORKBOOK_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_XLSX)


def main():
    status_rows = read_csv(STATUS_CSV)
    group_rows = sorted(read_csv(GROUP_VIEW_CSV), key=group_sort_key)
    major_rows = sorted(read_csv(MAJOR_VIEW_CSV), key=major_sort_key)
    group_ready_rows = read_csv(GROUP_READY_CSV)
    ready_by_group = {row.get("专业组出现ID", ""): row for row in group_ready_rows}

    observation_rows = [
        row for row in group_rows if row.get("是否进入机器初筛观察池", "") == "true"
    ]
    signal_rows = [
        row for row in major_rows if row.get("是否可作为机器初筛线索", "") == "true"
    ]
    sample_rows = build_quick_sample(major_rows)
    candidate_rows = build_candidate_template(group_rows)

    write_csv(GROUP_BROWSER_CSV, group_rows, GROUP_BROWSER_FIELDS)
    write_csv(MAJOR_BROWSER_CSV, major_rows, MAJOR_BROWSER_FIELDS)
    write_csv(OBSERVATION_GROUP_CSV, observation_rows, GROUP_BROWSER_FIELDS)
    write_csv(MACHINE_SIGNAL_CSV, signal_rows, MAJOR_BROWSER_FIELDS)
    write_csv(QUICK_SAMPLE_CSV, sample_rows, SAMPLE_EXTRA_FIELDS + MAJOR_BROWSER_FIELDS)
    write_csv(
        CANDIDATE_TEMPLATE_CSV,
        candidate_rows,
        CANDIDATE_TEMPLATE_EXTRA_FIELDS + GROUP_BROWSER_FIELDS,
    )
    write_workbook(
        status_rows,
        group_rows,
        major_rows,
        observation_rows,
        signal_rows,
        sample_rows,
        candidate_rows,
    )

    first_closure_group_count = sum(
        1
        for row in group_rows
        if ready_by_group.get(row.get("专业组出现ID", ""), {})
        .get("第一闭环覆盖状态", "")
        .startswith("Y-")
    )
    summary = {
        "status": "issue19_stable_foundation_browser_export_ready",
        "generated_by": "export_issue19_stable_foundation_browser.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "workbook": str(WORKBOOK_XLSX.relative_to(ROOT)),
        "group_browser_csv": str(GROUP_BROWSER_CSV.relative_to(ROOT)),
        "major_browser_csv": str(MAJOR_BROWSER_CSV.relative_to(ROOT)),
        "observation_group_csv": str(OBSERVATION_GROUP_CSV.relative_to(ROOT)),
        "machine_signal_csv": str(MACHINE_SIGNAL_CSV.relative_to(ROOT)),
        "quick_verification_sample_csv": str(QUICK_SAMPLE_CSV.relative_to(ROOT)),
        "candidate_discussion_template_csv": str(CANDIDATE_TEMPLATE_CSV.relative_to(ROOT)),
        "group_browser_rows": len(group_rows),
        "major_browser_rows": len(major_rows),
        "observation_group_rows": len(observation_rows),
        "machine_signal_major_rows": len(signal_rows),
        "quick_verification_sample_rows": len(sample_rows),
        "candidate_discussion_template_rows": len(candidate_rows),
        "first_closure_group_count": first_closure_group_count,
        "usage_boundary": "只用于候选发现、数据浏览和核验排队；不作为定稿依据。",
        "rapid_verification_strategy": [
            "先用 baseline/checksum/血缘审计证明结构化链路没有漂移。",
            "再用快速核验抽样覆盖偏好专业、城市、P0/P1风险、官网辅证、低风险抽检和历史线索接近项。",
            "对进入家庭讨论或志愿梯度的院校专业组做整组100%核验。",
        ],
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe(
        [
            GROUP_BROWSER_CSV,
            MAJOR_BROWSER_CSV,
            OBSERVATION_GROUP_CSV,
            MACHINE_SIGNAL_CSV,
            QUICK_SAMPLE_CSV,
            CANDIDATE_TEMPLATE_CSV,
            SUMMARY_JSON,
        ]
    )
    print(f"写出浏览工作簿：{WORKBOOK_XLSX.relative_to(ROOT)}")
    print(f"写出专业组浏览CSV：{GROUP_BROWSER_CSV.relative_to(ROOT)}")
    print(f"写出逐专业浏览CSV：{MAJOR_BROWSER_CSV.relative_to(ROOT)}")
    print(f"写出快速核验抽样CSV：{QUICK_SAMPLE_CSV.relative_to(ROOT)}")
    print(f"快速核验抽样行数：{len(sample_rows)}")


if __name__ == "__main__":
    main()
