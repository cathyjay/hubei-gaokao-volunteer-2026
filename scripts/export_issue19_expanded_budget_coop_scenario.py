#!/usr/bin/env python3
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

SCENARIO = WORKING / "family-preferences-expanded-2026-06-28.json"
GROUP_BROWSER = EXPORTS / "issue19-stable-foundation-group-browser.csv"
MAJOR_BROWSER = EXPORTS / "issue19-stable-foundation-major-browser.csv"

GROUP_OUTPUT = EXPORTS / "issue19-expanded-budget-coop-groups.csv"
MAJOR_OUTPUT = EXPORTS / "issue19-expanded-budget-coop-majors.csv"
WORKBOOK_OUTPUT = EXPORTS / "issue19-expanded-budget-coop-scenario.xlsx"
SUMMARY_OUTPUT = EXPORTS / "issue19-expanded-budget-coop-scenario-summary.json"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

GROUP_FIELDS = [
    "扩展预算场景专业组ID",
    "场景层级",
    "场景判断",
    "预算内中外合作或高收费专业数",
    "超过预算专业数",
    "待核费用专业数",
    "场景内最低学费",
    "场景内最高学费",
    "场景内偏好专业方向",
    "场景内专业摘要",
    "下一步核验动作",
    "是否可作为定稿依据",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "医学护理排除专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "机器筛选价值层级",
    "保真核验层级",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "组内完整专业清单索引",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

MAJOR_FIELDS = [
    "扩展预算场景专业ID",
    "专业组出现ID",
    "场景判断",
    "费用判断",
    "中外合作或高收费线索",
    "是否医学相关待确认",
    "下一步核验动作",
    "是否可作为定稿依据",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "人工核验优先级",
    "人工核验强度",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "专业行ID",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def stable_id(prefix, parts):
    return f"{prefix}-{hashlib.sha1('|'.join(str(part) for part in parts).encode('utf-8')).hexdigest()[:16]}"


def int_value(value):
    try:
        return int(str(value or "").strip())
    except ValueError:
        return 0


def tuition_value(value):
    text = str(value or "").replace(",", "").replace("，", "").replace("［", "").replace("[", "")
    if "万" in text:
        match = re.search(r"(\d+(?:\.\d+)?)\s*万", text)
        if match:
            return int(float(match.group(1)) * 10000)
    digits = re.findall(r"\d+", text)
    if not digits:
        return None
    # Pick the longest token so OCR fragments like "3+1, 28000" do not become 31.
    value = int(max(digits, key=len))
    # Tuition is annual yuan. Tiny fragments like "4" or "6" are usually OCR
    # residue from "4+0"/plan count/years, not a usable tuition reading.
    if value < 1000:
        return None
    return value


def has_coop_or_high_fee(row):
    text = "；".join(
        [
            row.get("专业名称及备注短摘", ""),
            row.get("专业风险类型", ""),
            row.get("家庭底线属性动作", ""),
            row.get("机器专业接受度初判", ""),
        ]
    )
    return any(token in text for token in ["中外合作", "合作办学", "高收费"])


def is_medical_pending(row):
    text = "；".join([row.get("专业名称及备注短摘", ""), row.get("专业风险类型", "")])
    return any(token in text for token in ["医学", "护理", "检验", "影像", "药学", "康复", "口腔", "中医"])


def scenario_judgement(row, max_tuition):
    tuition = tuition_value(row.get("学费OCR候选", ""))
    coop = has_coop_or_high_fee(row)
    high_tuition = tuition is not None and tuition > 15000
    if not coop and not high_tuition:
        return "", "", "", tuition
    if tuition is None:
        return "费用待核", "待核费用", "需要回原页和章程确认学费", tuition
    if tuition > max_tuition:
        return "超过预算", "超过预算", f"学费候选 {tuition} 元超过 {max_tuition} 元", tuition
    if coop:
        return "预算内中外合作/高收费可讨论", "预算内", f"学费候选 {tuition} 元，不超过 {max_tuition} 元；需核培养模式和证书", tuition
    return "预算内高学费可讨论", "预算内", f"学费候选 {tuition} 元，不超过 {max_tuition} 元；需核是否普通收费或特殊项目", tuition


def preference_rank(row):
    pref = row.get("专业偏好方向", "")
    if "数字媒体技术" in pref:
        return 0
    if "计算机" in pref:
        return 1
    if "师范" in pref:
        return 2
    return 3


def group_sort_key(row):
    judgement_order = {
        "预算内中外合作/高收费可讨论": 0,
        "预算内高学费可讨论": 1,
        "费用待核": 2,
        "超过预算": 3,
    }
    return (
        judgement_order.get(row.get("场景判断", ""), 9),
        -int_value(row.get("预算内中外合作或高收费专业数")),
        int_value(row.get("医学护理排除专业数")),
        row.get("历史线索分层", ""),
        row.get("院校代码", ""),
        row.get("院校专业组代码OCR规范化", ""),
    )


def major_sort_key(row):
    judgement_order = {
        "预算内中外合作/高收费可讨论": 0,
        "预算内高学费可讨论": 1,
        "费用待核": 2,
        "超过预算": 3,
    }
    return (
        judgement_order.get(row.get("场景判断", ""), 9),
        preference_rank(row),
        row.get("院校代码", ""),
        row.get("院校专业组代码OCR规范化", ""),
        int_value(row.get("专业组内专业序号")),
    )


def append_sheet(ws, rows, fields):
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, start=1):
        max_len = min(
            46,
            max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:200]]),
        )
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "private/",
        "private\\",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
        "人工读数",
        "已核准",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"扩展预算场景公开文件包含禁止内容：{hit}")


def main():
    scenario = json.loads(SCENARIO.read_text())
    max_tuition = int(scenario["budget"]["annual_upper_limit_yuan"])
    groups = read_csv(GROUP_BROWSER)
    majors = read_csv(MAJOR_BROWSER)
    group_by_id = {row.get("专业组出现ID", ""): row for row in groups}

    major_rows = []
    by_group = defaultdict(list)
    for row in majors:
        judgement, fee_status, reason, tuition = scenario_judgement(row, max_tuition)
        if not judgement:
            continue
        out = dict(row)
        out["扩展预算场景专业ID"] = stable_id("COOPMAJOR", [row.get("专业行ID", "")])
        out["场景判断"] = judgement
        out["费用判断"] = fee_status
        out["中外合作或高收费线索"] = reason
        out["是否医学相关待确认"] = "true" if is_medical_pending(row) else "false"
        out["下一步核验动作"] = "回看第19期原页、湖北官方系统、招生章程和高校官网；核学费、培养模式、证书、校区、专业组边界和调剂范围。"
        out["是否可作为定稿依据"] = "false"
        out["_tuition_value"] = tuition
        major_rows.append(out)
        by_group[row.get("专业组出现ID", "")].append(out)

    group_rows = []
    for group_id, rows in by_group.items():
        group = group_by_id.get(group_id, {})
        tuition_values = [row["_tuition_value"] for row in rows if row["_tuition_value"] is not None]
        budget_ok = [row for row in rows if row["场景判断"].startswith("预算内")]
        over = [row for row in rows if row["场景判断"] == "超过预算"]
        pending = [row for row in rows if row["场景判断"] == "费用待核"]
        judgement = "预算内中外合作/高收费可讨论" if budget_ok else ("费用待核" if pending else "超过预算")
        prefs = sorted({row.get("专业偏好方向", "") for row in rows if row.get("专业偏好方向", "")})
        summary = []
        for row in sorted(rows, key=major_sort_key):
            summary.append(
                f"{row.get('专业代号OCR', '')} {row.get('专业名称及备注短摘', '')}"
                f"｜学费:{row.get('学费OCR候选', '')}"
                f"｜{row.get('场景判断', '')}"
            )
        out = dict(group)
        out["扩展预算场景专业组ID"] = stable_id("COOPGROUP", [group_id])
        out["场景层级"] = "7万预算中外合作/高收费专项讨论"
        out["场景判断"] = judgement
        out["预算内中外合作或高收费专业数"] = str(len(budget_ok))
        out["超过预算专业数"] = str(len(over))
        out["待核费用专业数"] = str(len(pending))
        out["场景内最低学费"] = str(min(tuition_values)) if tuition_values else ""
        out["场景内最高学费"] = str(max(tuition_values)) if tuition_values else ""
        out["场景内偏好专业方向"] = "；".join(prefs)
        out["场景内专业摘要"] = "；".join(summary)
        out["下一步核验动作"] = "先核是否真为2026湖北物理普通类计划，再核学费、培养模式、证书、校区、完整专业组和调剂范围。"
        out["是否可作为定稿依据"] = "false"
        group_rows.append(out)

    major_rows = sorted(major_rows, key=major_sort_key)
    group_rows = sorted(group_rows, key=group_sort_key)

    write_csv(MAJOR_OUTPUT, major_rows, MAJOR_FIELDS)
    write_csv(GROUP_OUTPUT, group_rows, GROUP_FIELDS)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_说明"
    ws.append(["项目", "说明"])
    for item in [
        ("用途", "7 万预算内中外合作/高收费专项讨论，不替代普通批稳定基座。"),
        ("边界", "所有行都不是定稿依据；必须回看第19期原页、湖北官方系统和招生章程。"),
        ("先看", "先看 01_专业组场景，再看 02_逐专业场景。"),
        ("注意", "医学相关仍与既有不学医底线冲突，未确认例外前不进主方案。"),
    ]:
        ws.append(item)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90
    append_sheet(wb.create_sheet("01_专业组场景"), group_rows, GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_逐专业场景"), major_rows, MAJOR_FIELDS)
    WORKBOOK_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_OUTPUT)

    summary = {
        "status": "issue19_expanded_budget_coop_scenario_ready",
        "generated_by": "export_issue19_expanded_budget_coop_scenario.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "scenario_source": str(SCENARIO.relative_to(ROOT)),
        "annual_upper_limit_yuan": max_tuition,
        "group_output": str(GROUP_OUTPUT.relative_to(ROOT)),
        "major_output": str(MAJOR_OUTPUT.relative_to(ROOT)),
        "workbook": str(WORKBOOK_OUTPUT.relative_to(ROOT)),
        "group_count": len(group_rows),
        "major_count": len(major_rows),
        "scenario_judgement_counts": dict(Counter(row["场景判断"] for row in major_rows)),
        "group_judgement_counts": dict(Counter(row["场景判断"] for row in group_rows)),
        "medical_pending_major_count": sum(row["是否医学相关待确认"] == "true" for row in major_rows),
        "usage_boundary": "只用于7万预算内中外合作/高收费专项讨论，不作为定稿依据。",
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    public_text_safe([GROUP_OUTPUT, MAJOR_OUTPUT, SUMMARY_OUTPUT])
    print(f"写出扩展预算专业组场景：{GROUP_OUTPUT.relative_to(ROOT)}")
    print(f"写出扩展预算逐专业场景：{MAJOR_OUTPUT.relative_to(ROOT)}")
    print(f"写出扩展预算工作簿：{WORKBOOK_OUTPUT.relative_to(ROOT)}")
    print(f"专业组数：{len(group_rows)}；专业数：{len(major_rows)}")


if __name__ == "__main__":
    main()
