#!/usr/bin/env python3
import csv
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

GROUP_BROWSER = EXPORTS / "issue19-stable-foundation-group-browser.csv"
MAJOR_BROWSER = EXPORTS / "issue19-stable-foundation-major-browser.csv"

ROUND1_GROUPS = EXPORTS / "issue19-round1-candidate-groups.csv"
ROUND1_SHORTLIST = EXPORTS / "issue19-round1-shortlist-groups.csv"
ROUND1_PRIORITY_CITY = EXPORTS / "issue19-round1-priority-city-watchlist.csv"
ROUND1_SHORTLIST_MAJORS = EXPORTS / "issue19-round1-shortlist-majors.csv"
ROUND1_WORKBOOK = EXPORTS / "issue19-round1-candidate-selection.xlsx"
ROUND1_SUMMARY = EXPORTS / "issue19-round1-candidate-selection-summary.json"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
PRIORITY_CITIES = {"武汉", "成都", "西安", "北京"}

GROUP_FIELDS = [
    "第一轮候选ID",
    "第一轮层级",
    "建议梯度",
    "入选理由",
    "风险提示",
    "下一步核验动作",
    "是否可作为定稿依据",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "医学护理排除专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "机器筛选价值层级",
    "保真核验层级",
    "历史线索分层",
    "同代码命中年份最大值",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "组内完整专业清单索引",
    "机器核验下一步",
    "人工核验下一步",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

MAJOR_FIELDS = [
    "第一轮候选ID",
    "第一轮层级",
    "建议梯度",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "来源页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: row.get(field, "") for field in fields} for row in rows])


def int_value(row, field):
    try:
        return int(str(row.get(field, "")).strip() or "0")
    except ValueError:
        return 0


def float_value(row, field, default=999):
    try:
        return float(str(row.get(field, "")).strip())
    except ValueError:
        return default


def history_code(row):
    return row.get("历史线索分层", "")[:2]


def history_band(row):
    code = history_code(row)
    diff = float_value(row, "历史最高等位分差")
    if code == "H0":
        return "历史待补"
    if code == "H1" or diff <= -20:
        return "保底观察"
    if code == "H2" or diff <= 0:
        return "稳妥观察"
    if code == "H3" or diff <= 15:
        return "稳冲观察"
    if code == "H4" or diff <= 30:
        return "冲刺观察"
    return "高冲暂缓"


def public_normal(row):
    return row.get("家庭底线属性动作", "") == "继续核公办普通学费-非民办线索"


def hard_ok(row):
    return (
        public_normal(row)
        and int_value(row, "医学护理排除专业数") == 0
        and int_value(row, "高收费或超预算专业数") == 0
        and int_value(row, "专业明细行数") > 0
    )


def has_preference(row):
    return int_value(row, "偏好专业数") > 0


def has_priority_city(row):
    return row.get("城市候选", "") in PRIORITY_CITIES


def stable_hash(row):
    key = "|".join(
        [
            row.get("专业组出现ID", ""),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ]
    )
    return hashlib.sha256(("issue19-round1|" + key).encode("utf-8")).hexdigest()[:16]


def risk_notes(row):
    notes = []
    if int_value(row, "特殊限制待核专业数") > 0:
        notes.append("组内有特殊限制待核")
    if "结构异常" in row.get("调剂初判", ""):
        notes.append("结构或组内边界需先核")
    if "特殊限制" in row.get("调剂初判", ""):
        notes.append("调剂前需核特殊限制")
    if history_code(row) == "H0":
        notes.append("缺同代码历史线索")
    if history_band(row) in {"冲刺观察", "高冲暂缓"}:
        notes.append("录取线距离偏冲")
    if not notes:
        notes.append("仍需完整组核验")
    return "；".join(notes)


def selection_reason(row):
    reasons = []
    if has_preference(row):
        parts = []
        if int_value(row, "数字媒体技术专业数") > 0:
            parts.append(f"数字媒体技术{int_value(row, '数字媒体技术专业数')}个")
        if int_value(row, "计算机类相关专业数") > 0:
            parts.append(f"计算机相关{int_value(row, '计算机类相关专业数')}个")
        if int_value(row, "师范类相关专业数") > 0:
            parts.append(f"师范相关{int_value(row, '师范类相关专业数')}个")
        reasons.append("命中偏好专业：" + "、".join(parts))
    if has_priority_city(row):
        reasons.append(f"命中优先城市：{row.get('城市候选')}")
    reasons.append(f"历史线索：{row.get('历史线索分层')}")
    reasons.append("家庭底线：公办/普通学费线索且无医学护理/高收费自动阻断")
    return "；".join(reasons)


def score(row):
    hist_score = {
        "H1": 38,
        "H2": 34,
        "H3": 22,
        "H4": 10,
        "H0": 6,
        "H5": -20,
    }.get(history_code(row), 0)
    value = hist_score
    value += int_value(row, "偏好专业数") * 7
    value += int_value(row, "数字媒体技术专业数") * 8
    value += int_value(row, "计算机类相关专业数") * 4
    value += int_value(row, "师范类相关专业数") * 4
    if has_priority_city(row):
        value += 8
    if row.get("调剂初判", "") == "可进入人工调剂接受度判断":
        value += 8
    if int_value(row, "特殊限制待核专业数") > 0:
        value -= 8
    if "结构异常" in row.get("调剂初判", ""):
        value -= 6
    value -= max(float_value(row, "历史最高等位分差", 0), 0) / 5
    return value


def sorted_candidates(rows):
    return sorted(
        rows,
        key=lambda row: (
            -score(row),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ),
    )


def decorate(row, layer):
    out = dict(row)
    out["第一轮候选ID"] = f"R1GROUP-{stable_hash(row)}"
    out["第一轮层级"] = layer
    out["建议梯度"] = history_band(row)
    out["入选理由"] = selection_reason(row)
    out["风险提示"] = risk_notes(row)
    out["下一步核验动作"] = "回看第19期原页和完整专业组；若进入讨论，再核湖北官方侧、招生章程、学费、计划数、选科和调剂范围。"
    out["是否可作为定稿依据"] = "false"
    return out


def take_unique(source, limit, selected_ids):
    picked = []
    for row in sorted_candidates(source):
        group_id = row.get("专业组出现ID", "")
        if group_id in selected_ids:
            continue
        picked.append(row)
        selected_ids.add(group_id)
        if len(picked) >= limit:
            break
    return picked


def build_selection(group_rows):
    base = [row for row in group_rows if hard_ok(row) and (has_preference(row) or has_priority_city(row))]
    main = [row for row in base if history_code(row) in {"H1", "H2", "H3", "H4"}]
    history_missing = [row for row in base if history_code(row) == "H0"]
    city_high = [row for row in base if has_priority_city(row) and history_code(row) == "H5"]

    selected_ids = set()
    shortlist = []
    quota_plan = [
        ("S1-保底稳妥优先讨论", [row for row in main if history_code(row) in {"H1", "H2"}], 35),
        ("S2-稳冲优先讨论", [row for row in main if history_code(row) == "H3"], 25),
        ("S3-冲刺保留讨论", [row for row in main if history_code(row) == "H4"], 15),
        ("S4-城市或历史待补观察", history_missing + city_high, 5),
    ]
    for layer, rows, limit in quota_plan:
        shortlist.extend(decorate(row, layer) for row in take_unique(rows, limit, selected_ids))

    all_ids = set()
    all_candidates = []
    for layer, rows in [
        ("C1-偏好/城市且历史H1-H4", main),
        ("C2-偏好/城市但历史待补", history_missing),
        ("C3-优先城市高冲暂缓", city_high),
    ]:
        for row in sorted_candidates(rows):
            group_id = row.get("专业组出现ID", "")
            if group_id in all_ids:
                continue
            all_candidates.append(decorate(row, layer))
            all_ids.add(group_id)

    shortlist_ids = {row.get("专业组出现ID", "") for row in shortlist}
    for row in all_candidates:
        if row.get("专业组出现ID", "") in shortlist_ids:
            row["第一轮层级"] = row["第一轮层级"] + "；已进精选80"

    return all_candidates, shortlist


def build_shortlist_majors(shortlist, major_rows):
    group_meta = {row.get("专业组出现ID", ""): row for row in shortlist}
    rows = []
    for row in major_rows:
        meta = group_meta.get(row.get("专业组出现ID", ""))
        if not meta:
            continue
        out = dict(row)
        out["第一轮候选ID"] = meta.get("第一轮候选ID", "")
        out["第一轮层级"] = meta.get("第一轮层级", "")
        out["建议梯度"] = meta.get("建议梯度", "")
        rows.append(out)
    return sorted(
        rows,
        key=lambda row: (
            row.get("第一轮候选ID", ""),
            int_value(row, "专业组内专业序号"),
            row.get("专业代号OCR", ""),
        ),
    )


def build_priority_city_watchlist(all_candidates):
    return sorted(
        [row for row in all_candidates if row.get("城市候选", "") in PRIORITY_CITIES],
        key=lambda row: (
            row.get("城市候选", ""),
            {
                "稳妥观察": 0,
                "保底观察": 1,
                "稳冲观察": 2,
                "冲刺观察": 3,
                "历史待补": 4,
                "高冲暂缓": 5,
            }.get(row.get("建议梯度", ""), 9),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ),
    )


def append_sheet(ws, rows, fields):
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, start=1):
        max_len = min(44, max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:200]]))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(all_candidates, shortlist, city_watchlist, majors):
    wb = Workbook()
    wb.remove(wb.active)
    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "第一轮候选专业组池，只用于压缩讨论范围，不作为定稿依据。"),
        ("先看", "先看 01_精选80专业组，再看 03_精选组内专业明细。"),
        ("怎么用", "家庭先判断完整专业组是否能接受调剂；不能只看想填的6个专业。"),
        ("下一步", "对愿意讨论的组做第19期原页、湖北官方侧、招生章程、计划数、学费、选科和备注限制核验。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 100
    append_sheet(wb.create_sheet("01_精选80专业组"), shortlist, GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_完整候选池"), all_candidates, GROUP_FIELDS)
    append_sheet(wb.create_sheet("03_优先城市观察"), city_watchlist, GROUP_FIELDS)
    append_sheet(wb.create_sheet("04_精选组内专业明细"), majors, MAJOR_FIELDS)
    ROUND1_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ROUND1_WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "private/",
        "private\\",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
        "已核准",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"第一轮候选公开文件包含禁止内容：{hit}")


def main():
    group_rows = read_csv(GROUP_BROWSER)
    major_rows = read_csv(MAJOR_BROWSER)
    all_candidates, shortlist = build_selection(group_rows)
    city_watchlist = build_priority_city_watchlist(all_candidates)
    shortlist_major_rows = build_shortlist_majors(shortlist, major_rows)

    write_csv(ROUND1_GROUPS, all_candidates, GROUP_FIELDS)
    write_csv(ROUND1_SHORTLIST, shortlist, GROUP_FIELDS)
    write_csv(ROUND1_PRIORITY_CITY, city_watchlist, GROUP_FIELDS)
    write_csv(ROUND1_SHORTLIST_MAJORS, shortlist_major_rows, MAJOR_FIELDS)
    write_workbook(all_candidates, shortlist, city_watchlist, shortlist_major_rows)

    summary = {
        "status": "issue19_round1_candidate_selection_ready",
        "generated_by": "build_issue19_round1_candidate_selection.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "selection_boundary": "第一轮候选专业组池，只用于家庭讨论和入围组核验，不作为定稿依据。",
        "candidate_group_rows": len(all_candidates),
        "shortlist_group_rows": len(shortlist),
        "priority_city_watchlist_rows": len(city_watchlist),
        "shortlist_major_rows": len(shortlist_major_rows),
        "workbook": str(ROUND1_WORKBOOK.relative_to(ROOT)),
        "candidate_groups_csv": str(ROUND1_GROUPS.relative_to(ROOT)),
        "shortlist_groups_csv": str(ROUND1_SHORTLIST.relative_to(ROOT)),
        "priority_city_watchlist_csv": str(ROUND1_PRIORITY_CITY.relative_to(ROOT)),
        "shortlist_majors_csv": str(ROUND1_SHORTLIST_MAJORS.relative_to(ROOT)),
        "priority_cities": sorted(PRIORITY_CITIES),
        "hard_filters": [
            "公办/普通学费机器线索",
            "组内医学护理排除专业数为0",
            "组内高收费或超预算专业数为0",
            "必须有逐专业明细",
        ],
        "shortlist_quota": {
            "S1_H1_H2": 35,
            "S2_H3": 25,
            "S3_H4": 15,
            "S4_city_or_history_missing": 5,
        },
    }
    ROUND1_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe([ROUND1_GROUPS, ROUND1_SHORTLIST, ROUND1_PRIORITY_CITY, ROUND1_SHORTLIST_MAJORS, ROUND1_SUMMARY])
    print(f"写出第一轮候选池：{ROUND1_GROUPS.relative_to(ROOT)}")
    print(f"写出第一轮精选专业组：{ROUND1_SHORTLIST.relative_to(ROOT)}")
    print(f"写出第一轮优先城市观察：{ROUND1_PRIORITY_CITY.relative_to(ROOT)}")
    print(f"写出第一轮精选组内专业明细：{ROUND1_SHORTLIST_MAJORS.relative_to(ROOT)}")
    print(f"写出第一轮候选工作簿：{ROUND1_WORKBOOK.relative_to(ROOT)}")
    print(
        f"候选组={len(all_candidates)}；精选组={len(shortlist)}；"
        f"优先城市观察={len(city_watchlist)}；精选组内专业明细={len(shortlist_major_rows)}"
    )


if __name__ == "__main__":
    main()
