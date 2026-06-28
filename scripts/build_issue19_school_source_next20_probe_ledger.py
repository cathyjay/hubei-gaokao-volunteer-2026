#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
NEXT20 = ROOT / "data/exports/issue19-data-foundation-next-execution-v1-school-source-next20.csv"
LIVE_LEDGER = ROOT / "data/working/issue19-school-source-live-20260629-ledger.csv"
RETAINED_REUSE_LEDGER = ROOT / "data/working/issue19-c4-c6-retained-source-reuse-public-ledger.csv"
ZHINENGDAYI_LEDGER = ROOT / "data/working/issue19-next20-zhinengdayi-official-source-fetch-public-ledger.csv"

OUTPUT_CSV = ROOT / "data/working/issue19-school-source-next20-official-probe-public-ledger.csv"
SUMMARY_OUTPUT = ROOT / "data/working/issue19-school-source-next20-official-probe-summary.json"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "是否可作为定稿依据",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

PUBLIC_HIDDEN_FILE_SUFFIXES = (".jpg", ".jpeg", ".png")

OUTPUT_FIELDS = [
    "next20官网源探测ID",
    "来源next20执行表",
    "来源live补源账本",
    "来源C4C6复用审计账本",
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "最终可用",
    "可进入下一阶段",
    "是否可作为定稿依据",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
    "执行建议序号",
    "院校代码",
    "院校名称",
    "机会优先级",
    "机会类型",
    "官网辅证自动动作",
    "涉及招生明细数",
    "涉及专业组数",
    "官网源探测层级",
    "官网源类型",
    "官方URL",
    "本地公开来源文件",
    "本地公开来源文件集合SHA256",
    "结构化湖北相关行数",
    "结构化湖北物理行数",
    "结构化计划数合计",
    "可核字段公开摘要",
    "字段局限公开摘要",
    "已有C4C6计划数一致候选数",
    "已有C4C6官网可补OCR计划数候选数",
    "已有C4C6计划数冲突候选数",
    "已有live补源结论",
    "已有live湖北物理计划状态",
    "当前自动判断",
    "下一步自动动作",
    "人工最小核验动作",
    "PDF原页核页状态",
    "湖北官方系统或省招办计划核验状态",
    "高校官网源刷新状态",
    "字段事实写回状态",
    "公开安全策略",
]


def rel(path):
    return str(path.relative_to(ROOT))


def sha_text(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha_paths(paths):
    text = "；".join(sorted(path for path in paths if path))
    return sha_text(text)


def public_file_refs(paths):
    visible = [
        path for path in paths
        if not str(path).lower().endswith(PUBLIC_HIDDEN_FILE_SUFFIXES)
    ]
    hidden_count = len(paths) - len(visible)
    if hidden_count:
        visible.append(f"图片类证据文件已留存但公开账本隐藏路径×{hidden_count}")
    return "；".join(visible)


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def plan_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return 0


def json_rows(path):
    obj = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(obj, dict) and isinstance(obj.get("data"), dict):
        rows = obj["data"].get("zsjhList") or obj["data"].get("list") or []
    elif isinstance(obj, dict):
        rows = obj.get("list") or []
    else:
        rows = []
    return rows


def json_metric(paths):
    total_rows = 0
    plan_sum = 0
    for path in paths:
        rows = json_rows(ROOT / path)
        total_rows += len(rows)
        for row in rows:
            plan_sum += plan_int(row.get("jhrs") or row.get("zsjhs"))
    return total_rows, plan_sum


def zhinengdayi_metric(path):
    obj = json.loads((ROOT / path).read_text(encoding="utf-8"))
    rows = obj.get("data", []) if isinstance(obj, dict) else []
    plan_sum = sum(plan_int(row.get("enrollNum")) for row in rows)
    return len(rows), plan_sum


def csv_metric(path, plan_fields):
    rows = read_csv(ROOT / path)
    plan_sum = 0
    for row in rows:
        for field in plan_fields:
            if row.get(field):
                plan_sum += plan_int(row.get(field))
                break
    return len(rows), plan_sum


def xlsx_metric_hdu(path):
    try:
        import openpyxl
    except ImportError:
        return 0, 0
    wb = openpyxl.load_workbook(ROOT / path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    total = 0
    for row in rows[1:]:
        if len(row) >= 12 and row[0] == "湖北" and row[5] == "物理类":
            count += 1
            total += plan_int(row[11])
    return count, total


def xlsx_metric_ksu(path):
    try:
        import openpyxl
    except ImportError:
        return 0, 0
    wb = openpyxl.load_workbook(ROOT / path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    hubei_index = headers.index("湖北")
    count = 0
    total = 0
    for row in rows[3:]:
        value = row[hubei_index]
        if isinstance(value, (int, float)) and value:
            count += 1
            total += int(value)
    return count, total


def source_catalog():
    whpu_rows, whpu_sum = zhinengdayi_metric(
        "data/external/issue19-next20-official-sources/whpu-2026-hubei-physics-normal.json"
    )
    hbnu_rows, hbnu_sum = zhinengdayi_metric(
        "data/external/issue19-next20-official-sources/hbnu-2026-hubei-physics-group-normal.json"
    )
    xupt_rows, xupt_sum = json_metric(
        [
            "data/external/issue19-b0-b1-official-sources/xupt-2026-hubei-physics-normal.json",
            "data/external/issue19-b0-b1-official-sources/xupt-2026-hubei-physics-sino.json",
        ]
    )
    hdu_rows, hdu_sum = xlsx_metric_hdu(
        "data/external/issue19-b0-b1-official-sources/hdu-2026-hubei-plan.xlsx"
    )
    ksu_rows, ksu_sum = xlsx_metric_ksu(
        "data/external/issue19-b0-b1-official-sources/ksu-2026-undergraduate-plan.xlsx"
    )
    return {
        "C125": {
            "level": "S0-已有官方智能答疑API湖北物理结构化源",
            "kind": "official_zhinengdayi_json",
            "url": "https://zsb.whpu.edu.cn/whpu；https://admin.zhinengdayi.com/front/enroll/findEnrollPlanDetail?sCode=VFUSRX&cityName=湖北&year=2026&enrollType=普通文理&science=物理类&batch=本科普通批",
            "files": [
                "data/external/issue19-next20-official-sources/whpu-2026-hubei-physics-normal.json",
                "data/external/issue19-school-source-live-20260629/whpu-zsb-app-shell.html",
                "data/external/issue19-school-source-live-20260629/whpu-2026-plan-entry-shell.html",
                "data/external/issue19-school-source-live-20260629/whpu-2026-hubei-fill-suggestion-shell.html",
            ],
            "hubei_rows": whpu_rows,
            "physics_rows": whpu_rows,
            "plan_sum": whpu_sum,
            "fields": "年份；省份；科类；批次；专业名；计划数；选科备注",
            "limits": "官方招生系统API未给湖北院校专业组代码、专业代号、学费或校区；仍需第19期原页和湖北官方侧。",
        },
        "F582": {
            "level": "S3-仅外省计划，未取得湖北计划",
            "kind": "official_html_jilin_only",
            "url": "https://bzkzs.ccut.edu.cn/zsxx/bzkzsxx/zsjh.htm；https://bzkzs.ccut.edu.cn/info/1073/1491.htm；https://bzkzs.ccut.edu.cn/info/1019/1492.htm",
            "files": [
                "data/external/issue19-school-source-live-20260629/ccut-zsjh-list.html",
                "data/external/issue19-school-source-live-20260629/ccut-2026-jilin-plan.html",
                "data/external/issue19-school-source-live-20260629/ccut-2026-undergraduate-guide.html",
            ],
            "hubei_rows": 0,
            "physics_rows": 0,
            "plan_sum": 0,
            "fields": "可见招生简章和吉林省计划入口",
            "limits": "不能用吉林省计划迁移到湖北；湖北物理计划仍缺。",
        },
        "C133": {
            "level": "S0-已有官方智能答疑API湖北物理结构化源",
            "kind": "official_zhinengdayi_json",
            "url": "https://zhinengdayi.com/hbnu；https://admin.zhinengdayi.com/front/enroll/findEnrollPlanDetail?sCode=LSPJPH&cityName=湖北&year=2026&enrollType=普通文理&science=物理组&batch=-",
            "files": [
                "data/external/issue19-next20-official-sources/hbnu-2026-hubei-physics-group-normal.json",
                "data/external/issue19-school-source-live-20260629/hbnu-zsb-logout-shell.html",
                "data/external/issue19-school-source-live-20260629/hbnu-2026-undergraduate-charter.html",
            ],
            "hubei_rows": hbnu_rows,
            "physics_rows": hbnu_rows,
            "plan_sum": hbnu_sum,
            "fields": "年份；省份；科类；专业名；计划数；学费备注",
            "limits": "官方招生系统API批次字段为 '-'，未给湖北院校专业组代码、专业代号、选科或校区；仍需第19期原页和湖北官方侧。",
        },
        "C108": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_zsdata_json",
            "url": "https://zsdata.jhun.edu.cn/zsdata/lqxx/#/",
            "files": [
                "data/external/issue19-b0-b1-official-sources/jhun-2026-hubei-physics-normal-lqxx2.json",
                "data/external/issue19-b0-b1-official-sources/jhun-zsdata-index.html",
            ],
            "hubei_rows": 51,
            "physics_rows": 51,
            "plan_sum": 1495,
            "fields": "年份；省份；科类；批次；专业名；计划数；学制；校区；备注",
            "limits": "高校API未给湖北院校专业组代码边界；仍需第19期原页和湖北官方侧对齐。",
        },
        "H026": {
            "level": "S3-仅入口和信息公开页，未取得结构化计划",
            "kind": "official_entry_only",
            "url": "https://zsw.cuz.edu.cn/；https://xxgk.cuz.edu.cn/xxgkml/zkxx.htm",
            "files": [
                "data/external/issue19-school-source-live-20260629/cuz-zsw-home-20260629.html",
                "data/external/issue19-school-source-live-20260629/cuz-xxgk-zkxx-20260629.html",
            ],
            "hubei_rows": 0,
            "physics_rows": 0,
            "plan_sum": 0,
            "fields": "仅有招生入口和信息公开入口",
            "limits": "当前未取得2026湖北物理专业和人数明细。",
        },
        "A197": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_zsdata_json",
            "url": "学校官方招生系统接口留存",
            "files": ["data/external/issue19-b0-b1-official-sources/xbmu-2026-hubei-physics-normal-zsjh.json"],
            "hubei_rows": 30,
            "physics_rows": 30,
            "plan_sum": 95,
            "fields": "年份；省份；科类；专业名；计划数；校区",
            "limits": "未给湖北院校专业组代码边界；需核组边界、限制条款和第19期原页。",
        },
        "K486": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_zsdata_json",
            "url": "http://zsdata.xupt.edu.cn/zsdata/lqxx/#/",
            "files": [
                "data/external/issue19-b0-b1-official-sources/xupt-2026-hubei-physics-normal.json",
                "data/external/issue19-b0-b1-official-sources/xupt-2026-hubei-physics-sino.json",
                "data/external/issue19-b0-b1-official-sources/xupt-zsjh-type.json",
            ],
            "hubei_rows": xupt_rows,
            "physics_rows": xupt_rows,
            "plan_sum": xupt_sum,
            "fields": "年份；省份；科类；批次；类别；专业名；专业代号；计划数；选科",
            "limits": "高校API未给湖北专业组完整边界；普通类和中外合作需分开核。",
        },
        "H945": {
            "level": "S0-已有官网静态表结构化线索",
            "kind": "official_static_html_table",
            "url": "https://zs.unn.edu.cn/bkzn/zsjh.htm",
            "files": [
                "data/external/issue19-b0-b1-official-sources/unn-2026-undergraduate-plan-page.html",
                "data/external/issue19-b0-b1-official-sources/unn-home.html",
            ],
            "hubei_rows": 20,
            "physics_rows": 20,
            "plan_sum": 0,
            "fields": "专业名；湖北计划线索；学费；选科；学院；页面专业组编号",
            "limits": "页面专业组编号不是湖北志愿系统专业组代码；计划数冲突需回原页和湖北官方侧。",
        },
        "K465": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_zsdata_json",
            "url": "https://zsdata.xauat.edu.cn/lqxx/s/api/front/lqxx/getList",
            "files": ["data/external/issue19-c4-c6-official-sources/xauat-2026-hubei-physics-normal.json"],
            "hubei_rows": 20,
            "physics_rows": 20,
            "plan_sum": 80,
            "fields": "年份；省份；科类；类别；专业名；计划数；校内专业组；选科",
            "limits": "高校API未给湖北院校专业组代码、专业代号、学费和校区。",
        },
        "K753": {
            "level": "S1-已有官方XLSX湖北列，科类边界待核",
            "kind": "official_xlsx_hubei_column",
            "url": "学校官方招生计划附件留存",
            "files": [
                "data/external/issue19-b0-b1-official-sources/ksu-2026-undergraduate-plan.xlsx",
                "data/external/issue19-b0-b1-official-sources/ksu-2026-undergraduate-plan-page.html",
            ],
            "hubei_rows": ksu_rows,
            "physics_rows": 0,
            "plan_sum": ksu_sum,
            "fields": "专业名；湖北列计划数",
            "limits": "XLSX为省份列计划，未在表内直接拆湖北物理/历史；不能单独核专业组边界。",
        },
        "H450": {
            "level": "S0-已有官网PDF抽取湖北物理结构化源",
            "kind": "official_pdf_extracted",
            "url": "https://zs.sdtbu.edu.cn/info/2018/6284.htm",
            "files": [
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-plan-page.html",
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-province-major-plan.pdf",
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-hubei-physics-plan-extracted.csv",
            ],
            "hubei_rows": csv_metric(
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[0],
            "physics_rows": csv_metric(
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[0],
            "plan_sum": csv_metric(
                "data/external/issue19-b0-b1-official-sources/sdtbu-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[1],
            "fields": "专业名；科类；湖北计划数；总计划数",
            "limits": "由PDF网格和OCR联合抽取；未给湖北院校专业组代码、专业代号和校区。",
        },
        "A032": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_ajax_json",
            "url": "http://lqcx.blcu.edu.cn/f/ajax_zsjh",
            "files": ["data/external/issue19-c4-c6-official-sources/blcu-2026-hubei-physics-normal.json"],
            "hubei_rows": 14,
            "physics_rows": 14,
            "plan_sum": 34,
            "fields": "年份；省份；科类；批次；专业代号；专业名；计划数；学制；学费；校区；选科",
            "limits": "高校API未给湖北院校专业组代码；仍需对齐省招办组边界。",
        },
        "F902": {
            "level": "S0-已有官方图片转录湖北物理结构化源",
            "kind": "official_image_extracted",
            "url": "http://zs.jstu.edu.cn/5639/list.htm；https://mp.weixin.qq.com/s/-GH34DZnqSgWapAiTmokfg",
            "files": [
                "data/external/issue19-b0-b1-official-sources/jsut-zs-home-2026-plan-links.html",
                "data/external/issue19-b0-b1-official-sources/jsut-2026-hubei-plan.jpg",
                "data/external/issue19-b0-b1-official-sources/jsut-2026-hubei-physics-plan-extracted.csv",
            ],
            "hubei_rows": csv_metric(
                "data/external/issue19-b0-b1-official-sources/jsut-2026-hubei-physics-plan-extracted.csv",
                ["计划数"],
            )[0],
            "physics_rows": csv_metric(
                "data/external/issue19-b0-b1-official-sources/jsut-2026-hubei-physics-plan-extracted.csv",
                ["计划数"],
            )[0],
            "plan_sum": csv_metric(
                "data/external/issue19-b0-b1-official-sources/jsut-2026-hubei-physics-plan-extracted.csv",
                ["计划数"],
            )[1],
            "fields": "专业名；科类；学制；学费；计划数",
            "limits": "图片转录未给湖北院校专业组代码和专业代号；冲突包需双人核页。",
        },
        "K179": {
            "level": "S3-仅招生工作入口，未取得结构化计划",
            "kind": "official_entry_only",
            "url": "https://www.cdnu.edu.cn/zjc/zsgz.htm；https://www.cdnu.edu.cn/",
            "files": ["data/external/issue19-school-source-live-20260629/cdnu-zsgz-entry-20260629.html"],
            "hubei_rows": 0,
            "physics_rows": 0,
            "plan_sum": 0,
            "fields": "仅有招生工作入口",
            "limits": "未取得2026湖北物理专业和人数明细。",
        },
        "K487": {
            "level": "S0-已有官网PDF抽取湖北物理结构化源",
            "kind": "official_pdf_extracted",
            "url": "https://zb.xaau.edu.cn/info/1039/2831.htm",
            "files": [
                "data/external/issue19-school-source-live-20260629/xaau-2026-undergraduate-plan-page.html",
                "data/external/issue19-school-source-live-20260629/xaau-2026-province-major-plan.pdf",
                "data/external/issue19-school-source-live-20260629/xaau-2026-hubei-physics-plan-extracted.csv",
            ],
            "hubei_rows": csv_metric(
                "data/external/issue19-school-source-live-20260629/xaau-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[0],
            "physics_rows": csv_metric(
                "data/external/issue19-school-source-live-20260629/xaau-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[0],
            "plan_sum": csv_metric(
                "data/external/issue19-school-source-live-20260629/xaau-2026-hubei-physics-plan-extracted.csv",
                ["湖北计划数"],
            )[1],
            "fields": "专业代号；专业名；科类；湖北计划数；官网总计划数",
            "limits": "第19期K487存在OCR串读线索；必须回原页和湖北官方侧。",
        },
        "H775": {
            "level": "S4-官网访问失败，未取得结构化计划",
            "kind": "official_site_unreachable",
            "url": "https://zs.sgu.edu.cn/",
            "files": [],
            "hubei_rows": 0,
            "physics_rows": 0,
            "plan_sum": 0,
            "fields": "无",
            "limits": "招生域名HTTPS访问失败；需换网络、人工打开或回第19期原页和湖北官方侧。",
        },
        "A033": {
            "level": "S0-已有官方API湖北物理结构化源",
            "kind": "official_ajax_json",
            "url": "https://zszx.cuc.edu.cn/f/ajax_zsjh",
            "files": ["data/external/issue19-b0-b1-official-sources/cuc-2026-hubei-physics-normal-zsjh.json"],
            "hubei_rows": 14,
            "physics_rows": 14,
            "plan_sum": 34,
            "fields": "年份；省份；科类；批次；专业组线索；专业代号；专业名；计划数；学制；学费；校区；选科",
            "limits": "高校API给校内组名线索，但仍需湖北院校专业组代码和第19期原页闭环。",
        },
        "H001": {
            "level": "S0-已有官方XLSX湖北物理结构化源",
            "kind": "official_xlsx_physics_rows",
            "url": "杭州电子科技大学官网2026湖北计划页留存",
            "files": [
                "data/external/issue19-b0-b1-official-sources/hdu-2026-hubei-plan-page.html",
                "data/external/issue19-b0-b1-official-sources/hdu-2026-hubei-plan.xlsx",
            ],
            "hubei_rows": hdu_rows,
            "physics_rows": hdu_rows,
            "plan_sum": hdu_sum,
            "fields": "省份；专业名；学制；科类；批次；选科；计划数；就学地点",
            "limits": "XLSX未给湖北院校专业组代码；仍需第19期原页和湖北官方侧。",
        },
    }


def live_by_code():
    if not LIVE_LEDGER.exists():
        return {}
    rows = read_csv(LIVE_LEDGER)
    result = {}
    for row in rows:
        result.setdefault(row.get("院校代码", ""), row)
    return result


def reuse_by_code():
    if not RETAINED_REUSE_LEDGER.exists():
        return {}
    result = {}
    for row in read_csv(RETAINED_REUSE_LEDGER):
        code = row.get("院校代码", "")
        result.setdefault(code, []).append(row)
    return result


def sum_reuse(rows, field):
    return sum(plan_int(row.get(field)) for row in rows)


def build_rows():
    catalog = source_catalog()
    live = live_by_code()
    reuse = reuse_by_code()
    out = []
    for next_row in read_csv(NEXT20):
        code = next_row["院校代码"]
        info = catalog.get(
            code,
            {
                "level": "S9-未纳入本轮探测目录",
                "kind": "unknown",
                "url": "",
                "files": [],
                "hubei_rows": 0,
                "physics_rows": 0,
                "plan_sum": 0,
                "fields": "待补",
                "limits": "待补官方来源。",
            },
        )
        files = [path for path in info.get("files", []) if path]
        live_row = live.get(code, {})
        reuse_rows = reuse.get(code, [])
        current = (
            "已有结构化高校侧辅证，可进入候选diff和核页优先级判断"
            if str(info["level"]).startswith("S0")
            else "已有官方入口或附件线索，但仍需继续补结构化或补源"
            if str(info["level"]).startswith("S1")
            else "仍需继续找2026湖北物理类分省分专业计划源"
        )
        out.append(
            {
                "next20官网源探测ID": f"NEXT20PROBE-{int(next_row['执行建议序号']):02d}-{code}",
                "来源next20执行表": rel(NEXT20),
                "来源live补源账本": rel(LIVE_LEDGER),
                "来源C4C6复用审计账本": rel(RETAINED_REUSE_LEDGER),
                "来源期号": SOURCE_ISSUE,
                "来源PDF_SHA256": SOURCE_PDF_SHA256,
                "生成日期": "2026-06-29",
                "数据阶段": "issue19_school_source_next20_official_probe_public_ledger",
                "主表粒度": "next20高校侧辅证任务×官方源探测状态",
                **{field: "false" for field in FALSE_FIELDS},
                "执行建议序号": next_row["执行建议序号"],
                "院校代码": code,
                "院校名称": next_row["院校名称"],
                "机会优先级": next_row["机会优先级"],
                "机会类型": next_row["机会类型"],
                "官网辅证自动动作": next_row["官网辅证自动动作"],
                "涉及招生明细数": next_row["涉及招生明细数"],
                "涉及专业组数": next_row["涉及专业组数"],
                "官网源探测层级": info["level"],
                "官网源类型": info["kind"],
                "官方URL": info["url"],
                "本地公开来源文件": public_file_refs(files),
                "本地公开来源文件集合SHA256": sha_paths(files),
                "结构化湖北相关行数": str(info["hubei_rows"]),
                "结构化湖北物理行数": str(info["physics_rows"]),
                "结构化计划数合计": str(info["plan_sum"]),
                "可核字段公开摘要": info["fields"],
                "字段局限公开摘要": info["limits"],
                "已有C4C6计划数一致候选数": str(sum_reuse(reuse_rows, "计划数一致候选数")),
                "已有C4C6官网可补OCR计划数候选数": str(sum_reuse(reuse_rows, "官网可补OCR计划数候选数")),
                "已有C4C6计划数冲突候选数": str(sum_reuse(reuse_rows, "计划数冲突候选数")),
                "已有live补源结论": live_row.get("自动补源结论", ""),
                "已有live湖北物理计划状态": live_row.get("湖北物理计划状态", ""),
                "当前自动判断": current,
                "下一步自动动作": (
                    "把高校侧结构化源接入候选diff；冲突和补缺只作为核页提示。"
                    if str(info["level"]).startswith("S0")
                    else "继续解析官方入口、XLSX/PDF/API或图片；不能迁移外省计划。"
                ),
                "人工最小核验动作": (
                    "若进入候选组，逐组核第19期原页、湖北官方侧、招生章程、完整调剂范围。"
                ),
                "PDF原页核页状态": "pending_manual_pdf_review",
                "湖北官方系统或省招办计划核验状态": "pending_hubei_official_review",
                "高校官网源刷新状态": "school_source_probe_not_verified",
                "字段事实写回状态": "blocked_until_pdf_hubei_official_review",
                "公开安全策略": "公开层只保存学校级URL、公开来源相对路径、计数、状态和门禁；不保存逐专业字段值、截图、OCR原文、登录态或复核正文。",
            }
        )
    return out


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main():
    rows = build_rows()
    write_csv(OUTPUT_CSV, rows)
    unique_codes = sorted({row["院校代码"] for row in rows})
    level_counter = Counter(row["官网源探测层级"] for row in rows)
    unique_level_counter = Counter()
    seen = set()
    for row in rows:
        code = row["院校代码"]
        if code in seen:
            continue
        seen.add(code)
        unique_level_counter[row["官网源探测层级"]] += 1
    summary = {
        "status": "issue19_school_source_next20_official_probe_not_final",
        "generated_by": "build_issue19_school_source_next20_probe_ledger.py",
        "source_next20": rel(NEXT20),
        "source_live_ledger": rel(LIVE_LEDGER),
        "source_c4c6_reuse_ledger": rel(RETAINED_REUSE_LEDGER),
        "source_zhinengdayi_ledger": rel(ZHINENGDAYI_LEDGER),
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "generated_at": "2026-06-29",
        "output_csv": rel(OUTPUT_CSV),
        "next20_action_row_count": len(rows),
        "unique_school_count": len(unique_codes),
        "structured_action_row_count": sum(1 for row in rows if row["官网源探测层级"].startswith("S0")),
        "structured_unique_school_count": sum(
            1
            for code in unique_codes
            if any(row["院校代码"] == code and row["官网源探测层级"].startswith("S0") for row in rows)
        ),
        "needs_source_unique_school_count": sum(
            1
            for code in unique_codes
            if any(row["院校代码"] == code and row["官网源探测层级"].startswith(("S2", "S3", "S4")) for row in rows)
        ),
        "level_counts_by_action_row": dict(level_counter),
        "level_counts_by_unique_school": dict(unique_level_counter),
        "total_structured_hubei_rows_by_action": sum(plan_int(row["结构化湖北相关行数"]) for row in rows),
        "total_structured_physics_rows_by_action": sum(plan_int(row["结构化湖北物理行数"]) for row in rows),
        "total_plan_sum_by_action": sum(plan_int(row["结构化计划数合计"]) for row in rows),
        "field_writeback_allowed_count": 0,
        "recommendation_basis_allowed_count": 0,
        "final_available_count": 0,
        "policy": "本账本只汇总高校侧辅证探测状态；高校官网/API/PDF/XLSX/图片不能替代第19期原页、湖北官方系统或省招办计划。",
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    public_text = OUTPUT_CSV.read_text(encoding="utf-8", errors="ignore") + SUMMARY_OUTPUT.read_text(
        encoding="utf-8", errors="ignore"
    )
    forbidden = [
        "/Users/",
        "/private/",
        "private/",
        "Authorization",
        "Bearer ",
        "Cookie",
        "access_token",
        "password",
        "secret",
        "人工读数",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    if any(token in public_text for token in forbidden):
        raise SystemExit("next20官网源探测公开账本包含禁止公开内容")
    print(f"写出 next20 官网源探测公开账本：{rel(OUTPUT_CSV)}")
    print(f"写出摘要：{rel(SUMMARY_OUTPUT)}")
    print(f"next20任务行：{len(rows)}，学校数：{len(unique_codes)}")
    print(f"已有结构化高校侧辅证任务行：{summary['structured_action_row_count']}")


if __name__ == "__main__":
    main()
