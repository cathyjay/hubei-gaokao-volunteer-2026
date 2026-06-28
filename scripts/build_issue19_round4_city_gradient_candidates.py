#!/usr/bin/env python3
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from build_issue19_round2_updated_preferences import (
    GROUP_BROWSER,
    MAJOR_BROWSER,
    ROOT,
    SOURCE_PDF_SHA256,
    aggregate_major_signals,
    append_sheet,
    count,
    direction_tags,
    float_value,
    has_clinical_pause,
    has_health_or_vet_special,
    has_main_direction,
    has_nursing_or_animal_block,
    history_band,
    history_code,
    int_value,
    main_ok,
    major_role,
    merge_signals,
    public_normal,
    public_value,
    read_csv,
    stable_hash,
)


EXPORTS = ROOT / "data/exports"
WORKING = ROOT / "data/working"

MOE_ATTRIBUTES = WORKING / "issue19-moe-school-attribute-major-detail.csv"
HISTORICAL_LINES = WORKING / "issue19-major-line-historical-toudang-sidecar.csv"

ROUND4_GROUPS = EXPORTS / "issue19-round4-city-gradient-candidate-groups.csv"
ROUND4_PRIORITY = EXPORTS / "issue19-round4-city-gradient-priority120-groups.csv"
ROUND4_MAJORS = EXPORTS / "issue19-round4-city-gradient-major-details.csv"
ROUND4_CITY_SUMMARY = EXPORTS / "issue19-round4-city-gradient-city-summary.csv"
ROUND4_HISTORY_MISSING = EXPORTS / "issue19-round4-city-gradient-history-missing-groups.csv"
ROUND4_HIGH_RUSH_PAUSED = EXPORTS / "issue19-round4-city-gradient-high-rush-paused-groups.csv"
ROUND4_WORKBOOK = EXPORTS / "issue19-round4-city-gradient.xlsx"
ROUND4_SUMMARY = EXPORTS / "issue19-round4-city-gradient-summary.json"

SOURCE_POLICY = "直接来源于稳定底座、教育部学校属性表和三年投档线旁挂表；不引用 Round3 输出。"
CITY_POLICY = "城市按教育部学校所在地展示，仅作讨论维度；实际校区和办学地点必须再核第19期原页、湖北官方侧和招生章程。"
REACHABLE_POLICY = "主表纳入 H1/H2/H3/H4；H0 历史待补和 H5 高冲暂缓均单列索引，不混入可能够得着主表。"
FAMILY_POLICY = "优先公办本科、普通学费，普通主线年学费上限15000；护理/助产、动物医学/兽医/动物科学、临床/口腔/中医、医技/康复不进入普通主线。"

EQUIVALENT_SCORES = {"2025": "494", "2024": "497", "2023": "481"}
CANDIDATE_SCORE = "515"
CANDIDATE_RANK = "91723"

GROUP_FIELDS = [
    "第四轮候选ID",
    "第四轮来源口径",
    "讨论集合",
    "是否进入优先讨论",
    "是否可作为定稿依据",
    "城市",
    "城市来源",
    "教育部所在地",
    "教育部所在地命中城市",
    "原城市候选",
    "城市口径提醒",
    "所在地使用边界",
    "城市字段状态",
    "校区字段状态",
    "省份",
    "主管部门",
    "学校登记备注",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组号",
    "历史线索分层",
    "冲稳保",
    "冲稳保解释",
    "与515关系",
    "2026考生分数",
    "2026考生位次",
    "2025等位分",
    "2024等位分",
    "2023等位分",
    "2025投档分",
    "2025等位分差",
    "2024投档分",
    "2024等位分差",
    "2023投档分",
    "2023等位分差",
    "同代码命中年份数",
    "同代码命中年份",
    "三年投档分范围",
    "等位分差摘要",
    "历史线使用口径",
    "再选要求跨年变化",
    "历史组名疑似不一致",
    "公办民办机器线索",
    "家庭底线属性动作",
    "专业明细行数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "计算机AI软件专业数",
    "电子信息网络专业数",
    "机械自动化机器人专业数",
    "师范类相关专业数",
    "工商旅游管理专业数",
    "环境工程科学专业数",
    "农业不含动物相关专业数",
    "高收费或超预算专业数",
    "护理助产专业数",
    "动物医学兽医专业数",
    "临床口腔中医暂缓专业数",
    "医技护理康复专业数",
    "特殊限制待核专业数",
    "调剂初判",
    "完整专业组接受度初判",
    "组内专业摘录",
    "办学地点线索",
    "来源页码",
    "版面列",
    "保真核验层级",
    "机器核验下一步",
    "人工核验下一步",
    "下一步核验动作",
    "讨论排序分",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

MAJOR_FIELDS = [
    "第四轮候选ID",
    "讨论集合",
    "城市",
    "冲稳保",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组内专业序号",
    "专业代号",
    "专业名称及备注",
    "专业偏好方向",
    "第四轮方向标签",
    "专业角色判断",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
]

CITY_SUMMARY_FIELDS = [
    "城市",
    "省份",
    "主表专业组数",
    "优先讨论专业组数",
    "保底观察",
    "稳妥观察",
    "稳冲观察",
    "冲刺观察",
    "历史待补",
]


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: public_value(row.get(field, "")) for field in fields} for row in rows])


def unique_values(rows, field, limit=None):
    values = []
    seen = set()
    for row in rows:
        value = str(row.get(field, "")).strip()
        if value and value not in seen:
            seen.add(value)
            values.append(value)
        if limit and len(values) >= limit:
            break
    return values


def join_values(rows, field, limit=None):
    return "；".join(unique_values(rows, field, limit=limit))


def aggregate_moe(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[row.get("专业组出现ID", "")].append(row)
    result = {}
    for group_id, group_rows in grouped.items():
        location = join_values(group_rows, "教育部所在地")
        city_candidate = join_values(group_rows, "教育部所在地命中城市") or join_values(group_rows, "城市候选")
        city = location or city_candidate or "未识别"
        result[group_id] = {
            "城市": city,
            "城市来源": "教育部所在地" if location else ("教育部所在地命中城市/城市候选" if city_candidate else "未识别"),
            "教育部所在地": location,
            "教育部所在地命中城市": join_values(group_rows, "教育部所在地命中城市"),
            "原城市候选": join_values(group_rows, "城市候选"),
            "所在地使用边界": join_values(group_rows, "所在地使用边界"),
            "城市字段状态": join_values(group_rows, "城市字段状态"),
            "校区字段状态": join_values(group_rows, "校区字段状态"),
            "省份": join_values(group_rows, "教育部省份分组") or join_values(group_rows, "教育部所在地"),
            "主管部门": join_values(group_rows, "教育部主管部门"),
            "学校登记备注": join_values(group_rows, "教育部备注"),
            "办学地点线索": join_values(group_rows, "校区/办学地点候选", limit=6),
        }
    return result


def aggregate_history(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[row.get("专业组出现ID", "")].append(row)
    result = {}
    fields = [
        "2023投档分",
        "2023等位分差",
        "2024投档分",
        "2024等位分差",
        "2025投档分",
        "2025等位分差",
        "同代码命中年份数",
        "同代码命中年份",
        "三年投档分范围",
        "等位分差摘要",
        "历史线使用口径",
        "再选要求是否跨年变化",
        "历史院校专业组名称是否疑似不一致",
    ]
    for group_id, group_rows in grouped.items():
        result[group_id] = {field: join_values(group_rows, field) for field in fields}
    return result


def agriculture_without_animal(row):
    return max(count(row, "农业动医兽医专业数") - count(row, "动物医学兽医专业数"), 0)


def group_score(row):
    hist = {"H1": 42, "H2": 46, "H3": 38, "H4": 28, "H0": 16, "H5": -15}.get(history_code(row), 0)
    score = hist
    score += count(row, "数字媒体技术专业数") * 12
    score += count(row, "计算机AI软件专业数") * 5
    score += count(row, "计算机类相关专业数") * 4
    score += count(row, "电子信息网络专业数") * 4
    score += count(row, "机械自动化机器人专业数") * 4
    score += count(row, "师范类相关专业数") * 4
    score += count(row, "工商旅游管理专业数") * 2
    score += count(row, "环境工程科学专业数") * 2
    score += agriculture_without_animal(row) * 2
    if row.get("调剂初判", "") == "可进入人工调剂接受度判断":
        score += 8
    if "结构异常" in row.get("调剂初判", ""):
        score -= 12
    if count(row, "特殊限制待核专业数") > 0:
        score -= 8
    score -= max(float_value(row, "历史最高等位分差", 0), 0) / 4
    return round(score, 3)


def sorted_groups(rows):
    gradient_order = {"保底观察": 1, "稳妥观察": 2, "稳冲观察": 3, "冲刺观察": 4, "历史待补": 5, "高冲暂缓": 6}
    return sorted(
        rows,
        key=lambda row: (
            row.get("城市", "未识别"),
            gradient_order.get(row.get("冲稳保", ""), 99),
            -float_value(row, "讨论排序分", 0),
            row.get("院校代码", ""),
            row.get("院校专业组代码", ""),
            row.get("专业组出现ID", ""),
        ),
    )


def priority_groups(rows):
    selected = []
    selected_ids = set()
    quotas = [
        ({"H1", "H2"}, 50),
        ({"H3"}, 35),
        ({"H4"}, 35),
    ]
    by_score = sorted(rows, key=lambda row: (-float_value(row, "讨论排序分", 0), row.get("院校代码", "")))
    for codes, limit in quotas:
        count_taken = 0
        for row in by_score:
            if row.get("专业组出现ID") in selected_ids or history_code(row) not in codes:
                continue
            out = dict(row)
            out["是否进入优先讨论"] = "true"
            selected.append(out)
            selected_ids.add(row.get("专业组出现ID"))
            count_taken += 1
            if count_taken >= limit:
                break
    return sorted_groups(selected)


def classification_explain(row):
    band = history_band(row)
    max_diff = row.get("历史最高等位分差", "")
    if history_code(row) == "H0":
        return "无同代码历史投档线，先作为历史待补组，不按分差下结论。"
    return f"{band}：按同代码历史最高等位分差 {max_diff} 判断；分差为各年投档分减该年515位次等位分。"


def relation_to_515(row, history):
    if history_code(row) == "H0":
        return "515分、91723位次暂无同代码历史线可直接比；需补同校同类组或官方投档线证据。"
    summary = history.get("等位分差摘要", "")
    max_diff = row.get("历史最高等位分差", "")
    if summary:
        return f"515对应等位分：2025={EQUIVALENT_SCORES['2025']}、2024={EQUIVALENT_SCORES['2024']}、2023={EQUIVALENT_SCORES['2023']}；{summary}；历史最高等位分差={max_diff}。"
    return f"515对应等位分：2025={EQUIVALENT_SCORES['2025']}、2024={EQUIVALENT_SCORES['2024']}、2023={EQUIVALENT_SCORES['2023']}；历史最高等位分差={max_diff}。"


def acceptance_note(row):
    notes = []
    if count(row, "护理助产专业数") > 0:
        notes.append("护理/助产暂不纳入")
    if count(row, "动物医学兽医专业数") > 0:
        notes.append("动物医学/兽医/动物科学暂不纳入")
    if count(row, "临床口腔中医暂缓专业数") > 0:
        notes.append("临床/口腔/中医暂缓")
    if count(row, "医技护理康复专业数") > 0:
        notes.append("医技/康复暂不进普通主线")
    if count(row, "高收费或超预算专业数") > 0:
        notes.append("高收费或超预算")
    if "结构异常" in row.get("调剂初判", ""):
        notes.append("组内结构需先核")
    if not notes:
        notes.append("主线机器门禁通过，仍需完整专业组人工核验。")
    return "；".join(notes)


def next_action(row):
    if history_code(row) == "H0":
        return "先补历史投档映射，再核第19期原页、湖北官方侧、招生章程、学费、计划数、选科、校区和调剂范围。"
    return "核第19期原页和完整专业组；核湖北官方系统/省招办计划；核高校招生章程、学费、计划数、选科、校区、语种/单科/体检限制和调剂范围。"


def decorate(row, moe_map, history_map, discussion_set):
    group_id = row.get("专业组出现ID", "")
    moe = moe_map.get(group_id, {})
    history = history_map.get(group_id, {})
    out = {
        "第四轮候选ID": f"R4GROUP-{stable_hash('issue19-round4-city-gradient', row)}",
        "第四轮来源口径": SOURCE_POLICY,
        "讨论集合": discussion_set,
        "是否进入优先讨论": "false",
        "是否可作为定稿依据": "false",
        "城市": moe.get("城市", row.get("城市候选", "") or "未识别"),
        "城市来源": moe.get("城市来源", "城市候选/未识别"),
        "教育部所在地": moe.get("教育部所在地", ""),
        "教育部所在地命中城市": moe.get("教育部所在地命中城市", ""),
        "原城市候选": moe.get("原城市候选", row.get("城市候选", "")),
        "城市口径提醒": CITY_POLICY,
        "所在地使用边界": moe.get("所在地使用边界", ""),
        "城市字段状态": moe.get("城市字段状态", ""),
        "校区字段状态": moe.get("校区字段状态", ""),
        "省份": moe.get("省份", ""),
        "主管部门": moe.get("主管部门", ""),
        "学校登记备注": moe.get("学校登记备注", ""),
        "院校代码": row.get("院校代码", ""),
        "院校名称": row.get("院校名称OCR", ""),
        "院校专业组代码": row.get("院校专业组代码OCR规范化", ""),
        "专业组号": row.get("专业组号OCR", ""),
        "历史线索分层": row.get("历史线索分层", ""),
        "冲稳保": history_band(row),
        "冲稳保解释": classification_explain(row),
        "与515关系": relation_to_515(row, history),
        "2026考生分数": CANDIDATE_SCORE,
        "2026考生位次": CANDIDATE_RANK,
        "2025等位分": EQUIVALENT_SCORES["2025"],
        "2024等位分": EQUIVALENT_SCORES["2024"],
        "2023等位分": EQUIVALENT_SCORES["2023"],
        "2025投档分": history.get("2025投档分", ""),
        "2025等位分差": history.get("2025等位分差", ""),
        "2024投档分": history.get("2024投档分", ""),
        "2024等位分差": history.get("2024等位分差", ""),
        "2023投档分": history.get("2023投档分", ""),
        "2023等位分差": history.get("2023等位分差", ""),
        "同代码命中年份数": history.get("同代码命中年份数", row.get("同代码命中年份最大值", "")),
        "同代码命中年份": history.get("同代码命中年份", ""),
        "三年投档分范围": history.get("三年投档分范围", ""),
        "等位分差摘要": history.get("等位分差摘要", ""),
        "历史线使用口径": history.get("历史线使用口径", ""),
        "再选要求跨年变化": history.get("再选要求是否跨年变化", ""),
        "历史组名疑似不一致": history.get("历史院校专业组名称是否疑似不一致", ""),
        "公办民办机器线索": row.get("公办民办机器线索", ""),
        "家庭底线属性动作": row.get("家庭底线属性动作", ""),
        "专业明细行数": row.get("专业明细行数", ""),
        "数字媒体技术专业数": row.get("数字媒体技术专业数", ""),
        "计算机类相关专业数": row.get("计算机类相关专业数", ""),
        "计算机AI软件专业数": row.get("计算机AI软件专业数", ""),
        "电子信息网络专业数": row.get("电子信息网络专业数", ""),
        "机械自动化机器人专业数": row.get("机械自动化机器人专业数", ""),
        "师范类相关专业数": row.get("师范类相关专业数", ""),
        "工商旅游管理专业数": row.get("工商旅游管理专业数", ""),
        "环境工程科学专业数": row.get("环境工程科学专业数", ""),
        "农业不含动物相关专业数": str(agriculture_without_animal(row)),
        "高收费或超预算专业数": row.get("高收费或超预算专业数", ""),
        "护理助产专业数": row.get("护理助产专业数", ""),
        "动物医学兽医专业数": row.get("动物医学兽医专业数", ""),
        "临床口腔中医暂缓专业数": row.get("临床口腔中医暂缓专业数", ""),
        "医技护理康复专业数": row.get("医技护理康复专业数", ""),
        "特殊限制待核专业数": row.get("特殊限制待核专业数", ""),
        "调剂初判": row.get("调剂初判", ""),
        "完整专业组接受度初判": acceptance_note(row),
        "组内专业摘录": row.get("组内完整专业清单索引", ""),
        "办学地点线索": moe.get("办学地点线索", ""),
        "来源页码": row.get("来源页码", ""),
        "版面列": row.get("版面列", ""),
        "保真核验层级": row.get("保真核验层级", ""),
        "机器核验下一步": row.get("机器核验下一步", ""),
        "人工核验下一步": row.get("人工核验下一步", ""),
        "下一步核验动作": next_action(row),
        "讨论排序分": str(group_score(row)),
        "专业组出现ID": group_id,
        "稳定基座专业组筛选ID": row.get("稳定基座专业组筛选ID", ""),
    }
    return out


def build_major_rows(group_rows, major_rows):
    group_by_id = {row.get("专业组出现ID", ""): row for row in group_rows}
    rows = []
    for row in major_rows:
        group = group_by_id.get(row.get("专业组出现ID", ""))
        if not group:
            continue
        out = {
            "第四轮候选ID": group.get("第四轮候选ID", ""),
            "讨论集合": group.get("讨论集合", ""),
            "城市": group.get("城市", ""),
            "冲稳保": group.get("冲稳保", ""),
            "院校代码": row.get("院校代码", ""),
            "院校名称": row.get("院校名称OCR", ""),
            "院校专业组代码": row.get("院校专业组代码OCR规范化", ""),
            "专业组内专业序号": row.get("专业组内专业序号", ""),
            "专业代号": row.get("专业代号OCR", ""),
            "专业名称及备注": row.get("专业名称及备注短摘", ""),
            "专业偏好方向": row.get("专业偏好方向", ""),
            "第四轮方向标签": "；".join(direction_tags(row)),
            "专业角色判断": major_role(row),
            "机器专业接受度初判": row.get("机器专业接受度初判", ""),
            "专业风险类型": row.get("专业风险类型", ""),
            "再选科目OCR候选": row.get("再选科目OCR候选", ""),
            "专业计划数OCR候选": row.get("专业计划数OCR候选", ""),
            "学费OCR候选": row.get("学费OCR候选", ""),
            "字段缺口数": row.get("字段缺口数", ""),
            "字段缺口字段": row.get("字段缺口字段", ""),
            "人工核验优先级": row.get("人工核验优先级", ""),
            "人工核验强度": row.get("人工核验强度", ""),
            "是否必须100%人工核验": row.get("是否必须100%人工核验", ""),
            "PDF原页核验状态": row.get("PDF原页核验状态", ""),
            "湖北官方系统核验状态": row.get("湖北官方系统核验状态", ""),
            "高校官网证据匹配状态": row.get("高校官网证据匹配状态", ""),
            "历史线索分层": row.get("历史线索分层", ""),
            "历史最高等位分差": row.get("历史最高等位分差", ""),
            "历史最低等位分差": row.get("历史最低等位分差", ""),
            "来源页码": row.get("来源页码", ""),
            "版面列": row.get("版面列", ""),
            "专业行ID": row.get("专业行ID", ""),
            "专业组出现ID": row.get("专业组出现ID", ""),
        }
        rows.append(out)
    return rows


def city_summary(rows, priority_rows):
    priority_ids = {row.get("专业组出现ID") for row in priority_rows}
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row.get("城市", "未识别"), row.get("省份", ""))].append(row)
    summary = []
    for (city, province), group_rows in sorted(grouped.items()):
        counter = Counter(row.get("冲稳保", "") for row in group_rows)
        summary.append(
            {
                "城市": city,
                "省份": province,
                "主表专业组数": len(group_rows),
                "优先讨论专业组数": sum(row.get("专业组出现ID") in priority_ids for row in group_rows),
                "保底观察": counter.get("保底观察", 0),
                "稳妥观察": counter.get("稳妥观察", 0),
                "稳冲观察": counter.get("稳冲观察", 0),
                "冲刺观察": counter.get("冲刺观察", 0),
                "历史待补": counter.get("历史待补", 0),
            }
        )
    return summary


def write_workbook(candidate_rows, priority_rows, major_rows, city_rows, history_missing_rows, high_rush_rows):
    wb = Workbook()
    wb.remove(wb.active)
    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "Round4 按城市和冲稳保重建候选全集，用于家庭讨论和后续核验，不作为定稿依据。"),
        ("来源", SOURCE_POLICY),
        ("成绩口径", "考生 515 分、累计位次 91723；等位分按 2025=494、2024=497、2023=481 比较历史投档分。"),
        ("主表边界", REACHABLE_POLICY),
        ("家庭边界", FAMILY_POLICY),
        ("城市口径", CITY_POLICY),
        ("下一步", "先按城市筛一遍生活接受度，再逐组看完整专业组是否能接受调剂；H0 历史待补只做补证索引，不当作可冲稳保对象。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 120

    append_sheet(wb.create_sheet("01_按城市冲稳保总表"), candidate_rows, GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_优先讨论120组"), priority_rows, GROUP_FIELDS)
    append_sheet(wb.create_sheet("03_城市分布"), city_rows, CITY_SUMMARY_FIELDS)
    append_sheet(wb.create_sheet("04_组内专业明细"), major_rows, MAJOR_FIELDS)
    append_sheet(wb.create_sheet("05_历史待补组"), history_missing_rows, GROUP_FIELDS)
    append_sheet(wb.create_sheet("06_高冲暂缓索引"), high_rush_rows, GROUP_FIELDS)
    ROUND4_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ROUND4_WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "/home/",
        "/var/folders/",
        "/private/",
        "private/",
        "private\\",
        "file://",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"Round4 公开文件包含禁止内容：{hit}")


def main():
    group_rows = read_csv(GROUP_BROWSER)
    major_rows = read_csv(MAJOR_BROWSER)
    major_signals, _ = aggregate_major_signals(major_rows)
    enriched_groups = merge_signals(group_rows, major_signals)

    moe_map = aggregate_moe(read_csv(MOE_ATTRIBUTES))
    history_map = aggregate_history(read_csv(HISTORICAL_LINES))

    main_pool = [row for row in enriched_groups if main_ok(row)]
    candidate_source = [row for row in main_pool if history_code(row) in {"H1", "H2", "H3", "H4"}]
    history_missing_source = [row for row in main_pool if history_code(row) == "H0"]
    high_rush_source = [row for row in main_pool if history_code(row) == "H5"]

    candidate_rows = sorted_groups(
        decorate(row, moe_map, history_map, "主线可能够得着全集")
        for row in candidate_source
    )
    priority_rows = priority_groups(candidate_rows)
    priority_ids = {row.get("专业组出现ID") for row in priority_rows}
    candidate_rows = [
        {**row, "是否进入优先讨论": "true" if row.get("专业组出现ID") in priority_ids else "false"}
        for row in candidate_rows
    ]
    history_missing_rows = sorted_groups(
        decorate(row, moe_map, history_map, "历史待补附录") for row in history_missing_source
    )
    high_rush_rows = sorted_groups(
        decorate(row, moe_map, history_map, "高冲暂缓索引") for row in high_rush_source
    )
    city_rows = city_summary(candidate_rows, priority_rows)
    major_detail_rows = build_major_rows(candidate_rows, major_rows)

    write_csv(ROUND4_GROUPS, candidate_rows, GROUP_FIELDS)
    write_csv(ROUND4_PRIORITY, priority_rows, GROUP_FIELDS)
    write_csv(ROUND4_MAJORS, major_detail_rows, MAJOR_FIELDS)
    write_csv(ROUND4_CITY_SUMMARY, city_rows, CITY_SUMMARY_FIELDS)
    write_csv(ROUND4_HISTORY_MISSING, history_missing_rows, GROUP_FIELDS)
    write_csv(ROUND4_HIGH_RUSH_PAUSED, high_rush_rows, GROUP_FIELDS)
    write_workbook(candidate_rows, priority_rows, major_detail_rows, city_rows, history_missing_rows, high_rush_rows)

    summary = {
        "status": "issue19_round4_city_gradient_ready",
        "generated_by": "build_issue19_round4_city_gradient_candidates.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "source_policy": SOURCE_POLICY,
        "city_policy": CITY_POLICY,
        "reachable_policy": REACHABLE_POLICY,
        "family_policy": FAMILY_POLICY,
        "candidate_score": CANDIDATE_SCORE,
        "candidate_rank": CANDIDATE_RANK,
        "equivalent_scores": EQUIVALENT_SCORES,
        "main_pool_rows_before_reachable_filter": len(main_pool),
        "candidate_group_rows": len(candidate_rows),
        "priority_group_rows": len(priority_rows),
        "major_detail_rows": len(major_detail_rows),
        "city_summary_rows": len(city_rows),
        "history_missing_group_rows": len(history_missing_rows),
        "high_rush_paused_group_rows": len(high_rush_rows),
        "gradient_distribution": dict(Counter(row.get("冲稳保", "") for row in candidate_rows)),
        "priority_gradient_distribution": dict(Counter(row.get("冲稳保", "") for row in priority_rows)),
        "top_city_counts": Counter(row.get("城市", "") or "未识别" for row in candidate_rows).most_common(30),
        "outputs": {
            "workbook": str(ROUND4_WORKBOOK.relative_to(ROOT)),
            "groups_csv": str(ROUND4_GROUPS.relative_to(ROOT)),
            "priority_csv": str(ROUND4_PRIORITY.relative_to(ROOT)),
            "majors_csv": str(ROUND4_MAJORS.relative_to(ROOT)),
            "city_summary_csv": str(ROUND4_CITY_SUMMARY.relative_to(ROOT)),
            "history_missing_csv": str(ROUND4_HISTORY_MISSING.relative_to(ROOT)),
            "high_rush_paused_csv": str(ROUND4_HIGH_RUSH_PAUSED.relative_to(ROOT)),
        },
        "hard_filters_main": [
            "公办/普通学费机器线索",
            "组内高收费或超预算专业数为0",
            "必须有逐专业明细",
            "必须命中至少一个当前可讨论方向",
            "临床/口腔/中医等暂缓方向不进普通主线",
            "护理/助产、动物医学/兽医/动物科学不进普通主线",
            "医技/康复不进普通主线",
            "H0 历史待补不进主表",
            "H5 高冲暂缓不进主表",
        ],
    }
    ROUND4_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe(
        [
            ROUND4_GROUPS,
            ROUND4_PRIORITY,
            ROUND4_MAJORS,
            ROUND4_CITY_SUMMARY,
            ROUND4_HISTORY_MISSING,
            ROUND4_HIGH_RUSH_PAUSED,
            ROUND4_SUMMARY,
        ]
    )

    print(f"写出 Round4 城市冲稳保总表：{ROUND4_GROUPS.relative_to(ROOT)}")
    print(f"写出 Round4 优先讨论表：{ROUND4_PRIORITY.relative_to(ROOT)}")
    print(f"写出 Round4 工作簿：{ROUND4_WORKBOOK.relative_to(ROOT)}")
    print(
        f"主线池={len(main_pool)}；主表={len(candidate_rows)}；优先讨论={len(priority_rows)}；"
        f"专业明细={len(major_detail_rows)}；高冲暂缓={len(high_rush_rows)}"
    )


if __name__ == "__main__":
    main()
