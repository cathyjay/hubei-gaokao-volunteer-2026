#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

GROUP_BROWSER = EXPORTS / "issue19-stable-foundation-group-browser.csv"
MAJOR_BROWSER = EXPORTS / "issue19-stable-foundation-major-browser.csv"

ROUND2_GROUPS = EXPORTS / "issue19-round2-updated-preferences-candidate-groups.csv"
ROUND2_MAIN_SHORTLIST = EXPORTS / "issue19-round2-updated-preferences-main-shortlist-groups.csv"
ROUND2_SPECIAL_GROUPS = EXPORTS / "issue19-round2-updated-preferences-health-agri-special-groups.csv"
ROUND2_PRIORITY_CITY = EXPORTS / "issue19-round2-updated-preferences-priority-city-watchlist.csv"
ROUND2_SPECIFIC_WATCHLIST = EXPORTS / "issue19-round2-updated-preferences-specific-watchlist.csv"
ROUND2_MAIN_MAJORS = EXPORTS / "issue19-round2-updated-preferences-main-shortlist-majors.csv"
ROUND2_SPECIAL_MAJORS = EXPORTS / "issue19-round2-updated-preferences-special-majors.csv"
ROUND2_WORKBOOK = EXPORTS / "issue19-round2-updated-preferences.xlsx"
ROUND2_SUMMARY = EXPORTS / "issue19-round2-updated-preferences-summary.json"

SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
PRIORITY_CITIES = {"武汉", "成都", "西安", "北京"}

DIRECTION_KEYWORDS = {
    "机械自动化机器人": [
        "机械",
        "自动化",
        "机器人工程",
        "智能制造",
        "机械电子",
        "机械设计制造",
        "智能装备",
    ],
    "电子信息网络": [
        "电子信息",
        "电子科学",
        "信息工程",
        "通信工程",
        "信息网络",
        "网络工程",
        "物联网",
        "集成电路",
        "微电子",
        "光电信息",
    ],
    "计算机AI软件": [
        "计算机",
        "软件工程",
        "人工智能",
        "数据科学",
        "大数据",
        "智能科学",
        "信息安全",
        "网络空间",
    ],
    "农业动医兽医": [
        "动物医学",
        "动物科学",
        "兽医",
        "农学",
        "农业",
        "园艺",
        "植物保护",
        "水产",
        "智慧农业",
        "农业资源与环境",
    ],
    "工商旅游管理": [
        "工商管理",
        "市场营销",
        "会计",
        "财务管理",
        "人力资源",
        "审计",
        "物流管理",
        "电子商务",
        "旅游管理",
    ],
    "环境工程科学": [
        "环境工程",
        "环境科学",
        "环境生态",
        "环保",
        "生态",
        "农业资源与环境",
    ],
    "医技护理康复": ["医学影像", "医学检验", "护理", "康复", "助产"],
    "临床口腔中医暂缓": [
        "临床医学",
        "口腔医学",
        "中医学",
        "针灸推拿",
        "中西医",
        "麻醉学",
        "儿科学",
        "精神医学",
        "预防医学",
    ],
}

SPECIAL_ANIMAL_KEYWORDS = ["动物医学", "兽医", "动物科学"]
SPECIAL_NURSING_KEYWORDS = ["护理", "助产"]
SPECIFIC_SCHOOL_KEYWORDS = {
    "西安建筑科技大学": ["环境科学", "环境工程"],
    "贵州大学": ["旅游"],
    "北京农学院": ["农业资源与环境"],
    "厦门大学嘉庚学院": [],
    "辽宁工程技术大学": [],
}

GROUP_FIELDS = [
    "第二轮候选ID",
    "第二轮池",
    "第二轮层级",
    "建议梯度",
    "是否进入家庭讨论",
    "是否可作为定稿依据",
    "入选理由",
    "专项风险提示",
    "下一步核验动作",
    "体检和章程核验提示",
    "重点观察标签",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组号OCR",
    "城市候选",
    "公办民办机器线索",
    "家庭底线属性动作",
    "机器家庭匹配初判",
    "调剂初判",
    "偏好专业数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "师范类相关专业数",
    "机械自动化机器人专业数",
    "电子信息网络专业数",
    "计算机AI软件专业数",
    "农业动医兽医专业数",
    "动物医学兽医专业数",
    "护理助产专业数",
    "工商旅游管理专业数",
    "环境工程科学专业数",
    "医技护理康复专业数",
    "临床口腔中医暂缓专业数",
    "高收费或超预算专业数",
    "特殊限制待核专业数",
    "专业明细行数",
    "机器筛选价值层级",
    "保真核验层级",
    "历史线索分层",
    "同代码命中年份最大值",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "组内完整专业清单索引",
    "机器核验下一步",
    "人工核验下一步",
    "专业组出现ID",
    "稳定基座专业组筛选ID",
]

MAJOR_FIELDS = [
    "第二轮候选ID",
    "第二轮池",
    "第二轮层级",
    "建议梯度",
    "专业角色判断",
    "第二轮方向标签",
    "是否可作为定稿依据",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "专业偏好方向",
    "机器专业接受度初判",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "历史线索分层",
    "历史最高等位分差",
    "历史最低等位分差",
    "来源页码",
    "版面列",
    "专业行ID",
    "专业组出现ID",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: public_value(row.get(field, "")) for field in fields} for row in rows])


def public_value(value):
    return str(value).replace("最终候选", "本轮候选")


def int_value(row, field):
    try:
        return int(str(row.get(field, "")).strip() or "0")
    except ValueError:
        return 0


def float_value(row, field, default=999):
    try:
        return float(str(row.get(field, "")).strip())
    except ValueError:
        return default


def history_code(row):
    return row.get("历史线索分层", "")[:2]


def history_band(row):
    code = history_code(row)
    diff = float_value(row, "历史最高等位分差")
    if code == "H0":
        return "历史待补"
    if code == "H1" or diff <= -20:
        return "保底观察"
    if code == "H2" or diff <= 0:
        return "稳妥观察"
    if code == "H3" or diff <= 15:
        return "稳冲观察"
    if code == "H4" or diff <= 30:
        return "冲刺观察"
    return "高冲暂缓"


def public_normal(row):
    return row.get("家庭底线属性动作", "") == "继续核公办普通学费-非民办线索"


def has_priority_city(row):
    return row.get("城市候选", "") in PRIORITY_CITIES


def major_name_text(row):
    return "；".join(
        [
            row.get("专业名称及备注短摘", ""),
            row.get("专业偏好方向", ""),
        ]
    )


def text_of_major(row):
    return "；".join(
        [
            row.get("专业名称及备注短摘", ""),
            row.get("专业偏好方向", ""),
            row.get("机器专业接受度初判", ""),
            row.get("专业风险类型", ""),
        ]
    )


def direction_tags(row):
    text = major_name_text(row)
    tags = []
    if row.get("专业偏好方向"):
        tags.append(row.get("专业偏好方向"))
    for direction, keywords in DIRECTION_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            tags.append(direction)
    if any(keyword in text for keyword in SPECIAL_ANIMAL_KEYWORDS):
        tags.append("动物医学兽医专项")
    if any(keyword in text for keyword in SPECIAL_NURSING_KEYWORDS):
        tags.append("护理助产专项")
    return sorted(set(tags))


def stable_hash(prefix, row):
    key = "|".join(
        [
            row.get("专业组出现ID", ""),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ]
    )
    return hashlib.sha256((prefix + "|" + key).encode("utf-8")).hexdigest()[:16]


def aggregate_major_signals(major_rows):
    grouped = defaultdict(list)
    for row in major_rows:
        grouped[row.get("专业组出现ID", "")].append(row)

    signals = {}
    for group_id, rows in grouped.items():
        counter = Counter()
        tag_rows = defaultdict(int)
        for row in rows:
            tags = direction_tags(row)
            for tag in tags:
                tag_rows[tag] += 1
            text = major_name_text(row)
            if any(keyword in text for keyword in SPECIAL_ANIMAL_KEYWORDS):
                counter["动物医学兽医专业数"] += 1
            if any(keyword in text for keyword in SPECIAL_NURSING_KEYWORDS):
                counter["护理助产专业数"] += 1
        for key in DIRECTION_KEYWORDS:
            counter[f"{key}专业数"] = tag_rows[key]
        signals[group_id] = dict(counter)
    return signals, grouped


def merge_signals(group_rows, major_signals):
    rows = []
    for row in group_rows:
        out = dict(row)
        signals = major_signals.get(row.get("专业组出现ID", ""), {})
        for field in [
            "机械自动化机器人专业数",
            "电子信息网络专业数",
            "计算机AI软件专业数",
            "农业动医兽医专业数",
            "动物医学兽医专业数",
            "护理助产专业数",
            "工商旅游管理专业数",
            "环境工程科学专业数",
            "医技护理康复专业数",
            "临床口腔中医暂缓专业数",
        ]:
            out[field] = str(signals.get(field, 0))
        rows.append(out)
    return rows


def count(row, field):
    return int_value(row, field)


def has_main_direction(row):
    return any(
        count(row, field) > 0
        for field in [
            "数字媒体技术专业数",
            "计算机类相关专业数",
            "师范类相关专业数",
            "机械自动化机器人专业数",
            "电子信息网络专业数",
            "计算机AI软件专业数",
            "工商旅游管理专业数",
            "环境工程科学专业数",
            "农业动医兽医专业数",
        ]
    )


def has_health_or_vet_special(row):
    return count(row, "医技护理康复专业数") > 0


def has_nursing_or_animal_block(row):
    return count(row, "护理助产专业数") > 0 or count(row, "动物医学兽医专业数") > 0


def has_medtech_rehab_special(row):
    return count(row, "医技护理康复专业数") > 0


def agriculture_without_animal_count(row):
    return max(count(row, "农业动医兽医专业数") - count(row, "动物医学兽医专业数"), 0)


def has_clinical_pause(row):
    return count(row, "临床口腔中医暂缓专业数") > 0


def specific_watch_tags(row):
    school = row.get("院校名称OCR", "")
    group_code = row.get("院校专业组代码OCR规范化", "")
    majors = row.get("组内完整专业清单索引", "")
    tags = []
    for school_keyword, major_keywords in SPECIFIC_SCHOOL_KEYWORDS.items():
        if school_keyword not in school:
            continue
        if major_keywords and not any(keyword in majors for keyword in major_keywords):
            continue
        if school_keyword == "厦门大学嘉庚学院" and group_code[-2:] not in {"06", "07"}:
            continue
        tags.append(school_keyword)
    return tags


def main_ok(row):
    return (
        public_normal(row)
        and count(row, "高收费或超预算专业数") == 0
        and count(row, "专业明细行数") > 0
        and has_main_direction(row)
        and not has_clinical_pause(row)
        and not has_health_or_vet_special(row)
        and not has_nursing_or_animal_block(row)
    )


def special_ok(row):
    return (
        public_normal(row)
        and count(row, "高收费或超预算专业数") == 0
        and count(row, "专业明细行数") > 0
        and not has_clinical_pause(row)
        and not has_nursing_or_animal_block(row)
        and (
            has_medtech_rehab_special(row)
            or agriculture_without_animal_count(row) > 0
            or count(row, "环境工程科学专业数") > 0
            or "北京农学院" in row.get("院校名称OCR", "")
        )
    )


def score(row, special=False):
    hist_score = {
        "H1": 42,
        "H2": 38,
        "H3": 25,
        "H4": 13,
        "H0": 8,
        "H5": -20,
    }.get(history_code(row), 0)
    value = hist_score
    value += count(row, "数字媒体技术专业数") * 10
    value += count(row, "计算机类相关专业数") * 5
    value += count(row, "计算机AI软件专业数") * 5
    value += count(row, "电子信息网络专业数") * 4
    value += count(row, "机械自动化机器人专业数") * 4
    value += count(row, "师范类相关专业数") * 4
    value += count(row, "环境工程科学专业数") * 3
    value += count(row, "工商旅游管理专业数") * 2
    value += agriculture_without_animal_count(row) * (4 if special else 2)
    value += count(row, "医技护理康复专业数") * (4 if special else -8)
    value -= count(row, "动物医学兽医专业数") * 50
    value -= count(row, "护理助产专业数") * 50
    if has_priority_city(row):
        value += 8
    if specific_watch_tags(row):
        value += 10
    if row.get("调剂初判", "") == "可进入人工调剂接受度判断":
        value += 8
    if count(row, "特殊限制待核专业数") > 0:
        value -= 8
    if "结构异常" in row.get("调剂初判", ""):
        value -= 8
    if has_clinical_pause(row):
        value -= 30
    value -= max(float_value(row, "历史最高等位分差", 0), 0) / 5
    return value


def sorted_candidates(rows, special=False):
    return sorted(
        rows,
        key=lambda row: (
            -score(row, special=special),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
            row.get("专业组出现ID", ""),
        ),
    )


def risk_notes(row, special=False):
    notes = []
    if has_clinical_pause(row):
        notes.append("组内含临床/口腔/中医等暂缓方向")
    if count(row, "护理助产专业数") > 0:
        notes.append("护理/助产暂不纳入：职业内容、夜班体力、实习和薪酬预期待确认")
    if count(row, "医技护理康复专业数") > 0:
        notes.append("医技/康复需核体检、证书、实习和就业路径")
    if count(row, "动物医学兽医专业数") > 0:
        notes.append("动物医学/兽医暂不纳入：动物接触、实验解剖、考证和职业暴露待确认")
    if count(row, "特殊限制待核专业数") > 0:
        notes.append("组内有特殊限制待核")
    if "结构异常" in row.get("调剂初判", ""):
        notes.append("结构或组内边界需先核")
    if history_code(row) == "H0":
        notes.append("缺同代码历史线索")
    if history_band(row) in {"冲刺观察", "高冲暂缓"}:
        notes.append("录取线距离偏冲")
    if not notes:
        notes.append("仍需完整组核验")
    if special:
        notes.append("专项方向只进入了解池，不与普通主线混排")
    return "；".join(notes)


def direction_summary(row):
    parts = []
    for label, field in [
        ("数字媒体", "数字媒体技术专业数"),
        ("计算机/AI/软件", "计算机AI软件专业数"),
        ("电子/网络", "电子信息网络专业数"),
        ("机械/自动化/机器人", "机械自动化机器人专业数"),
        ("师范", "师范类相关专业数"),
        ("环境", "环境工程科学专业数"),
        ("农业", "农业动医兽医专业数"),
        ("动物医学/兽医", "动物医学兽医专业数"),
        ("护理/助产", "护理助产专业数"),
        ("工商/旅游管理", "工商旅游管理专业数"),
        ("医技/康复", "医技护理康复专业数"),
    ]:
        value = count(row, field)
        if value:
            parts.append(f"{label}{value}个")
    return "、".join(parts) if parts else "未命中新方向"


def selection_reason(row, pool):
    reasons = [f"命中方向：{direction_summary(row)}"]
    if has_priority_city(row):
        reasons.append(f"命中优先城市：{row.get('城市候选')}")
    tags = specific_watch_tags(row)
    if tags:
        reasons.append("命名观察：" + "、".join(tags))
    reasons.append(f"历史线索：{row.get('历史线索分层')}")
    reasons.append("本轮口径：公办/普通学费线索；所有结论仍需原页、官方侧和章程核验")
    if pool == "主线":
        reasons.append("主线未纳入临床/口腔/中医、护理、医技、动物医学/兽医等方向；护理和动物医学/兽医暂不进入候选池")
    return "；".join(reasons)


def body_exam_note(row):
    if has_nursing_or_animal_block(row):
        return "护理、动物医学/兽医暂不纳入本轮候选；若未来重新讨论，需先核职业接受度、动物接触或夜班实习、职业暴露、体检要求和招生章程。"
    if has_health_or_vet_special(row) or has_clinical_pause(row):
        return "体检摘要显示色觉正常、矫正视力4.80/4.80；医技、康复、农学等仍必须逐校核招生章程和体检指导意见。"
    if count(row, "特殊限制待核专业数") > 0:
        return "体检或特殊限制字段待核；需看招生章程、专业备注和第19期原页。"
    return "无新增体检自动阻断线索；仍需按最终入围专业逐项核章程。"


def decorate(row, layer, pool, enter_discussion=True):
    out = dict(row)
    out["第二轮候选ID"] = f"R2GROUP-{stable_hash('issue19-round2', row)}"
    out["第二轮池"] = pool
    out["第二轮层级"] = layer
    out["建议梯度"] = history_band(row)
    out["是否进入家庭讨论"] = "true" if enter_discussion else "false"
    out["是否可作为定稿依据"] = "false"
    out["入选理由"] = selection_reason(row, pool)
    out["专项风险提示"] = risk_notes(row, special=(pool != "主线"))
    out["下一步核验动作"] = "回看第19期原页和完整专业组；核湖北官方系统/省招办计划；核高校招生章程、学费、计划数、选科、校区、语种/单科/体检限制和调剂范围。"
    out["体检和章程核验提示"] = body_exam_note(row)
    out["重点观察标签"] = "；".join(specific_watch_tags(row))
    return out


def take_unique(source, limit, selected_ids, special=False):
    picked = []
    for row in sorted_candidates(source, special=special):
        group_id = row.get("专业组出现ID", "")
        if group_id in selected_ids:
            continue
        picked.append(row)
        selected_ids.add(group_id)
        if len(picked) >= limit:
            break
    return picked


def build_main_shortlist(main_pool):
    selected_ids = set()
    shortlist = []
    quota_plan = [
        ("M1-保底稳妥主线优先", [row for row in main_pool if history_code(row) in {"H1", "H2"}], 40),
        ("M2-稳冲主线优先", [row for row in main_pool if history_code(row) == "H3"], 30),
        ("M3-冲刺主线观察", [row for row in main_pool if history_code(row) == "H4"], 20),
        (
            "M4-城市/点名/历史待补观察",
            [
                row
                for row in main_pool
                if history_code(row) == "H0" or has_priority_city(row) or specific_watch_tags(row)
            ],
            10,
        ),
    ]
    for layer, rows, limit in quota_plan:
        shortlist.extend(
            decorate(row, layer, "主线")
            for row in take_unique(rows, limit, selected_ids)
        )
    return shortlist


def build_special_groups(special_pool):
    rows = []
    selected_ids = set()
    quota_plan = [
        (
            "S1-医技/康复专项了解",
            [row for row in special_pool if has_medtech_rehab_special(row)],
            35,
        ),
        (
            "S2-农业/环境专项了解",
            [
                row
                for row in special_pool
                if agriculture_without_animal_count(row) > 0
                or count(row, "环境工程科学专业数") > 0
            ],
            45,
        ),
        (
            "S3-北京农学院等点名农业专项观察",
            [row for row in special_pool if "北京农学院" in row.get("院校名称OCR", "")],
            10,
        ),
    ]
    for layer, source, limit in quota_plan:
        rows.extend(
            decorate(row, layer, "专项")
            for row in take_unique(source, limit, selected_ids, special=True)
        )
    return rows


def build_all_candidates(main_pool, special_pool, specific_pool):
    all_rows = []
    seen = set()
    for pool_name, layer, source, special in [
        ("主线", "C1-主线完整候选池", main_pool, False),
        ("专项", "C2-医技康复农业环境专项候选池", special_pool, True),
        ("点名观察", "C3-用户点名学校观察池", specific_pool, True),
    ]:
        for row in sorted_candidates(source, special=special):
            group_id = row.get("专业组出现ID", "")
            if (pool_name, group_id) in seen:
                continue
            all_rows.append(decorate(row, layer, pool_name, enter_discussion=pool_name != "点名观察" or public_normal(row)))
            seen.add((pool_name, group_id))
    return all_rows


def build_priority_city_watchlist(all_candidates):
    return sorted(
        [row for row in all_candidates if row.get("城市候选", "") in PRIORITY_CITIES],
        key=lambda row: (
            row.get("城市候选", ""),
            {
                "稳妥观察": 0,
                "保底观察": 1,
                "稳冲观察": 2,
                "冲刺观察": 3,
                "历史待补": 4,
                "高冲暂缓": 5,
            }.get(row.get("建议梯度", ""), 9),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ),
    )


def build_specific_watchlist(all_candidates):
    return sorted(
        [row for row in all_candidates if row.get("重点观察标签", "")],
        key=lambda row: (
            row.get("重点观察标签", ""),
            float_value(row, "历史最高等位分差"),
            row.get("院校代码", ""),
            row.get("院校专业组代码OCR规范化", ""),
        ),
    )


def major_role(row):
    tags = direction_tags(row)
    text = text_of_major(row)
    if "临床口腔中医暂缓" in tags:
        return "暂缓方向-临床口腔中医等"
    if "护理助产专项" in tags:
        return "暂不纳入-护理助产"
    if "医技护理康复" in tags:
        return "专项了解-医技护理康复"
    if "动物医学兽医专项" in tags:
        return "暂不纳入-动物医学兽医"
    if "农业动医兽医" in tags:
        return "可了解-农业相关"
    if any(tag in tags for tag in ["数字媒体技术", "计算机类相关", "计算机AI软件"]):
        return "主线偏好-计算机/数字媒体"
    if any(tag in tags for tag in ["机械自动化机器人", "电子信息网络"]):
        return "主线偏好-工科信息制造"
    if "师范类相关" in tags:
        return "可了解-师范"
    if "环境工程科学" in tags:
        return "可了解-环境"
    if "工商旅游管理" in tags:
        return "可了解-管理/旅游"
    if "中外合作" in text or "高收费" in text:
        return "费用或培养模式需核"
    if row.get("字段缺口数", "") not in {"", "0"}:
        return "字段缺口待核"
    return "调剂接受度待判断"


def build_major_rows(group_rows, major_rows):
    group_meta = {row.get("专业组出现ID", ""): row for row in group_rows}
    rows = []
    for row in major_rows:
        meta = group_meta.get(row.get("专业组出现ID", ""))
        if not meta:
            continue
        out = dict(row)
        out["第二轮候选ID"] = meta.get("第二轮候选ID", "")
        out["第二轮池"] = meta.get("第二轮池", "")
        out["第二轮层级"] = meta.get("第二轮层级", "")
        out["建议梯度"] = meta.get("建议梯度", "")
        out["专业角色判断"] = major_role(out)
        out["第二轮方向标签"] = "；".join(direction_tags(out))
        out["是否可作为定稿依据"] = "false"
        rows.append(out)
    return sorted(
        rows,
        key=lambda row: (
            row.get("第二轮候选ID", ""),
            int_value(row, "专业组内专业序号"),
            row.get("专业代号OCR", ""),
        ),
    )


def append_sheet(ws, rows, fields):
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([public_value(row.get(field, "")) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, start=1):
        max_len = min(46, max([len(str(field))] + [len(str(row.get(field, ""))) for row in rows[:200]]))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, max_len + 2)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(all_candidates, main_shortlist, special_groups, city_watchlist, specific_watchlist, main_majors, special_majors):
    wb = Workbook()
    wb.remove(wb.active)
    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "第二轮候选池基于最新体检和家庭偏好，只用于讨论和核验，不作为定稿依据。"),
        ("主线", "主线仍优先公办普通学费，且暂不混入护理、医技、动物医学/兽医、临床等方向。"),
        ("专项", "医技/康复、农业、环境相关作为专项了解池；护理、动物医学/兽医暂不纳入本轮候选。"),
        ("体检", "公开层只记录色觉正常、矫正视力4.80/4.80、有近视等摘要；不保存身份信息或体检截图。"),
        ("下一步", "家庭先看主线精选和专项表，挑出愿意继续核验的组；任何入围组都必须回看第19期原页、官方侧和招生章程。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 110
    append_sheet(wb.create_sheet("01_主线精选100专业组"), main_shortlist, GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_专项医技康复农业环境"), special_groups, GROUP_FIELDS)
    append_sheet(wb.create_sheet("03_点名学校观察"), specific_watchlist, GROUP_FIELDS)
    append_sheet(wb.create_sheet("04_优先城市观察"), city_watchlist, GROUP_FIELDS)
    append_sheet(wb.create_sheet("05_主线精选组内明细"), main_majors, MAJOR_FIELDS)
    append_sheet(wb.create_sheet("06_专项组内明细"), special_majors, MAJOR_FIELDS)
    append_sheet(wb.create_sheet("07_完整候选池"), all_candidates, GROUP_FIELDS)
    ROUND2_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ROUND2_WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "/home/",
        "/private/",
        "private/",
        "private\\",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"第二轮候选公开文件包含禁止内容：{hit}")


def main():
    group_rows = read_csv(GROUP_BROWSER)
    major_rows = read_csv(MAJOR_BROWSER)
    major_signals, major_rows_by_group = aggregate_major_signals(major_rows)
    enriched_groups = merge_signals(group_rows, major_signals)

    main_pool = [row for row in enriched_groups if main_ok(row)]
    special_pool = [row for row in enriched_groups if special_ok(row)]
    specific_pool = [row for row in enriched_groups if specific_watch_tags(row)]

    main_shortlist = build_main_shortlist(main_pool)
    special_groups = build_special_groups(special_pool)
    all_candidates = build_all_candidates(main_pool, special_pool, specific_pool)
    city_watchlist = build_priority_city_watchlist(all_candidates)
    specific_watchlist = build_specific_watchlist(all_candidates)

    main_major_rows = build_major_rows(main_shortlist, major_rows)
    special_major_rows = build_major_rows(special_groups + specific_watchlist, major_rows)

    write_csv(ROUND2_GROUPS, all_candidates, GROUP_FIELDS)
    write_csv(ROUND2_MAIN_SHORTLIST, main_shortlist, GROUP_FIELDS)
    write_csv(ROUND2_SPECIAL_GROUPS, special_groups, GROUP_FIELDS)
    write_csv(ROUND2_PRIORITY_CITY, city_watchlist, GROUP_FIELDS)
    write_csv(ROUND2_SPECIFIC_WATCHLIST, specific_watchlist, GROUP_FIELDS)
    write_csv(ROUND2_MAIN_MAJORS, main_major_rows, MAJOR_FIELDS)
    write_csv(ROUND2_SPECIAL_MAJORS, special_major_rows, MAJOR_FIELDS)
    write_workbook(
        all_candidates,
        main_shortlist,
        special_groups,
        city_watchlist,
        specific_watchlist,
        main_major_rows,
        special_major_rows,
    )

    direction_group_counts = {}
    direction_major_counts = {}
    for field in [
        "机械自动化机器人专业数",
        "电子信息网络专业数",
        "计算机AI软件专业数",
        "农业动医兽医专业数",
        "动物医学兽医专业数",
        "护理助产专业数",
        "工商旅游管理专业数",
        "环境工程科学专业数",
        "医技护理康复专业数",
        "临床口腔中医暂缓专业数",
    ]:
        direction_group_counts[field] = sum(1 for row in enriched_groups if count(row, field) > 0)
        direction_major_counts[field] = sum(count(row, field) for row in enriched_groups)

    summary = {
        "status": "issue19_round2_updated_preferences_ready",
        "generated_by": "build_issue19_round2_updated_preferences.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "selection_boundary": "第二轮候选池只用于家庭讨论和入围组核验，不作为定稿依据。",
        "updated_preference_basis": [
            "色觉正常、矫正视力4.80/4.80、有近视，公开层不保存身份信息或体检截图。",
            "临床医学/口腔/中医等医学主线暂缓；护理、动物医学/兽医暂不纳入；医技/康复仅专项了解。",
            "新增机械自动化、电子信息/网络、计算机/AI/软件、农业、工商/旅游管理、环境、师范等研究方向。",
        ],
        "candidate_group_rows": len(all_candidates),
        "main_pool_rows_before_shortlist": len(main_pool),
        "main_shortlist_group_rows": len(main_shortlist),
        "special_pool_rows_before_shortlist": len(special_pool),
        "special_group_rows": len(special_groups),
        "priority_city_watchlist_rows": len(city_watchlist),
        "specific_watchlist_rows": len(specific_watchlist),
        "main_shortlist_major_rows": len(main_major_rows),
        "special_major_rows": len(special_major_rows),
        "direction_group_counts": direction_group_counts,
        "direction_major_counts": direction_major_counts,
        "workbook": str(ROUND2_WORKBOOK.relative_to(ROOT)),
        "candidate_groups_csv": str(ROUND2_GROUPS.relative_to(ROOT)),
        "main_shortlist_csv": str(ROUND2_MAIN_SHORTLIST.relative_to(ROOT)),
        "special_groups_csv": str(ROUND2_SPECIAL_GROUPS.relative_to(ROOT)),
        "priority_city_watchlist_csv": str(ROUND2_PRIORITY_CITY.relative_to(ROOT)),
        "specific_watchlist_csv": str(ROUND2_SPECIFIC_WATCHLIST.relative_to(ROOT)),
        "main_majors_csv": str(ROUND2_MAIN_MAJORS.relative_to(ROOT)),
        "special_majors_csv": str(ROUND2_SPECIAL_MAJORS.relative_to(ROOT)),
        "priority_cities": sorted(PRIORITY_CITIES),
        "hard_filters_main": [
            "公办/普通学费机器线索",
            "组内高收费或超预算专业数为0",
            "必须有逐专业明细",
            "临床/口腔/中医等暂缓方向不进主线",
            "护理、动物医学/兽医暂不纳入；医技/康复不混入主线",
        ],
    }
    ROUND2_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe(
        [
            ROUND2_GROUPS,
            ROUND2_MAIN_SHORTLIST,
            ROUND2_SPECIAL_GROUPS,
            ROUND2_PRIORITY_CITY,
            ROUND2_SPECIFIC_WATCHLIST,
            ROUND2_MAIN_MAJORS,
            ROUND2_SPECIAL_MAJORS,
            ROUND2_SUMMARY,
        ]
    )
    print(f"写出第二轮候选池：{ROUND2_GROUPS.relative_to(ROOT)}")
    print(f"写出第二轮主线精选：{ROUND2_MAIN_SHORTLIST.relative_to(ROOT)}")
    print(f"写出第二轮专项了解：{ROUND2_SPECIAL_GROUPS.relative_to(ROOT)}")
    print(f"写出第二轮点名观察：{ROUND2_SPECIFIC_WATCHLIST.relative_to(ROOT)}")
    print(f"写出第二轮工作簿：{ROUND2_WORKBOOK.relative_to(ROOT)}")
    print(
        f"主线池={len(main_pool)}；主线精选={len(main_shortlist)}；"
        f"专项池={len(special_pool)}；专项精选={len(special_groups)}；"
        f"点名观察={len(specific_watchlist)}；完整候选={len(all_candidates)}"
    )


if __name__ == "__main__":
    main()
