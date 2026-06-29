#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

FOCUS55 = EXPORTS / "issue19-round4-priority-focus55-groups.csv"
PAUSED65 = EXPORTS / "issue19-round4-priority-focus55-paused65-groups.csv"
PRIORITY120 = EXPORTS / "issue19-round4-city-gradient-priority120-groups.csv"

OUTPUT = EXPORTS / "issue19-round4-family-explanation-board.csv"
FOCUS_OUTPUT = EXPORTS / "issue19-round4-family-explanation-focus55.csv"
PAUSED_OUTPUT = EXPORTS / "issue19-round4-family-explanation-paused65.csv"
SUMMARY_OUTPUT = EXPORTS / "issue19-round4-family-explanation-board-summary.json"
WORKBOOK_OUTPUT = EXPORTS / "issue19-round4-family-explanation-board.xlsx"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
DATA_STAGE = "issue19_round4_family_explanation_board"
GENERATED_AT = "2026-06-29"

GRADIENT_QUOTA = {
    "保底观察": 1,
    "稳妥观察": 24,
    "稳冲观察": 18,
    "冲刺观察": 12,
}

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "可否进入最终志愿方案",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

FIELDS = [
    "Round4家庭说明ID",
    "来源Round4重点55组",
    "来源Round4暂缓65组",
    "来源Round4优先120组",
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "任务粒度",
    *FALSE_FIELDS,
    "家庭阅读排序",
    "入选状态",
    "是否进入40到60重点核验",
    "为什么入选",
    "为什么暂缓",
    "暂缓主因",
    "同梯度入选线",
    "距同梯度入选线分差",
    "高分暂缓标记",
    "梯度配额标签",
    "梯度内排序",
    "压缩综合分",
    "院校专业组",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组号",
    "城市",
    "城市来源",
    "城市口径提醒",
    "省份",
    "冲稳保",
    "历史稳定性等级",
    "与515关系",
    "同代码命中年份数",
    "等位分差摘要",
    "专业匹配等级",
    "专业方向命中摘要",
    "完整组内专业数",
    "接受专业数",
    "勉强调剂专业数",
    "待核后判断专业数",
    "不能接受提示专业数",
    "完整组内专业接受度摘要",
    "调剂接受风险等级",
    "调剂风险说明",
    "费用限制风险",
    "高收费或超预算专业数",
    "公办民办机器线索",
    "家庭底线属性动作",
    "结构字段阻断等级",
    "字段缺口专业行数",
    "结构或归属未闭环专业行数",
    "PDF原页待核专业行数",
    "湖北官方系统待核专业行数",
    "官网辅证命中专业行数",
    "第一闭环明细任务数",
    "第一闭环双人复核任务数",
    "高校侧刷新任务数",
    "主要阻断原因",
    "家庭最小决策动作",
    "核验最小动作",
    "下一步核验动作",
    "来源页码",
    "版面列",
    "第四轮候选ID",
    "专业组出现ID",
    "公开安全策略",
]

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/home/",
    "/var/folders/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    "file://",
    "Authorization",
    "Bearer ",
    "Cookie",
    "Set-Cookie",
    "access_token",
    "refresh_token",
    "password",
    "secret",
    "api_key",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "人工读数",
    "已确认",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
    "可排序",
]


def clean(value):
    return "" if value is None else str(value).replace("\r", " ").replace("\n", " ").strip()


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in FIELDS} for row in rows])


def write_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def stable_id(prefix, parts):
    text = "|".join(clean(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(text.encode('utf-8')).hexdigest()[:16]}"


def as_int(value):
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def as_float(value):
    try:
        return float(clean(value) or "0")
    except ValueError:
        return 0.0


def source_path(path):
    return str(path.relative_to(ROOT))


def direction_summary(row):
    mapping = [
        ("数字媒体技术", "数字媒体技术专业数"),
        ("计算机/软件/AI", "计算机AI软件专业数"),
        ("计算机类", "计算机类相关专业数"),
        ("电子信息/网络", "电子信息网络专业数"),
        ("师范", "师范类相关专业数"),
        ("机械自动化", "机械自动化机器人专业数"),
        ("环境", "环境工程科学专业数"),
        ("农业不含动物", "农业不含动物相关专业数"),
    ]
    parts = [f"{label}{as_int(row.get(field))}" for label, field in mapping if as_int(row.get(field))]
    return "；".join(parts) or "未明显命中当前优先方向"


def pause_reason(row, threshold):
    if as_int(row.get("不能接受提示专业数")):
        return "P0-存在不能接受专业提示，先暂缓"
    if clean(row.get("专业匹配等级", "")).startswith(("P3", "P4")):
        return "P1-专业匹配弱于当前优先方向"
    if clean(row.get("调剂接受风险等级", "")).startswith(("T1", "T2", "T3")):
        return "P2-调剂或特殊限制风险较高"
    if clean(row.get("核验成本等级", "")).startswith(("V0", "V1")):
        return "P3-核验成本高，先留给更高收益组"
    if threshold is not None:
        return "P4-同梯度配额外，作为备选保留"
    return "P5-综合优先级暂缓"


def pause_explanation(row, threshold):
    if clean(row.get("是否进入40到60重点核验")) == "true":
        return ""
    reason = pause_reason(row, threshold)
    pieces = [reason]
    if as_int(row.get("不能接受提示专业数")):
        pieces.append(f"组内有{as_int(row.get('不能接受提示专业数'))}个不能接受提示")
    if threshold is not None:
        diff = as_float(row.get("压缩综合分")) - threshold
        pieces.append(f"同梯度入选线{threshold:.2f}，本组{as_float(row.get('压缩综合分')):.2f}，差值{diff:.2f}")
    pieces.append(clean(row.get("压缩副理由")))
    return "；".join(part for part in pieces if part)


def transfer_explanation(row):
    risk = clean(row.get("调剂接受风险等级"))
    if risk.startswith("T0"):
        return "存在不能接受专业提示，默认不适合无条件服从调剂。"
    if risk.startswith("T1"):
        return "有体检、语种、单科、专项或章程限制待核，先看限制再谈调剂。"
    if risk.startswith("T2"):
        return "组内结构或专业归属未闭环，先核完整组内专业。"
    if risk.startswith("T3"):
        return "组内专业较多，必须逐条判断能否接受调剂。"
    return "暂未见硬阻断，但仍需家里逐专业确认接受度。"


def selected_thresholds(focus_rows):
    thresholds = {}
    for gradient, quota in GRADIENT_QUOTA.items():
        scores = [
            as_float(row.get("压缩综合分"))
            for row in focus_rows
            if clean(row.get("冲稳保")) == gradient
        ]
        if len(scores) == quota:
            thresholds[gradient] = min(scores)
    return thresholds


def build_rows():
    focus_rows = read_csv(FOCUS55)
    paused_rows = read_csv(PAUSED65)
    priority_rows = read_csv(PRIORITY120)
    priority_ids = {clean(row.get("第四轮候选ID")) for row in priority_rows}
    thresholds = selected_thresholds(focus_rows)

    combined = []
    for row in focus_rows:
        combined.append((row, True))
    for row in paused_rows:
        combined.append((row, False))

    rows = []
    for row, selected in combined:
        gradient = clean(row.get("冲稳保"))
        threshold = thresholds.get(gradient)
        score = as_float(row.get("压缩综合分"))
        threshold_text = f"{threshold:.2f}" if threshold is not None else ""
        diff_text = f"{score - threshold:.2f}" if threshold is not None else ""
        high_paused = "true" if (not selected and threshold is not None and score > threshold) else "false"
        selection_state = "重点核验" if selected else "暂缓保留"
        acceptance_summary = (
            f"接受{as_int(row.get('优先了解专业数'))}；"
            f"勉强调剂{as_int(row.get('可调剂了解专业数'))}；"
            f"待核{as_int(row.get('待核后判断专业数'))}；"
            f"不能{as_int(row.get('不能接受提示专业数'))}"
        )
        rows.append({
            "Round4家庭说明ID": stable_id("R4FAMILYEXPLAIN", [row.get("第四轮候选ID"), selection_state]),
            "来源Round4重点55组": source_path(FOCUS55),
            "来源Round4暂缓65组": source_path(PAUSED65),
            "来源Round4优先120组": source_path(PRIORITY120),
            "来源期号": SOURCE_ISSUE,
            "来源PDF_SHA256": SOURCE_PDF_SHA256,
            "生成日期": GENERATED_AT,
            "数据阶段": DATA_STAGE,
            "主表粒度": "Round4优先120组家庭阅读说明",
            "任务粒度": "院校专业组×入选或暂缓理由×调剂风险摘要",
            **{field: "false" for field in FALSE_FIELDS},
            "家庭阅读排序": "",
            "入选状态": selection_state,
            "是否进入40到60重点核验": "true" if selected else "false",
            "为什么入选": "；".join(
                part
                for part in [row.get("压缩主理由"), row.get("压缩副理由")]
                if clean(part)
            ) if selected else "",
            "为什么暂缓": pause_explanation(row, threshold),
            "暂缓主因": "" if selected else pause_reason(row, threshold),
            "同梯度入选线": threshold_text,
            "距同梯度入选线分差": diff_text,
            "高分暂缓标记": high_paused,
            "梯度配额标签": row.get("梯度配额标签", ""),
            "梯度内排序": row.get("梯度内排序", ""),
            "压缩综合分": row.get("压缩综合分", ""),
            "院校专业组": f"{row.get('院校代码')} {row.get('院校名称')} {row.get('院校专业组代码')} {row.get('专业组号')}",
            "院校代码": row.get("院校代码", ""),
            "院校名称": row.get("院校名称", ""),
            "院校专业组代码": row.get("院校专业组代码", ""),
            "专业组号": row.get("专业组号", ""),
            "城市": row.get("城市", ""),
            "城市来源": row.get("城市来源", ""),
            "城市口径提醒": "城市只是学校所在地或机器线索；实际校区必须核第19期原页、湖北官方计划和招生章程。",
            "省份": row.get("省份", ""),
            "冲稳保": gradient,
            "历史稳定性等级": row.get("历史稳定性等级", ""),
            "与515关系": row.get("与515关系", ""),
            "同代码命中年份数": row.get("同代码命中年份数", ""),
            "等位分差摘要": row.get("等位分差摘要", ""),
            "专业匹配等级": row.get("专业匹配等级", ""),
            "专业方向命中摘要": direction_summary(row),
            "完整组内专业数": row.get("专业明细行数", ""),
            "接受专业数": row.get("优先了解专业数", ""),
            "勉强调剂专业数": row.get("可调剂了解专业数", ""),
            "待核后判断专业数": row.get("待核后判断专业数", ""),
            "不能接受提示专业数": row.get("不能接受提示专业数", ""),
            "完整组内专业接受度摘要": acceptance_summary,
            "调剂接受风险等级": row.get("调剂接受风险等级", ""),
            "调剂风险说明": transfer_explanation(row),
            "费用限制风险": row.get("费用限制风险", ""),
            "高收费或超预算专业数": row.get("高收费或超预算专业数", ""),
            "公办民办机器线索": row.get("公办民办机器线索", ""),
            "家庭底线属性动作": row.get("家庭底线属性动作", ""),
            "结构字段阻断等级": row.get("结构字段阻断等级", ""),
            "字段缺口专业行数": row.get("字段缺口专业行数", ""),
            "结构或归属未闭环专业行数": row.get("结构或归属未闭环专业行数", ""),
            "PDF原页待核专业行数": row.get("PDF原页待核专业行数", ""),
            "湖北官方系统待核专业行数": row.get("湖北官方系统待核专业行数", ""),
            "官网辅证命中专业行数": row.get("官网辅证命中专业行数", ""),
            "第一闭环明细任务数": row.get("第一闭环明细任务数", ""),
            "第一闭环双人复核任务数": row.get("第一闭环双人复核任务数", ""),
            "高校侧刷新任务数": row.get("高校侧刷新任务数", ""),
            "主要阻断原因": row.get("主要阻断原因", ""),
            "家庭最小决策动作": row.get("家庭最小决策动作", ""),
            "核验最小动作": row.get("核验最小动作", ""),
            "下一步核验动作": row.get("下一步核验动作", ""),
            "来源页码": row.get("来源页码", ""),
            "版面列": row.get("版面列", ""),
            "第四轮候选ID": row.get("第四轮候选ID", ""),
            "专业组出现ID": row.get("专业组出现ID", ""),
            "公开安全策略": "该说明表只用于家庭阅读和核验排序；所有组仍需PDF原页、湖北官方侧、高校章程和完整组内专业闭环。",
        })

    rows.sort(
        key=lambda r: (
            0 if r["入选状态"] == "重点核验" else 1,
            as_int(r.get("压缩序号", "")) if r.get("压缩序号") else 999,
            clean(r.get("冲稳保")),
            -as_float(r.get("压缩综合分")),
            clean(r.get("院校代码")),
        )
    )
    focus_order = 0
    paused_order = 0
    for row in rows:
        if row["入选状态"] == "重点核验":
            focus_order += 1
            row["家庭阅读排序"] = str(focus_order)
        else:
            paused_order += 1
            row["家庭阅读排序"] = f"暂缓{paused_order}"

    missing = {row["第四轮候选ID"] for row in rows} - priority_ids
    if missing:
        raise SystemExit(f"说明表出现不在 priority120 内的候选ID：{sorted(missing)[:3]}")

    return rows


def write_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "120组说明"
    focus_ws = wb.create_sheet("重点55")
    paused_ws = wb.create_sheet("暂缓65")
    summary_ws = wb.create_sheet("说明")

    def write_sheet(sheet, sheet_rows):
        sheet.append(FIELDS)
        for row in sheet_rows:
            sheet.append([row.get(field, "") for field in FIELDS])
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = {
            "A": 24,
            "V": 42,
            "W": 42,
            "AA": 28,
            "AB": 16,
            "AC": 18,
            "AD": 36,
            "AG": 16,
            "AO": 38,
            "AX": 38,
            "BH": 42,
        }
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

    write_sheet(ws, rows)
    write_sheet(focus_ws, [row for row in rows if row["入选状态"] == "重点核验"])
    write_sheet(paused_ws, [row for row in rows if row["入选状态"] == "暂缓保留"])
    summary_ws.append(["项目", "当前值", "说明"])
    summary_items = [
        ("使用边界", "非最终", "只用于家庭阅读、讨论和核验排序，不作为定稿依据。"),
        ("重点核验组", "55", "来自 Round4 priority focus55。"),
        ("暂缓保留组", "65", "来自 Round4 paused65，仍可回捞。"),
        ("硬门禁", "全 false", "最终、推荐、写回、替代湖北官方计划均未放行。"),
    ]
    for item in summary_items:
        summary_ws.append(item)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    summary_ws.column_dimensions["A"].width = 18
    summary_ws.column_dimensions["B"].width = 18
    summary_ws.column_dimensions["C"].width = 72
    wb.save(WORKBOOK_OUTPUT)


def ensure_public_safe(paths):
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    return [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]


def main():
    rows = build_rows()
    focus_rows = [row for row in rows if row["入选状态"] == "重点核验"]
    paused_rows = [row for row in rows if row["入选状态"] == "暂缓保留"]
    write_csv(OUTPUT, rows)
    write_csv(FOCUS_OUTPUT, focus_rows)
    write_csv(PAUSED_OUTPUT, paused_rows)
    write_workbook(rows)

    unsafe_tokens = ensure_public_safe([OUTPUT, FOCUS_OUTPUT, PAUSED_OUTPUT])
    if unsafe_tokens:
        raise SystemExit(f"公开产物包含禁止词：{unsafe_tokens[:5]}")

    summary = {
        "status": "issue19_round4_family_explanation_board_not_final",
        "generated_by": "build_issue19_round4_family_explanation_board.py",
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "generated_at": GENERATED_AT,
        "source_focus55": source_path(FOCUS55),
        "source_paused65": source_path(PAUSED65),
        "source_priority120": source_path(PRIORITY120),
        "output_table": source_path(OUTPUT),
        "output_focus55": source_path(FOCUS_OUTPUT),
        "output_paused65": source_path(PAUSED_OUTPUT),
        "output_workbook": source_path(WORKBOOK_OUTPUT),
        "row_count": len(rows),
        "focus_count": len(focus_rows),
        "paused_count": len(paused_rows),
        "unique_candidate_count": len({row["第四轮候选ID"] for row in rows}),
        "state_counts": dict(Counter(row["入选状态"] for row in rows)),
        "focus_gradient_counts": dict(Counter(row["冲稳保"] for row in focus_rows)),
        "paused_reason_counts": dict(Counter(row["暂缓主因"] for row in paused_rows)),
        "high_score_paused_count": sum(row["高分暂缓标记"] == "true" for row in paused_rows),
        "acceptance_totals": {
            "接受专业数": sum(as_int(row["接受专业数"]) for row in rows),
            "勉强调剂专业数": sum(as_int(row["勉强调剂专业数"]) for row in rows),
            "待核后判断专业数": sum(as_int(row["待核后判断专业数"]) for row in rows),
            "不能接受提示专业数": sum(as_int(row["不能接受提示专业数"]) for row in rows),
        },
        "all_false_gate": all(row[field] == "false" for row in rows for field in FALSE_FIELDS),
        "public_boundary": "该表把 Round4 120 组解释为重点核验或暂缓保留；所有组仍非最终，不能作为志愿定稿依据。",
    }
    write_json(SUMMARY_OUTPUT, summary)


if __name__ == "__main__":
    main()
