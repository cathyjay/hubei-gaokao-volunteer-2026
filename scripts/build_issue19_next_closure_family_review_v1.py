#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

FIRST_PAGE = EXPORTS / "issue19-closure-and-shortlist-v1-first-closure-page-sides.csv"
FIRST_DETAIL = EXPORTS / "issue19-closure-and-shortlist-v1-first-closure-detail-tasks.csv"
FIRST_FIELD = WORKING / "issue19-stable-foundation-first-closure-field-confirmation-public-ledger.csv"
PRIORITY_GROUPS = EXPORTS / "issue19-closure-and-shortlist-v1-priority55-groups.csv"
PRIORITY_MAJORS = EXPORTS / "issue19-closure-and-shortlist-v1-priority55-major-details.csv"

OUT_PREFIX = "issue19-next-closure-family-review-v1"
OUT_WORKBOOK = EXPORTS / f"{OUT_PREFIX}.xlsx"
OUT_PAGE_PACK = EXPORTS / f"{OUT_PREFIX}-first-closure-page-pack.csv"
OUT_ACTION_PACK = EXPORTS / f"{OUT_PREFIX}-first-closure-action-pack.csv"
OUT_TASK_PACK = EXPORTS / f"{OUT_PREFIX}-first-closure-task-status.csv"
OUT_GROUP_REVIEW = EXPORTS / f"{OUT_PREFIX}-priority55-group-review.csv"
OUT_MAJOR_REVIEW = EXPORTS / f"{OUT_PREFIX}-priority55-major-review.csv"
OUT_SUMMARY = EXPORTS / f"{OUT_PREFIX}-summary.json"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

PAGE_PACK_FIELDS = [
    "核验包序号",
    "页码版面键",
    "来源页码",
    "版面列",
    "执行泳道",
    "人工核验批次",
    "页列任务数",
    "涉及院校数",
    "涉及专业组数",
    "涉及专业行数",
    "字段范围摘要",
    "冲突状态摘要",
    "PDFOCR提示任务数",
    "机器坐标提示任务数",
    "高校辅证线索任务数",
    "无候选需看图任务数",
    "双人复核任务数",
    "PDF原页状态",
    "湖北官方状态",
    "高校辅证状态",
    "字段写回状态",
    "公开核验动作",
    "私有填写内容",
    "完成条件",
    "下一步",
]

ACTION_PACK_FIELDS = [
    "核验包编号",
    "包类型",
    "物理批次序号",
    "执行顺序",
    "页面类别",
    "页列优先级",
    "来源期号",
    "来源文档哈希",
    "来源页码",
    "版面列",
    "页码版面",
    "候选动作桶",
    "字段动作泳道",
    "包内任务数",
    "原页必核任务数",
    "湖北官方必核任务数",
    "高校辅证需复核任务数",
    "双人复核任务数",
    "人工看图任务数",
    "候选冲突任务数",
    "一致待官核任务数",
    "候选确认任务数",
    "主要字段范围",
    "必须完成证据步骤",
    "原页记录状态",
    "湖北官方记录状态",
    "高校辅证记录状态",
    "双人复核状态",
    "三方一致性状态",
    "字段写回评估状态",
    "公开门禁状态",
    "阻断原因",
    "下一步动作",
    "完成条件",
    "私有工作台哈希",
    "关联明细任务编号集合",
    "关联公开账本编号集合",
]

TASK_PACK_FIELDS = [
    "核验任务序号",
    "页码版面键",
    "执行泳道",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业代号",
    "专业名称短摘",
    "字段范围",
    "任务来源类型",
    "公开冲突桶",
    "候选提示桶",
    "PDFOCR提示",
    "机器坐标提示",
    "高校辅证线索",
    "是否需要双人复核",
    "PDF原页记录状态",
    "湖北官方记录状态",
    "高校辅证记录状态",
    "字段写回状态",
    "人工动作",
    "完成条件",
    "专业行ID",
    "第一闭环任务ID",
]

GROUP_REVIEW_FIELDS = [
    "下一轮序号",
    "下一轮建议状态",
    "组级核验动作",
    "家庭讨论优先级",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "城市",
    "冲稳保",
    "历史线索分层",
    "专业明细行数",
    "优先了解专业数",
    "可调剂了解专业数",
    "待核后判断专业数",
    "不能接受提示专业数",
    "特殊限制专业数",
    "高费或中外合作提示数",
    "医学护理动物兽医提示数",
    "农林园艺提示数",
    "字段缺口专业数",
    "100%核验专业数",
    "建议放入6专业讨论专业数",
    "可服从调剂初判",
    "组内调剂判断状态",
    "调剂风险摘要",
    "家庭先决策问题",
    "家庭逐专业填写说明",
    "入选理由",
    "下一步核验动作",
    "是否可作为定稿依据",
    "来源页码",
    "版面列",
    "第四轮候选ID",
]

MAJOR_REVIEW_FIELDS = [
    "下一轮组序号",
    "下一轮建议状态",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "城市",
    "冲稳保",
    "专业组内专业序号",
    "专业代号",
    "专业名称及备注",
    "机器家庭接受度建议",
    "家庭可选项提示",
    "家庭最终接受度待填",
    "家庭最终接受度可填值",
    "是否建议放入6专业讨论",
    "建议6专业理由",
    "调剂风险提示",
    "服从调剂影响",
    "组内调剂判断状态",
    "家庭要回答的问题",
    "专业偏好方向",
    "专业角色判断",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "来源页码",
    "版面列",
    "专业行ID",
    "第四轮候选ID",
]


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, fields, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def digest(value):
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:16]


def yes(value):
    return str(value).strip().lower() == "true"


def int_value(value):
    try:
        return int(str(value).strip())
    except Exception:
        return 0


def money_value(value):
    try:
        return int(float(str(value).strip()))
    except Exception:
        return 0


def compact_counter(counter):
    parts = []
    for key, value in counter.most_common():
        if not key:
            key = "空"
        parts.append(f"{key}×{value}")
    return "；".join(parts)


def page_batch(lane):
    if lane.startswith("E0"):
        return "B1-E0冲突异常双人核页"
    if lane.startswith("E1"):
        return "B2-E1计划数补缺或偏大核页"
    if lane.startswith("E2"):
        return "B3-E2官网未匹配归属核页"
    return "B9-待人工确认"


def page_public_action(lane):
    if lane.startswith("E0"):
        return "先双人回看PDF原页，核专业行归属、计划数/学费/选科；再查湖北官方侧；最后用高校官网解释差异。"
    if lane.startswith("E1"):
        return "先回PDF原页补计划数或异常字段；再查湖北官方侧确认；必要时用高校官网补缺线索定位。"
    if lane.startswith("E2"):
        return "先确认专业名、专业代号和专业组边界是否串读；再核字段值和高校辅证匹配关系。"
    return "人工确认页列任务类型后再核。"


def page_completion(lane):
    if lane.startswith("E0"):
        return "同一页列冲突字段完成PDF原页读数、湖北官方侧字段值、高校辅证复核和第二人复核，仍不自动写回。"
    if lane.startswith("E1"):
        return "计划数/偏大字段完成PDF原页读数和湖北官方侧确认；若与高校辅证不一致，升级双人复核。"
    if lane.startswith("E2"):
        return "专业归属、组边界和字段三方一致性完成记录；若发现串读，整页列升级核验。"
    return "完成PDF原页、湖北官方侧和必要高校辅证记录。"


def action_public_type(bucket):
    if bucket.startswith("P0"):
        return "P0候选冲突"
    if bucket.startswith("P1"):
        return "P1缺候选或需看图"
    if bucket.startswith("P2"):
        return "P2候选一致但未闭环"
    if bucket.startswith("P3"):
        return "P3候选需人工确认"
    return "P9待判断"


def action_next_step(bucket):
    if bucket.startswith("P0"):
        return "双人先核PDF原页和湖北官方侧，再用高校辅证解释差异；不得直接写回。"
    if bucket.startswith("P1"):
        return "先看PDF原页补读缺候选字段，再查湖北官方侧确认；机器坐标只作定位提示。"
    if bucket.startswith("P2"):
        return "走快速三方闭环，但仍必须查湖北官方侧；一致也不能跳过官方计划。"
    if bucket.startswith("P3"):
        return "人工确认候选字段归属、字段列和专业行边界，再进入官方侧核验。"
    return "先判定动作桶后再核。"


def action_completion(bucket):
    if bucket.startswith("P0"):
        return "包内任务完成PDF原页、湖北官方侧、高校辅证和双人复核记录，三方一致性明确后才进入写回评估。"
    if bucket.startswith("P1"):
        return "包内缺候选字段完成PDF原页补读和湖北官方侧确认；仍冲突则升级双人复核。"
    if bucket.startswith("P2"):
        return "包内候选一致字段完成湖北官方侧确认和三方记录；不自动生成推荐依据。"
    if bucket.startswith("P3"):
        return "包内候选完成归属确认、原页核验和湖北官方侧记录；发现串读则升级整页列。"
    return "完成必要三方记录。"


def build_action_packages(field_rows, max_tasks=10):
    groups = defaultdict(list)
    for row in field_rows:
        key = (row.get("页码版面键", ""), row.get("PDFOCR提示审阅桶", "P9-待判断"))
        groups[key].append(row)

    packages = []
    sorted_groups = sorted(
        groups.items(),
        key=lambda item: (
            min(int_value(row.get("第一闭环执行顺序", "")) for row in item[1]),
            item[0][1],
        ),
    )
    for (page_key, bucket), rows in sorted_groups:
        rows = sorted(
            rows,
            key=lambda row: (
                int_value(row.get("字段确认公开账本总序", "")),
                row.get("第一闭环字段确认公开账本ID", ""),
            ),
        )
        for chunk_index, start in enumerate(range(0, len(rows), max_tasks), start=1):
            chunk = rows[start : start + max_tasks]
            first = chunk[0]
            package_seed = "|".join(row.get("第一闭环字段确认公开账本ID", "") for row in chunk)
            packages.append(
                {
                    "核验包编号": f"FCACT-{digest(package_seed)}",
                    "包类型": action_public_type(bucket),
                    "物理批次序号": chunk_index,
                    "执行顺序": min(int_value(row.get("第一闭环执行顺序", "")) for row in chunk),
                    "页面类别": first.get("执行泳道", ""),
                    "页列优先级": first.get("第一闭环页列优先级", ""),
                    "来源期号": first.get("来源期号", SOURCE_ISSUE),
                    "来源文档哈希": first.get("来源PDF_SHA256", SOURCE_PDF_SHA256),
                    "来源页码": first.get("来源页码", ""),
                    "版面列": first.get("版面列", ""),
                    "页码版面": page_key,
                    "候选动作桶": bucket,
                    "字段动作泳道": compact_counter(Counter(row.get("人工核验泳道", "") for row in chunk)),
                    "包内任务数": len(chunk),
                    "原页必核任务数": len(chunk),
                    "湖北官方必核任务数": len(chunk),
                    "高校辅证需复核任务数": sum(1 for row in chunk if row.get("高校辅证私有记录状态", "") == "pending_private_school_reading"),
                    "双人复核任务数": sum(1 for row in chunk if yes(row.get("是否需要双人复核", ""))),
                    "人工看图任务数": sum(1 for row in chunk if yes(row.get("是否需要人工直接看图", ""))),
                    "候选冲突任务数": sum(1 for row in chunk if row.get("PDFOCR提示审阅桶", "").startswith("P0")),
                    "一致待官核任务数": sum(1 for row in chunk if row.get("PDFOCR提示审阅桶", "").startswith("P2")),
                    "候选确认任务数": sum(1 for row in chunk if row.get("PDFOCR提示审阅桶", "").startswith("P3")),
                    "主要字段范围": compact_counter(Counter(row.get("字段名", "") for row in chunk)),
                    "必须完成证据步骤": "PDF原页记录；湖北官方侧记录；必要高校辅证；冲突任务双人复核；三方一致性记录。",
                    "原页记录状态": compact_counter(Counter(row.get("PDF原页私有记录状态", "") for row in chunk)),
                    "湖北官方记录状态": compact_counter(Counter(row.get("湖北官方私有记录状态", "") for row in chunk)),
                    "高校辅证记录状态": compact_counter(Counter(row.get("高校辅证私有记录状态", "") for row in chunk)),
                    "双人复核状态": compact_counter(Counter(row.get("双人复核公开状态", "") for row in chunk)),
                    "三方一致性状态": compact_counter(Counter(row.get("三方字段一致性公开状态", "") for row in chunk)),
                    "字段写回评估状态": compact_counter(Counter(row.get("字段事实写回评估状态", "") for row in chunk)),
                    "公开门禁状态": "not_final_blocked_until_private_readings_complete",
                    "阻断原因": "公开包只排核验顺序；字段读数、候选值、截图、坐标和人工备注必须留在私有工作台。",
                    "下一步动作": action_next_step(bucket),
                    "完成条件": action_completion(bucket),
                    "私有工作台哈希": digest(package_seed),
                    "关联明细任务编号集合": "；".join(row.get("稳定基座第一闭环明细任务ID", "") for row in chunk),
                    "关联公开账本编号集合": "；".join(row.get("第一闭环字段确认公开账本ID", "") for row in chunk),
                }
            )
    return packages


def group_status(group, majors):
    transfer = group.get("调剂初判", "")
    if "结构异常" in transfer:
        return "先核页后讨论"
    if "特殊限制" in transfer:
        return "先核限制后讨论"
    if "学费字段待核" in transfer:
        return "先核费用后讨论"
    if sum(1 for row in majors if major_acceptance(row) == "可接受") > 0:
        return "优先家庭讨论"
    return "可讨论但先看调剂"


def group_action(status):
    return {
        "优先家庭讨论": "家庭先看完整组内专业，标出可接受/勉强接受/不能接受；保留后再做100%原页和湖北官方侧核验。",
        "可讨论但先看调剂": "先判断是否能接受组内非偏好专业调剂，再决定是否投入核页时间。",
        "先核页后讨论": "先核PDF原页结构、专业组边界和串读风险，再进入家庭调剂讨论。",
        "先核限制后讨论": "先核体检、语种、单科、专项或备注限制；限制不能接受则整组暂缓。",
        "先核费用后讨论": "先核学费和合作办学属性；超预算或高收费不接受则暂缓。",
    }.get(status, "先补证再讨论。")


def family_priority(group, status):
    if status == "先核页后讨论":
        return "P0-先核页"
    if status in {"先核限制后讨论", "先核费用后讨论"}:
        return "P1-先核限制费用"
    gradient = group.get("冲稳保", "")
    if "保底" in gradient:
        return "P0-保底样本讨论"
    if "稳妥" in gradient:
        return "P1-稳妥主力讨论"
    if "稳冲" in gradient:
        return "P2-稳冲择优讨论"
    if "冲刺" in gradient:
        return "P3-冲刺择优讨论"
    return "P9-待判断"


def major_acceptance(row):
    name = row.get("专业名称及备注", "")
    risk = row.get("专业风险类型", "")
    fee = money_value(row.get("学费OCR候选", ""))
    if has_medical(name) or "医学" in risk:
        return "不能接受"
    if fee > 15000 or "中外合作" in name or "高收费" in name:
        return "不能接受"
    current = row.get("家庭接受度初判V1", "")
    if current.startswith("优先了解"):
        return "可接受"
    if current.startswith("可作为调剂了解"):
        return "勉强接受"
    if current.startswith("待核后判断"):
        return "待核后判断"
    return "待家庭判断"


def family_option(row):
    acceptance = major_acceptance(row)
    if acceptance in {"可接受", "勉强接受", "不能接受", "待核后判断"}:
        return acceptance
    return "待家庭确认"


def recommend_for_six(row):
    if major_acceptance(row) != "可接受":
        return False
    role = row.get("专业角色判断", "")
    preference = row.get("专业偏好方向", "")
    direction = row.get("第四轮方向标签", "")
    return bool(preference or "主线偏好" in role or direction)


def six_reason(row):
    if not recommend_for_six(row):
        if major_acceptance(row) == "勉强接受":
            return "可作为调剂容忍项，暂不优先放入6专业。"
        if major_acceptance(row) == "待核后判断":
            return "限制、字段或费用核清前不放入6专业。"
        if major_acceptance(row) == "不能接受":
            return "触发家庭底线或高风险，不放入6专业。"
        return "待家庭确认后再判断。"
    preference = row.get("专业偏好方向", "")
    if preference:
        return f"命中当前偏好方向：{preference}。"
    role = row.get("专业角色判断", "")
    return role or "机器初判为可接受专业。"


def transfer_impact(row):
    acceptance = major_acceptance(row)
    if acceptance == "不能接受":
        return "若该专业在组内且可能调剂到，服从调剂暂不安全；需整组暂缓或确认不服从风险。"
    if acceptance == "待核后判断":
        return "限制、费用或字段未闭环前，不能判断是否可接受调剂。"
    if acceptance == "勉强接受":
        return "可作为调剂容忍项，但必须由家庭确认是否真的能接受。"
    if acceptance == "可接受":
        return "可作为6专业或调剂可接受项，仍需原页、湖北官方侧和章程闭环。"
    return "待家庭确认。"


def group_transfer_status(accept_counter):
    if accept_counter.get("不能接受", 0):
        return "存在不能接受专业，服从调剂暂不安全"
    if accept_counter.get("待核后判断", 0) or accept_counter.get("待家庭判断", 0):
        return "存在待核后判断专业，需先核限制/字段再判断调剂"
    if accept_counter.get("勉强接受", 0):
        return "可进入调剂讨论，但需确认勉强接受专业是否真能接受"
    if accept_counter.get("可接受", 0):
        return "组内初判可讨论服从调剂，仍需三方核验"
    return "缺少可判断专业，暂不讨论服从调剂"


def group_adjustment_hint(accept_counter):
    status = group_transfer_status(accept_counter)
    if status.startswith("存在不能接受"):
        return "暂不建议服从调剂"
    if status.startswith("存在待核"):
        return "待核后再判断"
    if status.startswith("可进入"):
        return "可讨论服从调剂"
    if status.startswith("组内初判"):
        return "可讨论服从调剂"
    return "待家庭确认"


def has_medical(name):
    keywords = ["临床医学", "口腔医学", "护理", "药学", "中医学", "医学影像", "医学检验", "康复治疗", "动物医学", "兽医"]
    return any(keyword in name for keyword in keywords)


def is_agri(name):
    return any(keyword in name for keyword in ["农学", "园林", "园艺", "植物", "水产", "林学", "茶学"])


def is_special(row):
    risk = row.get("专业风险类型", "")
    name = row.get("专业名称及备注", "")
    return bool(risk) or "不招" in name or "单科" in name or "语种" in name


def family_question(row):
    name = row.get("专业名称及备注", "")
    risk = row.get("专业风险类型", "")
    questions = []
    if major_acceptance(row) == "可接受":
        questions.append("是否愿意把该专业作为组内优先专业")
    elif major_acceptance(row) == "勉强接受":
        questions.append("如果被调剂到该专业是否能接受")
    elif major_acceptance(row) == "待核后判断":
        questions.append("先核限制后再判断是否接受")
    if "体检" in risk or "色觉" in risk or "不招色" in name:
        questions.append("确认体检/色觉限制是否影响")
    if "语种" in risk or "单科" in risk or "英语" in name:
        questions.append("确认语种或单科要求是否满足")
    if money_value(row.get("学费OCR候选", "")) > 15000 or "中外合作" in name:
        questions.append("确认费用和合作办学是否接受")
    if has_medical(name):
        questions.append("当前不学医底线下原则上不接受")
    if is_agri(name):
        questions.append("确认农林园艺方向是否接受")
    return "；".join(dict.fromkeys(questions)) or "家庭直接标记可接受/勉强接受/不能接受"


def build_page_outputs(page_rows, detail_rows, field_rows):
    details_by_page = defaultdict(list)
    for row in detail_rows:
        details_by_page[row.get("页码版面键", "")].append(row)

    fields_by_task = {row.get("稳定基座第一闭环明细任务ID", ""): row for row in field_rows}
    fields_by_page = defaultdict(list)
    for row in field_rows:
        fields_by_page[row.get("页码版面键", "")].append(row)

    page_pack = []
    task_pack = []
    for page in page_rows:
        key = page.get("页码版面键", "")
        lane = page.get("执行泳道", "")
        page_fields = fields_by_page.get(key, [])
        page_details = details_by_page.get(key, [])
        field_counter = Counter(row.get("字段名", "") for row in page_fields)
        bucket_counter = Counter(row.get("候选提示综合桶", "") for row in page_fields)
        page_pack.append(
            {
                "核验包序号": page.get("执行顺序", ""),
                "页码版面键": key,
                "来源页码": page.get("来源页码", ""),
                "版面列": page.get("版面列", ""),
                "执行泳道": lane,
                "人工核验批次": page_batch(lane),
                "页列任务数": page.get("页列总任务数", ""),
                "涉及院校数": len({row.get("院校代码", "") for row in page_details if row.get("院校代码", "")}),
                "涉及专业组数": len({row.get("院校专业组代码OCR规范化", "") for row in page_details if row.get("院校专业组代码OCR规范化", "")}),
                "涉及专业行数": len({row.get("专业行ID", "") for row in page_details if row.get("专业行ID", "")}),
                "字段范围摘要": compact_counter(field_counter),
                "冲突状态摘要": compact_counter(bucket_counter),
                "PDFOCR提示任务数": sum(1 for row in page_fields if yes(row.get("是否有PDFOCR提示", ""))),
                "机器坐标提示任务数": sum(1 for row in page_fields if yes(row.get("是否有机器坐标提示", ""))),
                "高校辅证线索任务数": sum(1 for row in page_fields if yes(row.get("是否有高校辅证线索", ""))),
                "无候选需看图任务数": sum(1 for row in page_fields if yes(row.get("是否需要人工直接看图", ""))),
                "双人复核任务数": sum(1 for row in page_fields if yes(row.get("是否需要双人复核", ""))),
                "PDF原页状态": compact_counter(Counter(row.get("PDF原页私有记录状态", "") for row in page_fields)),
                "湖北官方状态": compact_counter(Counter(row.get("湖北官方私有记录状态", "") for row in page_fields)),
                "高校辅证状态": compact_counter(Counter(row.get("高校辅证私有记录状态", "") for row in page_fields)),
                "字段写回状态": compact_counter(Counter(row.get("字段事实写回评估状态", "") for row in page_fields)),
                "公开核验动作": page_public_action(lane),
                "私有填写内容": "PDF原页读数、湖北官方字段值、高校官网字段值、复核结论、复核备注只写入Git忽略的私有工作台。",
                "完成条件": page_completion(lane),
                "下一步": page.get("下一步", ""),
            }
        )

        for detail in page_details:
            task_id = detail.get("稳定基座第一闭环明细任务ID", "")
            field = fields_by_task.get(task_id, {})
            task_pack.append(
                {
                    "核验任务序号": detail.get("执行顺序", ""),
                    "页码版面键": key,
                    "执行泳道": lane,
                    "院校代码": detail.get("院校代码", ""),
                    "院校名称": detail.get("院校名称OCR", ""),
                    "院校专业组代码": detail.get("院校专业组代码OCR规范化", ""),
                    "专业代号": detail.get("专业代号OCR", ""),
                    "专业名称短摘": detail.get("专业名称及备注短摘", ""),
                    "字段范围": field.get("字段名", detail.get("字段名", "")),
                    "任务来源类型": detail.get("任务来源类型", ""),
                    "公开冲突桶": field.get("PDFOCR与高校辅证关系桶", ""),
                    "候选提示桶": field.get("候选提示综合桶", ""),
                    "PDFOCR提示": field.get("PDFOCR提示记录状态", ""),
                    "机器坐标提示": field.get("机器坐标提示记录状态", ""),
                    "高校辅证线索": field.get("是否有高校辅证线索", ""),
                    "是否需要双人复核": field.get("是否需要双人复核", detail.get("是否需要双人复核", "")),
                    "PDF原页记录状态": field.get("PDF原页私有记录状态", detail.get("PDF原页核页状态", "")),
                    "湖北官方记录状态": field.get("湖北官方私有记录状态", detail.get("湖北官方系统或省招办计划核验状态", "")),
                    "高校辅证记录状态": field.get("高校辅证私有记录状态", detail.get("高校官网或招生章程辅证状态", "")),
                    "字段写回状态": field.get("字段事实写回评估状态", detail.get("字段事实写回状态", "")),
                    "人工动作": field.get("人工核验动作", detail.get("下一步", "")),
                    "完成条件": "完成PDF原页、湖北官方侧和必要高校辅证记录；冲突任务需双人复核；仍需人工判断是否写回。",
                    "专业行ID": detail.get("专业行ID", ""),
                    "第一闭环任务ID": task_id,
                }
            )
    return page_pack, task_pack


def build_family_outputs(group_rows, major_rows):
    majors_by_group = defaultdict(list)
    for row in major_rows:
        key = row.get("第四轮候选ID") or f"{row.get('院校代码')}|{row.get('院校专业组代码')}"
        majors_by_group[key].append(row)

    group_review = []
    major_review = []
    for index, group in enumerate(group_rows, start=1):
        key = group.get("第四轮候选ID") or f"{group.get('院校代码')}|{group.get('院校专业组代码')}"
        majors = majors_by_group.get(key, [])
        status = group_status(group, majors)
        accept_counter = Counter(major_acceptance(row) for row in majors)
        risk_counter = Counter(row.get("调剂风险初判V1", "") for row in majors)
        special_count = sum(1 for row in majors if is_special(row))
        expensive_count = sum(1 for row in majors if money_value(row.get("学费OCR候选", "")) > 15000 or "中外合作" in row.get("专业名称及备注", ""))
        medical_count = sum(1 for row in majors if has_medical(row.get("专业名称及备注", "")))
        agri_count = sum(1 for row in majors if is_agri(row.get("专业名称及备注", "")))
        field_gap_count = sum(1 for row in majors if row.get("字段缺口字段", ""))
        full_review_count = sum(1 for row in majors if yes(row.get("是否必须100%人工核验", "")) or row.get("人工核验强度", "").startswith("H0"))
        recommend_six_count = sum(1 for row in majors if recommend_for_six(row))

        family_questions = []
        if accept_counter.get("勉强接受", 0):
            family_questions.append("是否能接受组内非偏好专业调剂")
        if special_count:
            family_questions.append("体检/语种/单科/专项限制是否满足")
        if expensive_count:
            family_questions.append("费用或中外合作是否接受")
        if medical_count:
            family_questions.append("不学医底线下是否直接排除相关专业")
        if agri_count:
            family_questions.append("农林园艺方向是否接受")
        if field_gap_count:
            family_questions.append("字段缺口补齐前不进入最终排序")

        group_review.append(
            {
                "下一轮序号": index,
                "下一轮建议状态": status,
                "组级核验动作": group_action(status),
                "家庭讨论优先级": family_priority(group, status),
                "院校代码": group.get("院校代码", ""),
                "院校名称": group.get("院校名称", ""),
                "院校专业组代码": group.get("院校专业组代码", ""),
                "城市": group.get("城市", ""),
                "冲稳保": group.get("冲稳保", ""),
                "历史线索分层": group.get("历史线索分层", ""),
                "专业明细行数": len(majors),
                "优先了解专业数": accept_counter.get("可接受", 0),
                "可调剂了解专业数": accept_counter.get("勉强接受", 0),
                "待核后判断专业数": accept_counter.get("待核后判断", 0),
                "不能接受提示专业数": accept_counter.get("不能接受", 0),
                "特殊限制专业数": special_count,
                "高费或中外合作提示数": expensive_count,
                "医学护理动物兽医提示数": medical_count,
                "农林园艺提示数": agri_count,
                "字段缺口专业数": field_gap_count,
                "100%核验专业数": full_review_count,
                "建议放入6专业讨论专业数": recommend_six_count,
                "可服从调剂初判": group_adjustment_hint(accept_counter),
                "组内调剂判断状态": group_transfer_status(accept_counter),
                "调剂风险摘要": compact_counter(risk_counter),
                "家庭先决策问题": "；".join(family_questions) or "直接逐专业标记可接受/勉强接受/不能接受",
                "家庭逐专业填写说明": "逐专业只填：可接受、勉强接受、不能接受、待核后判断；全组无不能接受且待核项闭环后，才讨论服从调剂。",
                "入选理由": group.get("入选或暂缓理由", ""),
                "下一步核验动作": group.get("下一步核验动作", ""),
                "是否可作为定稿依据": "false",
                "来源页码": group.get("来源页码", ""),
                "版面列": group.get("版面列", ""),
                "第四轮候选ID": group.get("第四轮候选ID", ""),
            }
        )

        for major in majors:
            major_review.append(
                {
                    "下一轮组序号": index,
                    "下一轮建议状态": status,
                    "院校代码": major.get("院校代码", ""),
                    "院校名称": major.get("院校名称", ""),
                    "院校专业组代码": major.get("院校专业组代码", ""),
                    "城市": major.get("城市", ""),
                    "冲稳保": major.get("冲稳保", ""),
                    "专业组内专业序号": major.get("专业组内专业序号", ""),
                    "专业代号": major.get("专业代号", ""),
                    "专业名称及备注": major.get("专业名称及备注", ""),
                    "机器家庭接受度建议": major_acceptance(major),
                    "家庭可选项提示": family_option(major),
                    "家庭最终接受度待填": "待家庭确认",
                    "家庭最终接受度可填值": "可接受/勉强接受/不能接受/待核后判断",
                    "是否建议放入6专业讨论": "true" if recommend_for_six(major) else "false",
                    "建议6专业理由": six_reason(major),
                    "调剂风险提示": major.get("调剂风险初判V1", ""),
                    "服从调剂影响": transfer_impact(major),
                    "组内调剂判断状态": group_transfer_status(accept_counter),
                    "家庭要回答的问题": family_question(major),
                    "专业偏好方向": major.get("专业偏好方向", ""),
                    "专业角色判断": major.get("专业角色判断", ""),
                    "专业风险类型": major.get("专业风险类型", ""),
                    "再选科目OCR候选": major.get("再选科目OCR候选", ""),
                    "专业计划数OCR候选": major.get("专业计划数OCR候选", ""),
                    "学费OCR候选": major.get("学费OCR候选", ""),
                    "字段缺口字段": major.get("字段缺口字段", ""),
                    "人工核验优先级": major.get("人工核验优先级", ""),
                    "人工核验强度": major.get("人工核验强度", ""),
                    "是否必须100%人工核验": major.get("是否必须100%人工核验", ""),
                    "PDF原页核验状态": major.get("PDF原页核验状态", ""),
                    "湖北官方系统核验状态": major.get("湖北官方系统核验状态", ""),
                    "高校官网证据匹配状态": major.get("高校官网证据匹配状态", ""),
                    "来源页码": major.get("来源页码", ""),
                    "版面列": major.get("版面列", ""),
                    "专业行ID": major.get("专业行ID", ""),
                    "第四轮候选ID": major.get("第四轮候选ID", ""),
                }
            )
    return group_review, major_review


def add_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append(fields)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 4, 12), 42)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def write_workbook(summary, page_pack, action_pack, task_pack, group_review, major_review):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    summary_rows = []
    for section, values in summary.items():
        if isinstance(values, dict):
            for key, value in values.items():
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                summary_rows.append({"板块": section, "指标": key, "值": value})
        else:
            summary_rows.append({"板块": "summary", "指标": section, "值": values})

    add_sheet(wb, "00_摘要", summary_rows, ["板块", "指标", "值"])
    add_sheet(wb, "01_第一闭环页列核验包", page_pack, PAGE_PACK_FIELDS)
    add_sheet(wb, "02_第一闭环64小包", action_pack, ACTION_PACK_FIELDS)
    add_sheet(wb, "03_第一闭环任务状态", task_pack, TASK_PACK_FIELDS)
    add_sheet(wb, "04_下一轮55组讨论", group_review, GROUP_REVIEW_FIELDS)
    add_sheet(wb, "05_重点组完整专业", major_review, MAJOR_REVIEW_FIELDS)
    wb.save(OUT_WORKBOOK)


def main():
    page_rows = read_csv(FIRST_PAGE)
    detail_rows = read_csv(FIRST_DETAIL)
    field_rows = read_csv(FIRST_FIELD)
    group_rows = read_csv(PRIORITY_GROUPS)
    major_rows = read_csv(PRIORITY_MAJORS)

    page_pack, task_pack = build_page_outputs(page_rows, detail_rows, field_rows)
    action_pack = build_action_packages(field_rows)
    group_review, major_review = build_family_outputs(group_rows, major_rows)

    write_csv(OUT_PAGE_PACK, PAGE_PACK_FIELDS, page_pack)
    write_csv(OUT_ACTION_PACK, ACTION_PACK_FIELDS, action_pack)
    write_csv(OUT_TASK_PACK, TASK_PACK_FIELDS, task_pack)
    write_csv(OUT_GROUP_REVIEW, GROUP_REVIEW_FIELDS, group_review)
    write_csv(OUT_MAJOR_REVIEW, MAJOR_REVIEW_FIELDS, major_review)

    status_counts = Counter(row["下一轮建议状态"] for row in group_review)
    page_batch_counts = Counter(row["人工核验批次"] for row in page_pack)
    major_accept_counts = Counter(row["机器家庭接受度建议"] for row in major_review)
    group_transfer_counts = Counter(row["可服从调剂初判"] for row in group_review)
    group_transfer_status_counts = Counter(row["组内调剂判断状态"] for row in group_review)
    major_six_discussion_counts = Counter(row["是否建议放入6专业讨论"] for row in major_review)
    major_transfer_impact_counts = Counter(row["服从调剂影响"] for row in major_review)
    summary = {
        "status": "issue19_next_closure_family_review_v1_ready_not_final",
        "generated_by": Path(__file__).name,
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "usage_boundary": "用于第一闭环人工核验和55组家庭讨论准备；不确认字段事实，不作为最终志愿方案。",
        "outputs": {
            "workbook": str(OUT_WORKBOOK.relative_to(ROOT)),
            "first_closure_page_pack": str(OUT_PAGE_PACK.relative_to(ROOT)),
            "first_closure_action_pack": str(OUT_ACTION_PACK.relative_to(ROOT)),
            "first_closure_task_status": str(OUT_TASK_PACK.relative_to(ROOT)),
            "priority55_group_review": str(OUT_GROUP_REVIEW.relative_to(ROOT)),
            "priority55_major_review": str(OUT_MAJOR_REVIEW.relative_to(ROOT)),
        },
        "first_closure": {
            "page_side_count": len(page_pack),
            "action_package_count": len(action_pack),
            "task_count": len(task_pack),
            "page_batch_counts": dict(page_batch_counts),
            "action_package_counts": dict(Counter(row["包类型"] for row in action_pack)),
            "field_writeback_ready_count": 0,
            "recommendation_basis_allowed_count": 0,
        },
        "family_review": {
            "group_count": len(group_review),
            "major_detail_count": len(major_review),
            "next_status_counts": dict(status_counts),
            "major_acceptance_counts": dict(major_accept_counts),
            "group_transfer_decision_counts": dict(group_transfer_counts),
            "group_transfer_status_counts": dict(group_transfer_status_counts),
            "major_six_discussion_counts": dict(major_six_discussion_counts),
            "major_transfer_impact_counts": dict(major_transfer_impact_counts),
            "group_count_target_40_to_60_met": 40 <= len(group_review) <= 60,
        },
        "hashes": {
            "page_pack_sha16": digest(json.dumps(page_pack, ensure_ascii=False, sort_keys=True)),
            "action_pack_sha16": digest(json.dumps(action_pack, ensure_ascii=False, sort_keys=True)),
            "task_pack_sha16": digest(json.dumps(task_pack, ensure_ascii=False, sort_keys=True)),
            "group_review_sha16": digest(json.dumps(group_review, ensure_ascii=False, sort_keys=True)),
            "major_review_sha16": digest(json.dumps(major_review, ensure_ascii=False, sort_keys=True)),
        },
    }
    OUT_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_workbook(summary, page_pack, action_pack, task_pack, group_review, major_review)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
