#!/usr/bin/env python3
import csv
import json
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from build_issue19_round2_updated_preferences import (
    GROUP_BROWSER,
    MAJOR_BROWSER,
    ROOT,
    SOURCE_PDF_SHA256,
    aggregate_major_signals,
    append_sheet,
    count,
    direction_tags,
    float_value,
    has_clinical_pause,
    has_health_or_vet_special,
    has_main_direction,
    has_nursing_or_animal_block,
    history_code,
    major_role,
    merge_signals,
    public_normal,
    public_value,
    read_csv,
    stable_hash,
)
from build_issue19_round4_city_gradient_candidates import (
    CANDIDATE_RANK,
    CANDIDATE_SCORE,
    CITY_POLICY,
    CITY_SUMMARY_FIELDS,
    EQUIVALENT_SCORES,
    GROUP_FIELDS as ROUND4_GROUP_FIELDS,
    HISTORICAL_LINES,
    MAJOR_FIELDS as ROUND4_MAJOR_FIELDS,
    MOE_ATTRIBUTES,
    aggregate_history,
    aggregate_moe,
    decorate as decorate_round4,
    sorted_groups,
)
from export_issue19_expanded_budget_coop_scenario import tuition_value


EXPORTS = ROOT / "data/exports"
MAX_TUITION = 50000

PREFIX = "issue19-round4-50k-coop-city-gradient"
GROUPS = EXPORTS / f"{PREFIX}-candidate-groups.csv"
PRIORITY = EXPORTS / f"{PREFIX}-priority-groups.csv"
MAJORS = EXPORTS / f"{PREFIX}-major-details.csv"
CITY_SUMMARY = EXPORTS / f"{PREFIX}-city-summary.csv"
HISTORY_MISSING = EXPORTS / f"{PREFIX}-history-missing-groups.csv"
HIGH_RUSH_PAUSED = EXPORTS / f"{PREFIX}-high-rush-paused-groups.csv"
FEE_BLOCKED = EXPORTS / f"{PREFIX}-fee-pending-or-over-budget-groups.csv"
WORKBOOK = EXPORTS / f"{PREFIX}.xlsx"
SUMMARY = EXPORTS / f"{PREFIX}-summary.json"

SOURCE_POLICY = "直接来源于稳定底座、教育部学校属性表、三年投档线旁挂表和逐专业费用线索；不引用 Round3/Round4 输出。"
REACHABLE_POLICY = "主表纳入 H1/H2/H3/H4；H0 历史待补、H5 高冲暂缓和费用待核/超过5万均单列索引，不混入主表。"
FAMILY_POLICY = "在 Round4 其他底线不变的基础上，允许公办高收费或中外合作/港澳合作办学线索，年学费上限50000；仍不纳入民办、职业本科、护理/助产、动物医学/兽医/动物科学、临床/口腔/中医、医技/康复。"
FEE_POLICY = "只把逐专业学费解析不超过50000元且触发明确中外/国际合作或高学费线索的专业计入主表；外方费用、出国费用、非人民币币种、费用待核和超过50000元均单列。"
COOP_KEYWORDS = [
    "中外合作",
    "港澳合作",
    "国际本科",
    "中澳",
    "中英",
    "中美",
    "中德",
    "中俄",
    "中法",
    "中白",
]
INTERNATIONAL_CONTEXT_KEYWORDS = [
    "外方",
    "外国",
    "国外",
    "境外",
    "出国",
    "留学",
    "赴",
    "澳大利亚",
    "英国",
    "美国",
    "德国",
    "俄罗斯",
    "白俄罗斯",
    "法国",
    "韩国",
    "新西兰",
    "乌兹别克斯坦",
    "昆士兰",
    "麦考瑞",
]
COOP_PROGRAM_KEYWORDS = [
    "合作办学",
    "联合培养",
    "学分互认",
    "双学位",
    "国际本科学术互认",
    "中俄学院",
]
EXTRA_COST_PENDING_KEYWORDS = [
    "国外费用另计",
    "英国费用另计",
    "澳方学费",
    "外方学费",
    "外方授权方缴纳",
    "留学期间",
    "赴澳学习",
    "赴国外",
    "第三、四年需赴",
]
NON_YUAN_CURRENCY_KEYWORDS = [
    "学费币种",
    "林吉特",
    "美元",
    "澳元",
    "英镑",
]

FEE_GROUP_FIELDS = [
    "第四轮5万专项候选ID",
    "费用场景",
    "年学费上限",
    "5万内中外合作或高收费专业数",
    "超过5万专业数",
    "费用待核专业数",
    "场景内最低学费",
    "场景内最高学费",
    "5万内专业摘要",
    "费用风险提示",
] + [field for field in ROUND4_GROUP_FIELDS if field != "第四轮候选ID"]

FEE_MAJOR_FIELDS = [
    "第四轮5万专项候选ID",
    "费用场景",
    "费用判断",
    "解析学费",
    "中外合作或高收费线索",
] + [field for field in ROUND4_MAJOR_FIELDS if field != "第四轮候选ID"]


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: public_value(row.get(field, "")) for field in fields} for row in rows])


def allowed_school_attribute(row):
    if public_normal(row):
        return True
    action = row.get("家庭底线属性动作", "")
    if "民办线索" in action or "职业本科" in action:
        return False
    return action in {
        "单独讨论-中外合作或港澳合作办学线索",
    }


def fee_scenario_judgement(row, max_tuition):
    tuition = tuition_value(row.get("学费OCR候选", ""))
    text = row.get("专业名称及备注短摘", "")
    explicit_coop = any(keyword in text for keyword in COOP_KEYWORDS)
    international_program = any(keyword in text for keyword in COOP_PROGRAM_KEYWORDS) and any(
        keyword in text for keyword in INTERNATIONAL_CONTEXT_KEYWORDS
    )
    coop = explicit_coop or international_program
    high_tuition = tuition is not None and tuition > 15000
    if not coop and not high_tuition:
        return "", "", "", tuition
    if any(keyword in text for keyword in EXTRA_COST_PENDING_KEYWORDS + NON_YUAN_CURRENCY_KEYWORDS):
        return "费用待核", "待核费用", "存在外方费用、出国费用或非人民币币种线索，需回原页和章程确认总费用", tuition
    if tuition is None:
        return "费用待核", "待核费用", "需要回原页和章程确认学费", tuition
    if tuition > max_tuition:
        return "超过预算", "超过预算", f"学费候选 {tuition} 元超过 {max_tuition} 元", tuition
    if coop:
        return "预算内中外合作/高收费可讨论", "预算内", f"学费候选 {tuition} 元，不超过 {max_tuition} 元；需核培养模式和证书", tuition
    return "预算内高学费可讨论", "预算内", f"学费候选 {tuition} 元，不超过 {max_tuition} 元；需核是否普通收费或特殊项目", tuition


def base_ok(row):
    return (
        allowed_school_attribute(row)
        and count(row, "专业明细行数") > 0
        and has_main_direction(row)
        and not has_clinical_pause(row)
        and not has_health_or_vet_special(row)
        and not has_nursing_or_animal_block(row)
    )


def build_fee_maps(major_rows):
    by_group = defaultdict(list)
    by_major = {}
    for row in major_rows:
        judgement, fee_status, reason, tuition = fee_scenario_judgement(row, MAX_TUITION)
        if not judgement:
            by_major[row.get("专业行ID", "")] = {
                "费用场景": "普通/未触发高费合作线索",
                "费用判断": "",
                "解析学费": tuition_value(row.get("学费OCR候选", "")) or "",
                "中外合作或高收费线索": "",
                "_tuition_value": tuition_value(row.get("学费OCR候选", "")),
            }
            continue
        out = {
            "费用场景": judgement,
            "费用判断": fee_status,
            "解析学费": tuition or "",
            "中外合作或高收费线索": reason,
            "_tuition_value": tuition,
            "_major_summary": (
                f"{row.get('专业代号OCR', '')} {row.get('专业名称及备注短摘', '')}"
                f"｜学费:{row.get('学费OCR候选', '')}"
                f"｜{judgement}"
            ),
        }
        by_major[row.get("专业行ID", "")] = out
        by_group[row.get("专业组出现ID", "")].append(out)

    group_fee = {}
    for group_id, rows in by_group.items():
        budget_ok = [row for row in rows if str(row.get("费用场景", "")).startswith("预算内")]
        over = [row for row in rows if row.get("费用场景") == "超过预算"]
        pending = [row for row in rows if row.get("费用场景") == "费用待核"]
        tuition_values = [row["_tuition_value"] for row in rows if row.get("_tuition_value") is not None]
        budget_values = [row["_tuition_value"] for row in budget_ok if row.get("_tuition_value") is not None]
        group_fee[group_id] = {
            "5万内中外合作或高收费专业数": len(budget_ok),
            "超过5万专业数": len(over),
            "费用待核专业数": len(pending),
            "场景内最低学费": min(budget_values) if budget_values else "",
            "场景内最高学费": max(budget_values) if budget_values else "",
            "5万内专业摘要": "；".join(row.get("_major_summary", "") for row in budget_ok[:12]),
            "费用风险提示": fee_note(budget_ok, over, pending),
            "_has_scenario": True,
            "_budget_ok": len(budget_ok) > 0,
            "_blocked_by_fee": len(over) > 0 or len(pending) > 0,
            "_scenario_rows": rows,
            "_all_tuition_values": tuition_values,
        }
    return group_fee, by_major


def fee_note(budget_ok, over, pending):
    notes = []
    if budget_ok:
        notes.append(f"{len(budget_ok)}个5万内高收费/中外合作线索")
    if pending:
        notes.append(f"{len(pending)}个费用待核，暂不进主表")
    if over:
        notes.append(f"{len(over)}个超过5万，暂不进主表")
    if not notes:
        notes.append("未触发5万内高收费/中外合作线索")
    notes.append("仍需核第19期原页、湖北官方侧、招生章程、培养模式、证书和出国要求")
    return "；".join(notes)


def group_score(row, fee):
    hist = {"H1": 42, "H2": 46, "H3": 38, "H4": 28, "H0": 12, "H5": -15}.get(history_code(row), 0)
    value = hist
    value += int(fee.get("5万内中外合作或高收费专业数") or 0) * 16
    value += count(row, "数字媒体技术专业数") * 12
    value += count(row, "计算机AI软件专业数") * 5
    value += count(row, "计算机类相关专业数") * 4
    value += count(row, "电子信息网络专业数") * 4
    value += count(row, "机械自动化机器人专业数") * 4
    value += count(row, "师范类相关专业数") * 4
    if row.get("调剂初判", "") == "可进入人工调剂接受度判断":
        value += 8
    if "结构异常" in row.get("调剂初判", ""):
        value -= 12
    if count(row, "特殊限制待核专业数") > 0:
        value -= 8
    value -= max(float_value(row, "历史最高等位分差", 0), 0) / 4
    return round(value, 3)


def decorate(row, moe_map, history_map, fee_map, discussion_set):
    group_id = row.get("专业组出现ID", "")
    fee = fee_map.get(group_id, {})
    out = decorate_round4(row, moe_map, history_map, discussion_set)
    out.pop("第四轮候选ID", None)
    out["第四轮5万专项候选ID"] = f"R4FEEGROUP-{stable_hash('issue19-round4-50k-coop', row)}"
    out["第四轮来源口径"] = SOURCE_POLICY
    out["费用场景"] = "5万内中外合作/高收费专项"
    out["年学费上限"] = str(MAX_TUITION)
    out["5万内中外合作或高收费专业数"] = str(fee.get("5万内中外合作或高收费专业数", 0))
    out["超过5万专业数"] = str(fee.get("超过5万专业数", 0))
    out["费用待核专业数"] = str(fee.get("费用待核专业数", 0))
    out["场景内最低学费"] = str(fee.get("场景内最低学费", ""))
    out["场景内最高学费"] = str(fee.get("场景内最高学费", ""))
    out["5万内专业摘要"] = fee.get("5万内专业摘要", "")
    out["费用风险提示"] = fee.get("费用风险提示", "")
    out["完整专业组接受度初判"] = out.get("完整专业组接受度初判", "") + "；费用专项需确认同组所有可调剂专业均不突破家庭费用底线。"
    out["下一步核验动作"] = "核第19期原页、湖北官方系统/省招办计划、招生章程、学费、培养模式、证书、出国要求、校区、语种/单科/体检限制和调剂范围。"
    out["讨论排序分"] = str(group_score(row, fee))
    return out


def priority_groups(rows):
    selected = []
    selected_ids = set()
    quotas = [
        ({"H1", "H2"}, 50),
        ({"H3"}, 35),
        ({"H4"}, 35),
    ]
    by_score = sorted(rows, key=lambda row: (-float_value(row, "讨论排序分", 0), row.get("院校代码", "")))
    for codes, limit in quotas:
        count_taken = 0
        for row in by_score:
            if row.get("专业组出现ID") in selected_ids or history_code(row) not in codes:
                continue
            out = dict(row)
            out["是否进入优先讨论"] = "true"
            selected.append(out)
            selected_ids.add(row.get("专业组出现ID"))
            count_taken += 1
            if count_taken >= limit:
                break
    return sorted_groups(selected)


def build_major_rows(group_rows, major_rows, fee_by_major):
    group_by_id = {row.get("专业组出现ID", ""): row for row in group_rows}
    rows = []
    for row in major_rows:
        group = group_by_id.get(row.get("专业组出现ID", ""))
        if not group:
            continue
        fee = fee_by_major.get(row.get("专业行ID", ""), {})
        out = {
            "第四轮5万专项候选ID": group.get("第四轮5万专项候选ID", ""),
            "费用场景": fee.get("费用场景", "普通/未触发高费合作线索"),
            "费用判断": fee.get("费用判断", ""),
            "解析学费": fee.get("解析学费", ""),
            "中外合作或高收费线索": fee.get("中外合作或高收费线索", ""),
            "讨论集合": group.get("讨论集合", ""),
            "城市": group.get("城市", ""),
            "冲稳保": group.get("冲稳保", ""),
            "院校代码": row.get("院校代码", ""),
            "院校名称": row.get("院校名称OCR", ""),
            "院校专业组代码": row.get("院校专业组代码OCR规范化", ""),
            "专业组内专业序号": row.get("专业组内专业序号", ""),
            "专业代号": row.get("专业代号OCR", ""),
            "专业名称及备注": row.get("专业名称及备注短摘", ""),
            "专业偏好方向": row.get("专业偏好方向", ""),
            "第四轮方向标签": "；".join(direction_tags(row)),
            "专业角色判断": major_role(row),
            "机器专业接受度初判": row.get("机器专业接受度初判", ""),
            "专业风险类型": row.get("专业风险类型", ""),
            "再选科目OCR候选": row.get("再选科目OCR候选", ""),
            "专业计划数OCR候选": row.get("专业计划数OCR候选", ""),
            "学费OCR候选": row.get("学费OCR候选", ""),
            "字段缺口数": row.get("字段缺口数", ""),
            "字段缺口字段": row.get("字段缺口字段", ""),
            "人工核验优先级": row.get("人工核验优先级", ""),
            "人工核验强度": row.get("人工核验强度", ""),
            "是否必须100%人工核验": row.get("是否必须100%人工核验", ""),
            "PDF原页核验状态": row.get("PDF原页核验状态", ""),
            "湖北官方系统核验状态": row.get("湖北官方系统核验状态", ""),
            "高校官网证据匹配状态": row.get("高校官网证据匹配状态", ""),
            "历史线索分层": row.get("历史线索分层", ""),
            "历史最高等位分差": row.get("历史最高等位分差", ""),
            "历史最低等位分差": row.get("历史最低等位分差", ""),
            "来源页码": row.get("来源页码", ""),
            "版面列": row.get("版面列", ""),
            "专业行ID": row.get("专业行ID", ""),
            "专业组出现ID": row.get("专业组出现ID", ""),
        }
        rows.append(out)
    return rows


def city_summary(rows, priority_rows):
    priority_ids = {row.get("专业组出现ID") for row in priority_rows}
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row.get("城市", "未识别"), row.get("省份", ""))].append(row)
    summary = []
    for (city, province), group_rows in sorted(grouped.items()):
        counter = Counter(row.get("冲稳保", "") for row in group_rows)
        summary.append(
            {
                "城市": city,
                "省份": province,
                "主表专业组数": len(group_rows),
                "优先讨论专业组数": sum(row.get("专业组出现ID") in priority_ids for row in group_rows),
                "保底观察": counter.get("保底观察", 0),
                "稳妥观察": counter.get("稳妥观察", 0),
                "稳冲观察": counter.get("稳冲观察", 0),
                "冲刺观察": counter.get("冲刺观察", 0),
                "历史待补": counter.get("历史待补", 0),
            }
        )
    return summary


def write_workbook(candidate_rows, priority_rows, major_rows, city_rows, history_missing_rows, high_rush_rows, fee_blocked_rows):
    wb = Workbook()
    wb.remove(wb.active)
    guide = wb.create_sheet("00_说明")
    guide.append(["项目", "说明"])
    guide_rows = [
        ("定位", "Round4 5万内中外合作/高收费专项平行版，用于家庭讨论和后续核验，不作为定稿依据。"),
        ("来源", SOURCE_POLICY),
        ("成绩口径", f"考生 {CANDIDATE_SCORE} 分、累计位次 {CANDIDATE_RANK}；等位分按 2025={EQUIVALENT_SCORES['2025']}、2024={EQUIVALENT_SCORES['2024']}、2023={EQUIVALENT_SCORES['2023']} 比较历史投档分。"),
        ("主表边界", REACHABLE_POLICY),
        ("家庭边界", FAMILY_POLICY),
        ("费用边界", FEE_POLICY),
        ("城市口径", CITY_POLICY),
        ("下一步", "先看5万内专业摘要和完整专业组，再核学费、证书、培养模式和调剂范围；H0/H5/费用待核均不当作主表候选。"),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 120
    append_sheet(wb.create_sheet("01_按城市冲稳保总表"), candidate_rows, FEE_GROUP_FIELDS)
    append_sheet(wb.create_sheet("02_优先讨论组"), priority_rows, FEE_GROUP_FIELDS)
    append_sheet(wb.create_sheet("03_城市分布"), city_rows, CITY_SUMMARY_FIELDS)
    append_sheet(wb.create_sheet("04_组内专业明细"), major_rows, FEE_MAJOR_FIELDS)
    append_sheet(wb.create_sheet("05_历史待补组"), history_missing_rows, FEE_GROUP_FIELDS)
    append_sheet(wb.create_sheet("06_高冲暂缓索引"), high_rush_rows, FEE_GROUP_FIELDS)
    append_sheet(wb.create_sheet("07_费用待核或超5万"), fee_blocked_rows, FEE_GROUP_FIELDS)
    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK)


def public_text_safe(paths):
    forbidden = [
        "/Users/",
        "/home/",
        "/var/folders/",
        "/private/",
        "private/",
        "private\\",
        "file://",
        "ocr-runs",
        "rendered-pages",
        "身份证",
        "准考证",
        "报名号",
        "序列号",
        "手机号",
        "Authorization",
        "Bearer ",
        "Cookie",
        "Set-Cookie",
        "access_token",
        "refresh_token",
        "password",
        "secret",
        "api_key",
        "已确认",
        "已核准",
        "最终推荐",
        "最终方案",
        "可填报",
        "可排序",
    ]
    text = "\n".join(path.read_text(encoding="utf-8", errors="ignore") for path in paths)
    hit = [token for token in forbidden if token in text]
    if hit:
        raise SystemExit(f"Round4 5万专项公开文件包含禁止内容：{hit}")


def main():
    group_rows = read_csv(GROUP_BROWSER)
    major_rows = read_csv(MAJOR_BROWSER)
    major_signals, _ = aggregate_major_signals(major_rows)
    enriched_groups = merge_signals(group_rows, major_signals)
    fee_map, fee_by_major = build_fee_maps(major_rows)
    moe_map = aggregate_moe(read_csv(MOE_ATTRIBUTES))
    history_map = aggregate_history(read_csv(HISTORICAL_LINES))

    base_pool = [row for row in enriched_groups if base_ok(row)]
    budget_ready = [
        row
        for row in base_pool
        if fee_map.get(row.get("专业组出现ID", ""), {}).get("_budget_ok")
        and not fee_map.get(row.get("专业组出现ID", ""), {}).get("_blocked_by_fee")
    ]
    fee_blocked_source = [
        row
        for row in base_pool
        if fee_map.get(row.get("专业组出现ID", ""), {}).get("_has_scenario")
        and not (
            fee_map.get(row.get("专业组出现ID", ""), {}).get("_budget_ok")
            and not fee_map.get(row.get("专业组出现ID", ""), {}).get("_blocked_by_fee")
        )
    ]

    candidate_source = [row for row in budget_ready if history_code(row) in {"H1", "H2", "H3", "H4"}]
    history_missing_source = [row for row in budget_ready if history_code(row) == "H0"]
    high_rush_source = [row for row in budget_ready if history_code(row) == "H5"]

    candidate_rows = sorted_groups(
        decorate(row, moe_map, history_map, fee_map, "5万内高收费/中外合作主表")
        for row in candidate_source
    )
    priority_rows = priority_groups(candidate_rows)
    priority_ids = {row.get("专业组出现ID") for row in priority_rows}
    candidate_rows = [
        {**row, "是否进入优先讨论": "true" if row.get("专业组出现ID") in priority_ids else "false"}
        for row in candidate_rows
    ]
    history_missing_rows = sorted_groups(
        decorate(row, moe_map, history_map, fee_map, "历史待补附录") for row in history_missing_source
    )
    high_rush_rows = sorted_groups(
        decorate(row, moe_map, history_map, fee_map, "高冲暂缓索引") for row in high_rush_source
    )
    fee_blocked_rows = sorted_groups(
        decorate(row, moe_map, history_map, fee_map, "费用待核或超过5万附录") for row in fee_blocked_source
    )
    city_rows = city_summary(candidate_rows, priority_rows)
    major_detail_rows = build_major_rows(candidate_rows, major_rows, fee_by_major)

    write_csv(GROUPS, candidate_rows, FEE_GROUP_FIELDS)
    write_csv(PRIORITY, priority_rows, FEE_GROUP_FIELDS)
    write_csv(MAJORS, major_detail_rows, FEE_MAJOR_FIELDS)
    write_csv(CITY_SUMMARY, city_rows, CITY_SUMMARY_FIELDS)
    write_csv(HISTORY_MISSING, history_missing_rows, FEE_GROUP_FIELDS)
    write_csv(HIGH_RUSH_PAUSED, high_rush_rows, FEE_GROUP_FIELDS)
    write_csv(FEE_BLOCKED, fee_blocked_rows, FEE_GROUP_FIELDS)
    write_workbook(candidate_rows, priority_rows, major_detail_rows, city_rows, history_missing_rows, high_rush_rows, fee_blocked_rows)

    summary = {
        "status": "issue19_round4_50k_coop_city_gradient_ready",
        "generated_by": "build_issue19_round4_50k_coop_city_gradient_candidates.py",
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "source_policy": SOURCE_POLICY,
        "city_policy": CITY_POLICY,
        "reachable_policy": REACHABLE_POLICY,
        "family_policy": FAMILY_POLICY,
        "fee_policy": FEE_POLICY,
        "annual_upper_limit_yuan": MAX_TUITION,
        "candidate_score": CANDIDATE_SCORE,
        "candidate_rank": CANDIDATE_RANK,
        "equivalent_scores": EQUIVALENT_SCORES,
        "base_pool_rows_before_fee_filter": len(base_pool),
        "budget_ready_group_rows": len(budget_ready),
        "candidate_group_rows": len(candidate_rows),
        "priority_group_rows": len(priority_rows),
        "major_detail_rows": len(major_detail_rows),
        "city_summary_rows": len(city_rows),
        "history_missing_group_rows": len(history_missing_rows),
        "high_rush_paused_group_rows": len(high_rush_rows),
        "fee_blocked_group_rows": len(fee_blocked_rows),
        "gradient_distribution": dict(Counter(row.get("冲稳保", "") for row in candidate_rows)),
        "priority_gradient_distribution": dict(Counter(row.get("冲稳保", "") for row in priority_rows)),
        "top_city_counts": Counter(row.get("城市", "") or "未识别" for row in candidate_rows).most_common(30),
        "outputs": {
            "workbook": str(WORKBOOK.relative_to(ROOT)),
            "groups_csv": str(GROUPS.relative_to(ROOT)),
            "priority_csv": str(PRIORITY.relative_to(ROOT)),
            "majors_csv": str(MAJORS.relative_to(ROOT)),
            "city_summary_csv": str(CITY_SUMMARY.relative_to(ROOT)),
            "history_missing_csv": str(HISTORY_MISSING.relative_to(ROOT)),
            "high_rush_paused_csv": str(HIGH_RUSH_PAUSED.relative_to(ROOT)),
            "fee_blocked_csv": str(FEE_BLOCKED.relative_to(ROOT)),
        },
        "hard_filters_main": [
            "非民办/非职业本科；允许中外合作或港澳合作办学线索单独讨论",
            "至少一个逐专业费用线索为5万内中外合作/高收费/高学费",
            "同组高收费/合作线索中不得有费用待核或超过5万",
            "必须有逐专业明细",
            "必须命中至少一个当前可讨论方向",
            "临床/口腔/中医等暂缓方向不进主表",
            "护理/助产、动物医学/兽医/动物科学不进主表",
            "医技/康复不进主表",
            "H0 历史待补不进主表",
            "H5 高冲暂缓不进主表",
        ],
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text_safe([GROUPS, PRIORITY, MAJORS, CITY_SUMMARY, HISTORY_MISSING, HIGH_RUSH_PAUSED, FEE_BLOCKED, SUMMARY])
    print(f"写出 Round4 5万专项城市冲稳保总表：{GROUPS.relative_to(ROOT)}")
    print(f"写出 Round4 5万专项优先讨论表：{PRIORITY.relative_to(ROOT)}")
    print(f"写出 Round4 5万专项工作簿：{WORKBOOK.relative_to(ROOT)}")
    print(
        f"基础池={len(base_pool)}；5万费用就绪={len(budget_ready)}；主表={len(candidate_rows)}；"
        f"优先讨论={len(priority_rows)}；专业明细={len(major_detail_rows)}；费用阻断={len(fee_blocked_rows)}"
    )


if __name__ == "__main__":
    main()
