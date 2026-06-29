#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "data/exports"

GROUPS = EXPORTS / "issue19-round4-family-explanation-focus55.csv"
MAJORS = EXPORTS / "issue19-next-closure-family-review-v1-priority55-major-review.csv"

GROUP_OUTPUT = EXPORTS / "issue19-priority55-family-major-decision-group-summary.csv"
MAJOR_OUTPUT = EXPORTS / "issue19-priority55-family-major-decision-major-items.csv"
SUMMARY_OUTPUT = EXPORTS / "issue19-priority55-family-major-decision-workbook-summary.json"
WORKBOOK_OUTPUT = EXPORTS / "issue19-priority55-family-major-decision-workbook.xlsx"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"
DATA_STAGE = "issue19_priority55_family_major_decision_workbook"
GENERATED_AT = "2026-06-29"

FALSE_FIELDS = [
    "最终可用",
    "可进入下一阶段",
    "可否进入最终志愿方案",
    "是否允许作为志愿推荐依据",
    "是否允许自动写回主表",
    "是否允许官网证据替代湖北官方计划",
    "是否允许生成学校专业建议",
    "是否允许写回字段事实",
]

BASE_FIELDS = [
    "来源期号",
    "来源PDF_SHA256",
    "生成日期",
    "数据阶段",
    "主表粒度",
    "任务粒度",
    *FALSE_FIELDS,
]

GROUP_FIELDS = [
    "家庭专业决策组ID",
    *BASE_FIELDS,
    "家庭讨论排序",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组号",
    "院校专业组",
    "城市",
    "冲稳保",
    "历史稳定性等级",
    "与515关系",
    "同代码命中年份数",
    "完整组内专业数",
    "机器可接受专业数",
    "机器勉强接受专业数",
    "机器待核后判断专业数",
    "机器不能接受提示专业数",
    "建议放入6专业讨论数",
    "特殊限制待核专业数",
    "体检色觉风险专业数",
    "语种或单科限制专业数",
    "必须100%人工核验专业数",
    "高校官网证据已匹配专业数",
    "高校官网证据未匹配专业数",
    "字段缺口专业数",
    "调剂接受风险等级",
    "调剂风险说明",
    "家庭本轮要回答的问题",
    "家庭整组态度待填",
    "家庭整组态度可填值",
    "是否愿意服从调剂待填",
    "服从调剂可填值",
    "家庭整组备注待填",
    "核验最小动作",
    "公开填写边界",
    "关联专业集合SHA16",
    "公开安全策略",
]

MAJOR_FIELDS = [
    "家庭专业决策明细ID",
    *BASE_FIELDS,
    "家庭讨论排序",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "城市",
    "冲稳保",
    "专业组内专业序号",
    "专业代号",
    "专业名称及备注",
    "机器家庭接受度建议",
    "家庭最终接受度待填",
    "家庭最终接受度可填值",
    "是否建议放入6专业讨论",
    "建议6专业理由",
    "调剂风险提示",
    "服从调剂影响",
    "组内调剂判断状态",
    "家庭要回答的问题",
    "专业偏好方向",
    "专业角色判断",
    "专业风险类型",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "来源页码",
    "版面列",
    "专业行ID",
    "第四轮候选ID",
    "家庭备注待填",
    "公开填写边界",
    "公开安全策略",
]

FORBIDDEN_PUBLIC_TOKENS = [
    "/Users/",
    "/home/",
    "/var/folders/",
    "/private/",
    "private/",
    "private\\",
    "ocr-runs",
    "rendered-pages",
    "file://",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".tif",
    ".tiff",
    ".heic",
    "Authorization",
    "Bearer ",
    "Cookie",
    "Set-Cookie",
    "access_token",
    "refresh_token",
    "password",
    "secret",
    "api_key",
    "身份证",
    "准考证",
    "报名号",
    "序列号",
    "手机号",
    "OCR行文本",
    "字段确认值",
    "人工读数",
    "PDF原页人工读数",
    "湖北官方字段值",
    "高校官网或招生章程字段值",
    "已确认",
    "已核准",
    "最终推荐",
    "最终方案",
    "可填报",
    "可排序",
]


def clean(value):
    return "" if value is None else str(value).replace("\r", " ").replace("\n", " ").strip()


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in fields} for row in rows])


def stable_id(prefix, parts):
    text = "|".join(clean(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(text.encode('utf-8')).hexdigest()[:16]}"


def sha16(values):
    normalized = "\n".join(sorted({clean(value) for value in values if clean(value)}))
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:16]


def as_int(value):
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def base_row(grain, task_grain):
    row = {
        "来源期号": SOURCE_ISSUE,
        "来源PDF_SHA256": SOURCE_PDF_SHA256,
        "生成日期": GENERATED_AT,
        "数据阶段": DATA_STAGE,
        "主表粒度": grain,
        "任务粒度": task_grain,
    }
    row.update({field: "false" for field in FALSE_FIELDS})
    return row


def yes_count(rows, field):
    return sum(clean(row.get(field)) == "true" for row in rows)


def contains_count(rows, field, token):
    return sum(token in clean(row.get(field)) for row in rows)


def non_empty_count(rows, field):
    return sum(bool(clean(row.get(field))) for row in rows)


def public_policy():
    return "not_final；family_discussion_only；no_field_values；no_private_paths；no_ocr_values；no_recommendation"


def fill_boundary():
    return "本表只填家庭接受度和备注；不要填写计划数、学费、选科读数、截图路径或任何私人信息。"


def group_question(row, majors):
    pending = sum(clean(item.get("机器家庭接受度建议")) == "待核后判断" for item in majors)
    reluctant = sum(clean(item.get("机器家庭接受度建议")) == "勉强接受" for item in majors)
    if pending:
        return "先核限制/字段后，再判断是否愿意服从调剂。"
    if reluctant:
        return "逐条确认勉强接受专业是否真的能接受调剂。"
    return "确认可接受专业是否愿意作为6个专业志愿或调剂底线。"


def workbook_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title)
    ws.append(fields)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([clean(row.get(field, "")) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 28,
        "B": 28,
        "C": 66,
        "D": 18,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    return ws


def assert_public_safe(rows, label):
    text = json.dumps(rows, ensure_ascii=False)
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in text]
    if found:
        raise SystemExit(f"{label} contains forbidden public tokens: {found[:5]}")


def main():
    group_rows = read_csv(GROUPS)
    major_rows = read_csv(MAJORS)
    majors_by_candidate = defaultdict(list)
    for row in major_rows:
        majors_by_candidate[clean(row.get("第四轮候选ID"))].append(row)

    output_groups = []
    output_majors = []
    group_order_by_candidate = {}

    for index, group in enumerate(sorted(group_rows, key=lambda row: as_int(row.get("家庭阅读排序"))), start=1):
        candidate_id = clean(group.get("第四轮候选ID"))
        group_order_by_candidate[candidate_id] = index
        majors = sorted(majors_by_candidate[candidate_id], key=lambda row: as_int(row.get("专业组内专业序号")))
        accept = sum(clean(row.get("机器家庭接受度建议")) == "可接受" for row in majors)
        reluctant = sum(clean(row.get("机器家庭接受度建议")) == "勉强接受" for row in majors)
        pending = sum(clean(row.get("机器家庭接受度建议")) == "待核后判断" for row in majors)
        impossible = sum(clean(row.get("机器家庭接受度建议")) == "不能接受" for row in majors)
        row = base_row("Priority55家庭逐专业决策组汇总", "院校专业组×完整组内专业")
        row.update(
            {
                "家庭专业决策组ID": stable_id("P55-FAM-GROUP", [SOURCE_PDF_SHA256, candidate_id]),
                "家庭讨论排序": index,
                "院校代码": group.get("院校代码", ""),
                "院校名称": group.get("院校名称", ""),
                "院校专业组代码": group.get("院校专业组代码", ""),
                "专业组号": group.get("专业组号", ""),
                "院校专业组": group.get("院校专业组", ""),
                "城市": group.get("城市", ""),
                "冲稳保": group.get("冲稳保", ""),
                "历史稳定性等级": group.get("历史稳定性等级", ""),
                "与515关系": group.get("与515关系", ""),
                "同代码命中年份数": group.get("同代码命中年份数", ""),
                "完整组内专业数": len(majors),
                "机器可接受专业数": accept,
                "机器勉强接受专业数": reluctant,
                "机器待核后判断专业数": pending,
                "机器不能接受提示专业数": impossible,
                "建议放入6专业讨论数": yes_count(majors, "是否建议放入6专业讨论"),
                "特殊限制待核专业数": contains_count(majors, "调剂风险提示", "特殊限制"),
                "体检色觉风险专业数": contains_count(majors, "专业风险类型", "体检"),
                "语种或单科限制专业数": contains_count(majors, "专业风险类型", "语种"),
                "必须100%人工核验专业数": yes_count(majors, "是否必须100%人工核验"),
                "高校官网证据已匹配专业数": sum(clean(row.get("高校官网证据匹配状态")) == "matched" for row in majors),
                "高校官网证据未匹配专业数": sum(clean(row.get("高校官网证据匹配状态")) == "unmatched" for row in majors),
                "字段缺口专业数": non_empty_count(majors, "字段缺口字段"),
                "调剂接受风险等级": group.get("调剂接受风险等级", ""),
                "调剂风险说明": group.get("调剂风险说明", ""),
                "家庭本轮要回答的问题": group_question(group, majors),
                "家庭整组态度待填": "",
                "家庭整组态度可填值": "保留重点核验 / 暂缓 / 排除 / 待家里讨论",
                "是否愿意服从调剂待填": "",
                "服从调剂可填值": "可以 / 勉强可以 / 不可以 / 待核后判断",
                "家庭整组备注待填": "",
                "核验最小动作": group.get("核验最小动作", ""),
                "公开填写边界": fill_boundary(),
                "关联专业集合SHA16": sha16(row.get("专业行ID") for row in majors),
                "公开安全策略": public_policy(),
            }
        )
        output_groups.append(row)

    for source in sorted(
        major_rows,
        key=lambda row: (
            group_order_by_candidate.get(clean(row.get("第四轮候选ID")), 9999),
            as_int(row.get("专业组内专业序号")),
        ),
    ):
        row = base_row("Priority55家庭逐专业决策明细", "院校专业组×专业明细")
        row.update(
            {
                "家庭专业决策明细ID": stable_id(
                    "P55-FAM-MAJOR",
                    [SOURCE_PDF_SHA256, source.get("第四轮候选ID", ""), source.get("专业行ID", "")],
                ),
                "家庭讨论排序": group_order_by_candidate.get(clean(source.get("第四轮候选ID")), ""),
                "院校代码": source.get("院校代码", ""),
                "院校名称": source.get("院校名称", ""),
                "院校专业组代码": source.get("院校专业组代码", ""),
                "城市": source.get("城市", ""),
                "冲稳保": source.get("冲稳保", ""),
                "专业组内专业序号": source.get("专业组内专业序号", ""),
                "专业代号": source.get("专业代号", ""),
                "专业名称及备注": source.get("专业名称及备注", ""),
                "机器家庭接受度建议": source.get("机器家庭接受度建议", ""),
                "家庭最终接受度待填": "",
                "家庭最终接受度可填值": "可接受 / 勉强接受 / 不能接受 / 待了解",
                "是否建议放入6专业讨论": source.get("是否建议放入6专业讨论", ""),
                "建议6专业理由": source.get("建议6专业理由", ""),
                "调剂风险提示": source.get("调剂风险提示", ""),
                "服从调剂影响": source.get("服从调剂影响", ""),
                "组内调剂判断状态": source.get("组内调剂判断状态", ""),
                "家庭要回答的问题": source.get("家庭要回答的问题", ""),
                "专业偏好方向": source.get("专业偏好方向", ""),
                "专业角色判断": source.get("专业角色判断", ""),
                "专业风险类型": source.get("专业风险类型", ""),
                "人工核验优先级": source.get("人工核验优先级", ""),
                "人工核验强度": source.get("人工核验强度", ""),
                "是否必须100%人工核验": source.get("是否必须100%人工核验", ""),
                "PDF原页核验状态": source.get("PDF原页核验状态", ""),
                "湖北官方系统核验状态": source.get("湖北官方系统核验状态", ""),
                "高校官网证据匹配状态": source.get("高校官网证据匹配状态", ""),
                "来源页码": source.get("来源页码", ""),
                "版面列": source.get("版面列", ""),
                "专业行ID": source.get("专业行ID", ""),
                "第四轮候选ID": source.get("第四轮候选ID", ""),
                "家庭备注待填": "",
                "公开填写边界": fill_boundary(),
                "公开安全策略": public_policy(),
            }
        )
        output_majors.append(row)

    assert_public_safe(output_groups, "group_summary")
    assert_public_safe(output_majors, "major_items")

    write_csv(GROUP_OUTPUT, output_groups, GROUP_FIELDS)
    write_csv(MAJOR_OUTPUT, output_majors, MAJOR_FIELDS)

    wb = Workbook()
    wb.remove(wb.active)
    workbook_sheet(wb, "55组汇总", output_groups, GROUP_FIELDS)
    workbook_sheet(wb, "逐专业决策458", output_majors, MAJOR_FIELDS)
    wb.save(WORKBOOK_OUTPUT)

    acceptance_counts = Counter(row["机器家庭接受度建议"] for row in output_majors)
    summary = {
        "status": "issue19_priority55_family_major_decision_workbook_ready_not_final",
        "generated_by": "build_issue19_priority55_family_major_decision_workbook.py",
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "group_output": str(GROUP_OUTPUT.relative_to(ROOT)),
        "major_output": str(MAJOR_OUTPUT.relative_to(ROOT)),
        "workbook_output": str(WORKBOOK_OUTPUT.relative_to(ROOT)),
        "group_count": len(output_groups),
        "major_count": len(output_majors),
        "machine_acceptance_counts": dict(acceptance_counts),
        "suggested_six_major_count": yes_count(output_majors, "是否建议放入6专业讨论"),
        "must_100_percent_manual_review_count": yes_count(output_majors, "是否必须100%人工核验"),
        "risk_type_counts": dict(Counter(row["专业风险类型"] or "无公开风险提示" for row in output_majors)),
        "pdf_pending_count": sum(bool(row["PDF原页核验状态"]) for row in output_majors),
        "hubei_official_pending_count": sum(bool(row["湖北官方系统核验状态"]) for row in output_majors),
        "final_available_count": 0,
        "recommendation_basis_allowed_count": 0,
        "field_writeback_allowed_count": 0,
        "policy": "本工作簿只用于家庭逐专业接受度讨论，不是最终志愿方案；不得填写字段内容或私人信息。",
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    public_text = (
        GROUP_OUTPUT.read_text(encoding="utf-8", errors="ignore")
        + MAJOR_OUTPUT.read_text(encoding="utf-8", errors="ignore")
        + SUMMARY_OUTPUT.read_text(encoding="utf-8", errors="ignore")
    )
    found = [token for token in FORBIDDEN_PUBLIC_TOKENS if token in public_text]
    if found:
        raise SystemExit(f"public output contains forbidden tokens: {found[:5]}")

    print(f"wrote {GROUP_OUTPUT}")
    print(f"wrote {MAJOR_OUTPUT}")
    print(f"wrote {SUMMARY_OUTPUT}")
    print(f"wrote {WORKBOOK_OUTPUT}")


if __name__ == "__main__":
    main()
