#!/usr/bin/env python3
import csv
import hashlib
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING = ROOT / "data/working"
EXPORTS = ROOT / "data/exports"

FIRST_EXEC = WORKING / "issue19-stable-foundation-first-closure-execution-queue.csv"
FIRST_DETAIL = WORKING / "issue19-stable-foundation-first-closure-detail-packet.csv"
FIRST_PAGE_CAND = WORKING / "issue19-stable-foundation-first-closure-page-side-candidate-dashboard.csv"
FIRST_FIELD_LEDGER = WORKING / "issue19-stable-foundation-first-closure-field-confirmation-public-ledger.csv"
SCHOOL_SOURCE = WORKING / "issue19-school-source-opportunity-queue.csv"
LIVE_SOURCE_LEDGER = WORKING / "issue19-school-source-live-20260629-ledger.csv"
ROUND4_PRIORITY = EXPORTS / "issue19-round4-city-gradient-priority120-groups.csv"
ROUND4_MAJORS = EXPORTS / "issue19-round4-city-gradient-major-details.csv"

OUT_PREFIX = "issue19-closure-and-shortlist-v1"
OUT_SUMMARY = EXPORTS / f"{OUT_PREFIX}-summary.json"
OUT_WORKBOOK = EXPORTS / f"{OUT_PREFIX}.xlsx"
OUT_FIRST_PAGE = EXPORTS / f"{OUT_PREFIX}-first-closure-page-sides.csv"
OUT_FIRST_DETAIL = EXPORTS / f"{OUT_PREFIX}-first-closure-detail-tasks.csv"
OUT_SCHOOL_SOURCE = EXPORTS / f"{OUT_PREFIX}-school-source-tasks.csv"
OUT_SELECTED_GROUPS = EXPORTS / f"{OUT_PREFIX}-priority55-groups.csv"
OUT_SELECTED_MAJORS = EXPORTS / f"{OUT_PREFIX}-priority55-major-details.csv"
OUT_PAUSED_GROUPS = EXPORTS / f"{OUT_PREFIX}-paused65-groups.csv"

SOURCE_ISSUE = "湖北招生考试2026年19期·本科普通批（下）"
SOURCE_PDF_SHA256 = "ee61fc69389f24a9a7830167113cf0ddc0447f8fa4b2743cd3241be60a9bd86d"

PREFERRED_CITIES = {"武汉市": 10, "西安市": 8, "成都市": 8, "北京市": 8}
SECONDARY_CITIES = {"天津市": 3, "重庆市": 3, "郑州市": 2, "南京市": 2, "杭州市": 2, "长沙市": 2}
GRADIENT_QUOTA = {
    "保底观察": 1,
    "稳妥观察": 24,
    "稳冲观察": 18,
    "冲刺观察": 12,
}

FIRST_PAGE_FIELDS = [
    "执行顺序",
    "页码版面键",
    "来源页码",
    "版面列",
    "执行泳道",
    "第一闭环页列优先级",
    "页列总任务数",
    "PDF原页必核任务数",
    "湖北官方侧必核任务数",
    "高校辅证需复核任务数",
    "双人复核任务数",
    "专业计划数字段任务数",
    "学费字段任务数",
    "再选科目字段任务数",
    "PDFOCR候选任务数",
    "无PDFOCR需看图任务数",
    "PDFOCR与高校辅证冲突任务数",
    "C0冲突任务数",
    "C1官网补缺任务数",
    "C7官网未匹配任务数",
    "EXEC01冲突异常字段数",
    "EXEC02计划数偏大字段数",
    "EXEC03高校辅证字段数",
    "页列首要核验动作",
    "页列候选核验动作",
    "当前闭环状态",
    "当前阻断原因",
    "下一步",
    "完成条件",
]

FIRST_DETAIL_FIELDS = [
    "执行顺序",
    "页码版面键",
    "来源页码",
    "版面列",
    "任务来源类型",
    "第一闭环纳入原因",
    "院校代码",
    "院校名称OCR",
    "院校专业组代码OCR规范化",
    "专业组内专业序号",
    "专业代号OCR",
    "专业名称及备注短摘",
    "字段名",
    "OCR专业计划数候选",
    "最佳官网计划数",
    "OCR学费候选",
    "最佳官网学费",
    "OCR再选科目候选",
    "最佳官网选科",
    "官网证据强度",
    "官网来源状态",
    "官网证据匹配状态",
    "计划数核验状态",
    "差异字段集合",
    "是否裁图OCR与高校辅证冲突",
    "是否机器高校冲突",
    "人工核验泳道",
    "人工核验方式",
    "是否需要双人复核",
    "PDF原页核页状态",
    "湖北官方系统或省招办计划核验状态",
    "高校官网或招生章程辅证状态",
    "三方字段一致性状态",
    "字段事实写回状态",
    "当前闭环状态",
    "下一步",
    "专业行ID",
    "稳定基座第一闭环明细任务ID",
]

SCHOOL_SOURCE_FIELDS = [
    "执行建议序号",
    "机会优先级",
    "机会类型",
    "院校代码",
    "院校名称OCR",
    "高校侧刷新批次",
    "高校侧刷新任务类型",
    "官网辅证自动动作",
    "官网证据强度",
    "官网来源状态",
    "来源文件类型集合",
    "公开来源文件数量",
    "公开来源URL数量",
    "涉及招生明细数",
    "计划数冲突行数",
    "官网补缺候选行数",
    "强辅证抽检行数",
    "部分来源待结构化行数",
    "继续补源行数",
    "官网未匹配行数",
    "C4C6计划数一致候选数",
    "C4C6官网可补OCR计划数候选数",
    "C4C6计划数冲突候选数",
    "最新自动探针状态",
    "来源质量判断",
    "自动化下一步",
    "人工最小核验动作",
    "保真边界",
    "字段事实写回状态",
    "下一步",
]

SELECTED_GROUP_FIELDS = [
    "重点核验序号",
    "重点核验状态",
    "压缩结论",
    "入选或暂缓理由",
    "城市偏好提示",
    "核验优先级",
    "建议家庭动作",
    "是否可作为定稿依据",
    "组内硬排除专业数",
    "城市",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "冲稳保",
    "历史线索分层",
    "同代码命中年份数",
    "2025投档分",
    "2025等位分差",
    "2024投档分",
    "2024等位分差",
    "2023投档分",
    "2023等位分差",
    "专业明细行数",
    "数字媒体技术专业数",
    "计算机类相关专业数",
    "计算机AI软件专业数",
    "电子信息网络专业数",
    "机械自动化机器人专业数",
    "师范类相关专业数",
    "环境工程科学专业数",
    "农业不含动物相关专业数",
    "工商旅游管理专业数",
    "特殊限制待核专业数",
    "调剂初判",
    "完整专业组接受度初判",
    "组内专业摘录",
    "办学地点线索",
    "来源页码",
    "版面列",
    "机器核验下一步",
    "人工核验下一步",
    "下一步核验动作",
    "专业组出现ID",
    "第四轮候选ID",
]

SELECTED_MAJOR_FIELDS = [
    "重点核验序号",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "城市",
    "冲稳保",
    "专业组内专业序号",
    "专业代号",
    "专业名称及备注",
    "专业偏好方向",
    "专业角色判断",
    "家庭接受度初判V1",
    "调剂风险初判V1",
    "专业风险类型",
    "再选科目OCR候选",
    "专业计划数OCR候选",
    "学费OCR候选",
    "字段缺口数",
    "字段缺口字段",
    "人工核验优先级",
    "人工核验强度",
    "是否必须100%人工核验",
    "PDF原页核验状态",
    "湖北官方系统核验状态",
    "高校官网证据匹配状态",
    "来源页码",
    "版面列",
    "专业行ID",
    "第四轮候选ID",
]

LIVE_SOURCE_FIELDS = [
    "抓取日期",
    "院校代码",
    "院校名称",
    "机会队列类型",
    "自动补源结论",
    "官方来源URL",
    "本地留存文件",
    "结构化输出",
    "湖北物理计划状态",
    "保真边界",
    "下一步",
]


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows([{field: clean(row.get(field, "")) for field in fields} for row in rows])


def write_json(path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def clean(value):
    text = "" if value is None else str(value)
    return text.replace("\r", " ").replace("\n", " ").strip()


def as_int(value):
    try:
        return int(float(clean(value)))
    except ValueError:
        return 0


def as_float(value):
    try:
        return float(clean(value))
    except ValueError:
        return 0.0


def stable_id(*parts):
    return hashlib.sha1("|".join(clean(p) for p in parts).encode("utf-8")).hexdigest()[:16]


def page_key(row):
    page = clean(row.get("来源页码"))
    side = clean(row.get("版面列"))
    return f"{int(page):03d}-{side}" if page.isdigit() else f"{page}-{side}"


def index_by(rows, field):
    return {clean(row.get(field)): row for row in rows}


def city_bonus(city):
    city = clean(city)
    if city in PREFERRED_CITIES:
        return PREFERRED_CITIES[city]
    return SECONDARY_CITIES.get(city, 0)


def direction_score(row):
    fields = [
        ("数字媒体技术专业数", 8),
        ("计算机AI软件专业数", 5),
        ("计算机类相关专业数", 4),
        ("电子信息网络专业数", 3),
        ("师范类相关专业数", 3),
        ("机械自动化机器人专业数", 2),
        ("环境工程科学专业数", 1),
        ("农业不含动物相关专业数", 1),
    ]
    return sum(as_int(row.get(field)) * weight for field, weight in fields)


def risk_penalty(row):
    penalty = 0
    penalty += as_int(row.get("_hard_major_count")) * 25
    penalty += min(as_int(row.get("特殊限制待核专业数")) * 2, 10)
    penalty += min(as_int(row.get("专业明细行数")) / 8, 4)
    if clean(row.get("历史组名疑似不一致")) not in {"", "false", "无"}:
        penalty += 4
    if clean(row.get("再选要求跨年变化")) not in {"", "false", "无"}:
        penalty += 4
    return penalty


def shortlist_score(row):
    return as_float(row.get("讨论排序分")) + city_bonus(row.get("城市")) + direction_score(row) - risk_penalty(row)


def direction_summary(row):
    parts = []
    mapping = [
        ("数字媒体技术", "数字媒体技术专业数"),
        ("计算机/软件/AI", "计算机AI软件专业数"),
        ("计算机类", "计算机类相关专业数"),
        ("电子信息/网络", "电子信息网络专业数"),
        ("师范", "师范类相关专业数"),
        ("机械自动化", "机械自动化机器人专业数"),
        ("环境", "环境工程科学专业数"),
        ("农业不含动物", "农业不含动物相关专业数"),
        ("工商旅游管理", "工商旅游管理专业数"),
    ]
    for label, field in mapping:
        n = as_int(row.get(field))
        if n:
            parts.append(f"{label}{n}")
    return "；".join(parts) or "无明显当前偏好专业，主要作为梯度/调剂讨论"


def selected_reason(row):
    reasons = [
        clean(row.get("冲稳保")),
        clean(row.get("历史线索分层")),
        direction_summary(row),
    ]
    city = clean(row.get("城市"))
    if city in PREFERRED_CITIES:
        reasons.append(f"命中初始优先城市：{city}")
    elif city in SECONDARY_CITIES:
        reasons.append(f"城市可作为次级讨论：{city}")
    if as_int(row.get("特殊限制待核专业数")):
        reasons.append(f"有{as_int(row.get('特殊限制待核专业数'))}个特殊限制专业，需先核章程/原页")
    if as_int(row.get("_hard_major_count")):
        reasons.append(f"组内硬排除专业{as_int(row.get('_hard_major_count'))}条，原则上后续应降级")
    reasons.append("公办普通学费主线且未进入H0/H5附录")
    return "；".join(r for r in reasons if r)


def paused_reason(row, selected_by_gradient):
    reasons = []
    gradient = clean(row.get("冲稳保"))
    if as_int(row.get("_hard_major_count")):
        reasons.append(f"组内存在{as_int(row.get('_hard_major_count'))}条当前硬排除或医学护理相关风险，服从调剂前先暂缓")
    if selected_by_gradient.get(gradient, 0) >= GRADIENT_QUOTA.get(gradient, 0):
        reasons.append(f"{gradient}本轮名额已优先用于分数/方向/城市综合分更高的组")
    if city_bonus(row.get("城市")) == 0:
        reasons.append("城市未命中当前优先城市，只保留为备选")
    if direction_score(row) == 0:
        reasons.append("当前专业方向匹配较弱")
    if as_int(row.get("特殊限制待核专业数")):
        reasons.append("存在特殊限制待核，先不占用第一核验位")
    if not reasons:
        reasons.append("仍在Round4优先120组内，但本轮先让位给更需要优先核验的梯度组合")
    return "；".join(reasons)


def city_note(city):
    city = clean(city)
    if city in PREFERRED_CITIES:
        return f"命中已提到的优先城市：{city}"
    if city in SECONDARY_CITIES:
        return f"次级可讨论城市：{city}"
    return "非当前优先城市；只作为录取梯度、学校和专业匹配讨论"


def select_priority_groups(rows):
    by_gradient = defaultdict(list)
    for row in rows:
        row["_shortlist_score"] = shortlist_score(row)
        by_gradient[clean(row.get("冲稳保"))].append(row)
    selected = []
    selected_by_gradient = Counter()
    for gradient, quota in GRADIENT_QUOTA.items():
        candidates = [
            row for row in by_gradient.get(gradient, [])
            if as_int(row.get("_hard_major_count")) == 0
        ]
        if len(candidates) < quota:
            candidates = by_gradient.get(gradient, [])
        candidates = sorted(candidates, key=lambda r: (-r["_shortlist_score"], clean(r.get("院校代码")), clean(r.get("院校专业组代码"))))
        for row in candidates[:quota]:
            selected.append(row)
            selected_by_gradient[gradient] += 1
    if len(selected) < 55:
        selected_ids = {clean(row.get("第四轮候选ID")) for row in selected}
        rest = [row for row in rows if clean(row.get("第四轮候选ID")) not in selected_ids]
        rest.sort(key=lambda r: (-r["_shortlist_score"], clean(r.get("冲稳保")), clean(r.get("院校代码"))))
        selected.extend(rest[: 55 - len(selected)])
        for row in rest[: 55 - len(selected)]:
            selected_by_gradient[clean(row.get("冲稳保"))] += 1
    selected.sort(key=lambda r: (["保底观察", "稳妥观察", "稳冲观察", "冲刺观察"].index(clean(r.get("冲稳保"))) if clean(r.get("冲稳保")) in GRADIENT_QUOTA else 9, -r["_shortlist_score"]))
    selected_ids = {clean(row.get("第四轮候选ID")) for row in selected}
    paused = [row for row in rows if clean(row.get("第四轮候选ID")) not in selected_ids]
    paused.sort(key=lambda r: (-r["_shortlist_score"], clean(r.get("冲稳保"))))
    return selected, paused, selected_by_gradient


def group_output_row(row, index, status, reason):
    return {
        "重点核验序号": str(index) if status == "重点核验" else "",
        "重点核验状态": status,
        "压缩结论": "本轮进入55组重点核验" if status == "重点核验" else "暂缓保留在Round4优先120组备选",
        "入选或暂缓理由": reason,
        "城市偏好提示": city_note(row.get("城市")),
        "核验优先级": priority_label(row),
        "建议家庭动作": family_action(row, status),
        **{field: clean(row.get(field)) for field in SELECTED_GROUP_FIELDS if field not in {
            "重点核验序号",
            "重点核验状态",
            "压缩结论",
            "入选或暂缓理由",
            "城市偏好提示",
            "核验优先级",
            "建议家庭动作",
        }},
        "组内硬排除专业数": str(as_int(row.get("_hard_major_count"))),
    }


def priority_label(row):
    gradient = clean(row.get("冲稳保"))
    if gradient == "保底观察":
        return "P0-保底样本先核"
    if gradient == "稳妥观察":
        return "P1-稳妥主力先核"
    if gradient == "稳冲观察":
        return "P2-稳冲择优核"
    return "P3-冲刺择优核"


def family_action(row, status):
    if status != "重点核验":
        return "暂不投入人工核验时间；家庭点名喜欢时再拉回"
    return "看完整组内专业，标出可接受/勉强接受/不能接受，再决定是否值得核原页和章程"


def major_acceptance(row):
    machine = clean(row.get("机器专业接受度初判"))
    risk = clean(row.get("专业风险类型"))
    direction = clean(row.get("专业偏好方向"))
    if "默认不能接受" in machine or "医学/护理" in risk:
        return "不能接受-当前硬排除或医学护理相关"
    if "特殊限制" in machine or risk:
        return "待核后判断-有体检/语种/单科/专项等限制"
    if direction:
        return "优先了解-命中当前专业方向"
    return "可作为调剂了解-未命中偏好但暂未见硬排除"


def transfer_risk(row, group_row):
    if "不能接受" in major_acceptance(row):
        return "同组若存在此类专业，服从调剂需谨慎"
    if as_int(group_row.get("特殊限制待核专业数")):
        return "组内有特殊限制待核，服从调剂前需核完整专业组"
    if as_int(group_row.get("专业明细行数")) > 12:
        return "组内专业较多，需逐条确认调剂接受度"
    return "暂未见硬排除项，但仍需原页/章程闭环"


def append_sheet(wb, title, rows, fields, tab_color=None):
    ws = wb.create_sheet(title)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(fields)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([clean(row.get(field)) for field in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, field in enumerate(fields, 1):
        max_len = min(max([len(clean(field))] + [len(clean(row.get(field))) for row in rows[:200]]) + 2, 42)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(10, max_len)
    return ws


def build_first_page_rows(exec_rows, page_cand_by_id):
    out = []
    for row in exec_rows:
        cand = page_cand_by_id.get(clean(row.get("稳定基座第一闭环页列包ID")), {})
        merged = {**row}
        for field in [
            "PDFOCR候选任务数",
            "无PDFOCR需看图任务数",
            "PDFOCR与高校辅证冲突任务数",
            "页列候选核验动作",
        ]:
            merged[field] = cand.get(field, "")
        merged["页码版面键"] = page_key(row)
        merged["当前闭环状态"] = "待核：PDF原页、湖北官方侧、高校辅证和字段写回均未闭环"
        out.append({field: clean(merged.get(field)) for field in FIRST_PAGE_FIELDS})
    return out


def build_detail_rows(detail_rows, exec_by_page):
    out = []
    for row in detail_rows:
        key = page_key(row)
        exec_row = exec_by_page.get(key, {})
        merged = {**row}
        merged["执行顺序"] = exec_row.get("执行顺序", "")
        merged["页码版面键"] = key
        merged["当前闭环状态"] = "线索待核，不可写回字段事实"
        out.append({field: clean(merged.get(field)) for field in FIRST_DETAIL_FIELDS})
    out.sort(key=lambda r: (as_int(r.get("执行顺序")), as_int(r.get("来源页码")), clean(r.get("版面列")), clean(r.get("院校代码")), clean(r.get("专业代号OCR"))))
    return out


def build_school_source_rows(rows):
    out = []
    for row in rows:
        out.append({field: clean(row.get(field)) for field in SCHOOL_SOURCE_FIELDS})
    out.sort(key=lambda r: (as_int(r.get("执行建议序号")), clean(r.get("院校代码")), clean(r.get("高校侧刷新任务类型"))))
    return out


def build_shortlist(priority_rows, major_rows):
    hard_major_count_by_group = Counter()
    for row in major_rows:
        if "不能接受" in major_acceptance(row):
            hard_major_count_by_group[clean(row.get("第四轮候选ID"))] += 1
    for row in priority_rows:
        row["_hard_major_count"] = str(hard_major_count_by_group.get(clean(row.get("第四轮候选ID")), 0))
    selected, paused, selected_by_gradient = select_priority_groups(priority_rows)
    selected_group_rows = [
        group_output_row(row, idx, "重点核验", selected_reason(row))
        for idx, row in enumerate(selected, 1)
    ]
    paused_group_rows = [
        group_output_row(row, "", "暂缓", paused_reason(row, selected_by_gradient))
        for row in paused
    ]
    selected_ids = {clean(row.get("第四轮候选ID")) for row in selected}
    group_by_id = {clean(row.get("第四轮候选ID")): row for row in selected}
    order_by_id = {clean(row.get("第四轮候选ID")): str(idx) for idx, row in enumerate(selected, 1)}
    selected_major_rows = []
    for row in major_rows:
        cid = clean(row.get("第四轮候选ID"))
        if cid not in selected_ids:
            continue
        group_row = group_by_id[cid]
        out = {field: clean(row.get(field)) for field in SELECTED_MAJOR_FIELDS}
        out["重点核验序号"] = order_by_id[cid]
        out["家庭接受度初判V1"] = major_acceptance(row)
        out["调剂风险初判V1"] = transfer_risk(row, group_row)
        selected_major_rows.append(out)
    selected_major_rows.sort(key=lambda r: (as_int(r.get("重点核验序号")), as_int(r.get("专业组内专业序号"))))
    return selected_group_rows, paused_group_rows, selected_major_rows


def main():
    exec_rows = read_csv(FIRST_EXEC)
    detail_rows = read_csv(FIRST_DETAIL)
    page_cand_rows = read_csv(FIRST_PAGE_CAND)
    field_rows = read_csv(FIRST_FIELD_LEDGER)
    school_rows = read_csv(SCHOOL_SOURCE)
    live_source_rows = read_csv(LIVE_SOURCE_LEDGER) if LIVE_SOURCE_LEDGER.exists() else []
    priority_rows = read_csv(ROUND4_PRIORITY)
    major_rows = read_csv(ROUND4_MAJORS)

    page_cand_by_packet = index_by(page_cand_rows, "稳定基座第一闭环页列包ID")
    exec_by_page = {page_key(row): row for row in exec_rows}

    first_page_rows = build_first_page_rows(exec_rows, page_cand_by_packet)
    first_detail_rows = build_detail_rows(detail_rows, exec_by_page)
    school_source_rows = build_school_source_rows(school_rows)
    selected_group_rows, paused_group_rows, selected_major_rows = build_shortlist(priority_rows, major_rows)

    write_csv(OUT_FIRST_PAGE, first_page_rows, FIRST_PAGE_FIELDS)
    write_csv(OUT_FIRST_DETAIL, first_detail_rows, FIRST_DETAIL_FIELDS)
    write_csv(OUT_SCHOOL_SOURCE, school_source_rows, SCHOOL_SOURCE_FIELDS)
    write_csv(OUT_SELECTED_GROUPS, selected_group_rows, SELECTED_GROUP_FIELDS)
    write_csv(OUT_SELECTED_MAJORS, selected_major_rows, SELECTED_MAJOR_FIELDS)
    write_csv(OUT_PAUSED_GROUPS, paused_group_rows, SELECTED_GROUP_FIELDS)

    summary_rows = [
        {
            "项目": "第一闭环页列",
            "当前数量": str(len(first_page_rows)),
            "说明": "37个PDF页码×版面列，覆盖206条最高优先级字段/辅证任务",
        },
        {
            "项目": "第一闭环逐任务",
            "当前数量": str(len(first_detail_rows)),
            "说明": "逐专业招生明细×字段/官网辅证任务；仍全部待PDF原页和湖北官方侧闭环",
        },
        {
            "项目": "高校官网辅证任务",
            "当前数量": str(len(school_source_rows)),
            "说明": "36所学校、80个高校侧辅证动作；只做double check和补源，不替代省招办计划",
        },
        {
            "项目": "Round4重点核验组",
            "当前数量": str(len(selected_group_rows)),
            "说明": "从Round4优先120组压缩到55组，用于下一轮家庭讨论和逐组核验",
        },
        {
            "项目": "重点组逐专业明细",
            "当前数量": str(len(selected_major_rows)),
            "说明": "展开55组的完整组内专业，标记家庭接受度和调剂风险初判",
        },
    ]

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    append_sheet(wb, "00_怎么看", summary_rows, ["项目", "当前数量", "说明"], "70AD47")
    append_sheet(wb, "01_第一闭环页列", first_page_rows, FIRST_PAGE_FIELDS, "5B9BD5")
    append_sheet(wb, "02_第一闭环逐任务", first_detail_rows, FIRST_DETAIL_FIELDS, "5B9BD5")
    append_sheet(wb, "03_高校官网辅证任务", school_source_rows, SCHOOL_SOURCE_FIELDS, "ED7D31")
    append_sheet(wb, "04_重点核验55组", selected_group_rows, SELECTED_GROUP_FIELDS, "70AD47")
    append_sheet(wb, "05_重点组完整专业", selected_major_rows, SELECTED_MAJOR_FIELDS, "70AD47")
    append_sheet(wb, "06_暂缓65组", paused_group_rows, SELECTED_GROUP_FIELDS, "A5A5A5")
    append_sheet(wb, "07_新增官网源尝试", live_source_rows, LIVE_SOURCE_FIELDS, "FFC000")
    wb.save(OUT_WORKBOOK)

    gradient_counts = Counter(row.get("冲稳保") for row in selected_group_rows)
    acceptance_counts = Counter(row.get("家庭接受度初判V1") for row in selected_major_rows)
    first_lane_counts = Counter(row.get("执行泳道") for row in first_page_rows)
    school_priority_counts = Counter(row.get("机会类型") for row in school_source_rows)
    summary = {
        "status": "issue19_closure_and_shortlist_v1_ready_not_final",
        "generated_by": Path(__file__).name,
        "source_issue": SOURCE_ISSUE,
        "source_pdf_sha256": SOURCE_PDF_SHA256,
        "usage_boundary": "只用于字段事实闭环推进、官网辅证double check和Round4重点核验排序；不作为定稿依据。",
        "outputs": {
            "workbook": str(OUT_WORKBOOK.relative_to(ROOT)),
            "first_closure_page_sides": str(OUT_FIRST_PAGE.relative_to(ROOT)),
            "first_closure_detail_tasks": str(OUT_FIRST_DETAIL.relative_to(ROOT)),
            "school_source_tasks": str(OUT_SCHOOL_SOURCE.relative_to(ROOT)),
            "priority55_groups": str(OUT_SELECTED_GROUPS.relative_to(ROOT)),
            "priority55_major_details": str(OUT_SELECTED_MAJORS.relative_to(ROOT)),
            "paused65_groups": str(OUT_PAUSED_GROUPS.relative_to(ROOT)),
        },
        "first_closure": {
            "page_side_count": len(first_page_rows),
            "detail_task_count": len(first_detail_rows),
            "field_confirmation_public_rows": len(field_rows),
            "lane_counts": dict(first_lane_counts),
            "pdf_review_pending_count": len(first_detail_rows),
            "hubei_official_review_pending_count": len(first_detail_rows),
            "field_writeback_ready_count": 0,
        },
        "school_source": {
            "task_count": len(school_source_rows),
            "unique_school_count": len({row.get("院校代码") for row in school_source_rows}),
            "opportunity_type_counts": dict(school_priority_counts),
            "live_source_attempt_rows": len(live_source_rows),
        },
        "round4_shortlist": {
            "source_priority_group_count": len(priority_rows),
            "selected_group_count": len(selected_group_rows),
            "paused_group_count": len(paused_group_rows),
            "selected_major_detail_count": len(selected_major_rows),
            "gradient_counts": dict(gradient_counts),
            "major_acceptance_counts": dict(acceptance_counts),
        },
        "gates": {
            "all_outputs_not_final_basis": True,
            "field_writeback_ready_count": 0,
            "recommendation_basis_allowed_count": 0,
            "final_available_count": 0,
        },
    }
    write_json(OUT_SUMMARY, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
